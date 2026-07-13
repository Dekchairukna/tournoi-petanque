from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
from flask_socketio import SocketIO, emit

from flask_login import LoginManager, current_user, login_user, logout_user, login_required
from datetime import datetime, timedelta, date
from flask import session
from PIL import Image
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from models import db, Match, Team, Event, User
import os
from urllib.parse import quote
from urllib.request import urlopen, Request
import re
import sys
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))
import pandas as pd
from flask_migrate import Migrate
from swiss_logic import generate_pairings, generate_manual_pairings
from standings import calculate_standings
from flask import Blueprint
from dotenv import load_dotenv
from sqlalchemy import func, text
from functools import wraps
from collections import defaultdict
from routes.match import match_bp  # import blueprint ที่สร้างในไฟล์ routes/match.py
from flask_wtf.file import FileField, FileAllowed
import json
import hashlib
import secrets
import qrcode
from i18n import SUPPORTED_LANGS, TEXT_TRANSLATIONS, translate
# สำหรับ Excel
import io
import math
import random
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from itsdangerous import URLSafeSerializer, BadSignature

load_dotenv()
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "your_secret_key")
BASE_DIR = os.path.abspath(os.path.dirname(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
os.makedirs(INSTANCE_DIR, exist_ok=True)
os.makedirs(UPLOAD_DIR, exist_ok=True)


def _ensure_sqlite_db_is_writable(default_db_path: str) -> str:
    """
    SQLite needs write permission on both the .db file and its folder
    because it creates -journal / -wal files during INSERT/UPDATE.
    If the project is opened from a read-only location, copy the DB to a
    writable folder in the user's home directory and use that copy instead.
    """
    default_db_path = os.path.abspath(default_db_path)
    default_dir = os.path.dirname(default_db_path)

    def _try_make_writable(path: str, mode: int):
        try:
            if os.path.exists(path):
                os.chmod(path, mode)
        except Exception:
            pass

    _try_make_writable(default_dir, 0o755)
    _try_make_writable(default_db_path, 0o644)

    try:
        os.makedirs(default_dir, exist_ok=True)
        test_path = os.path.join(default_dir, ".write_test")
        with open(test_path, "w", encoding="utf-8") as f:
            f.write("ok")
        os.remove(test_path)
        if os.path.exists(default_db_path) and not os.access(default_db_path, os.W_OK):
            raise PermissionError("SQLite database file is not writable")
        return default_db_path
    except Exception:
        # Fallback for read-only folders, external drives, or packaged deployments.
        writable_dir = os.path.join(os.path.expanduser("~"), ".tournoi_petanque")
        os.makedirs(writable_dir, exist_ok=True)
        writable_db_path = os.path.join(writable_dir, "tournoi.db")
        if os.path.exists(default_db_path) and not os.path.exists(writable_db_path):
            import shutil
            shutil.copy2(default_db_path, writable_db_path)
        _try_make_writable(writable_db_path, 0o644)
        print(f"[DB] Using writable SQLite database: {writable_db_path}")
        return writable_db_path


database_url = os.environ.get("DATABASE_URL")
if database_url:
    if database_url.startswith("postgres://"):
        database_url = database_url.replace("postgres://", "postgresql://", 1)
    app.config['SQLALCHEMY_DATABASE_URI'] = database_url
else:
    DB_PATH = _ensure_sqlite_db_is_writable(os.path.join(INSTANCE_DIR, 'tournoi.db'))
    app.config['SQLALCHEMY_DATABASE_URI'] = f"sqlite:///{DB_PATH}"

app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config["UPLOAD_FOLDER"] = UPLOAD_DIR

socketio = SocketIO(
    app,
    async_mode="threading",
    cors_allowed_origins="*",
    ping_timeout=60,
    ping_interval=25,
    max_http_buffer_size=10_000_000,
    logger=False,
    engineio_logger=False
)

db.init_app(app)  # ✅ ตรงนี้สำคัญ
migrate = Migrate(app, db)
migrate.init_app(app, db)
app.register_blueprint(match_bp, url_prefix='/match')  # ลงทะเบียน blueprint
event_bp = Blueprint('event', __name__)  # สร้าง blueprint event ถ้ามี
# Setup Flask-Login
login_manager = LoginManager()
login_manager.login_view = "login"
login_manager.init_app(app)

def get_current_lang():
    lang = request.args.get("lang") or session.get("lang") or "th"
    if lang not in SUPPORTED_LANGS:
        lang = "th"
    session["lang"] = lang
    return lang


@app.context_processor
def inject_i18n():
    lang = get_current_lang()
    return {
        "current_lang": lang,
        "supported_langs": SUPPORTED_LANGS,
        "text_translations": TEXT_TRANSLATIONS,
        "t": lambda key: translate(key, lang),
    }


@app.route("/set-language/<lang>")
def set_language(lang):
    if lang not in SUPPORTED_LANGS:
        lang = "th"
    session["lang"] = lang
    next_url = request.args.get("next") or request.referrer or url_for("index")
    return redirect(next_url)


def ensure_runtime_columns():
    """เพิ่มคอลัมน์ใหม่ให้ฐานข้อมูลเดิมโดยไม่ลบข้อมูลเก่า"""
    dialect = db.engine.dialect.name
    if dialect == 'postgresql':
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS pending_team1_score INTEGER"))
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS pending_team2_score INTEGER"))
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS pending_is_submitted BOOLEAN DEFAULT FALSE NOT NULL"))
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS pending_submitted_by_id INTEGER"))
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS pending_submitted_at TIMESTAMP"))
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS team1_signature TEXT"))
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS team2_signature TEXT"))
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS scorer_signature TEXT"))
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS score_ends TEXT"))
        db.session.execute(text("ALTER TABLE matches ADD COLUMN IF NOT EXISTS scorecard_token VARCHAR(80)"))
        db.session.execute(text("CREATE UNIQUE INDEX IF NOT EXISTS ix_matches_scorecard_token ON matches (scorecard_token)"))
        db.session.execute(text("ALTER TABLE events ADD COLUMN IF NOT EXISTS is_finished BOOLEAN DEFAULT FALSE NOT NULL"))
        db.session.execute(text("ALTER TABLE events ADD COLUMN IF NOT EXISTS finished_at TIMESTAMP"))
        db.session.execute(text("ALTER TABLE events ADD COLUMN IF NOT EXISTS finished_by_id INTEGER"))
    else:
        event_existing = {row[1] for row in db.session.execute(text("PRAGMA table_info(events)"))}
        if 'is_finished' not in event_existing:
            db.session.execute(text("ALTER TABLE events ADD COLUMN is_finished BOOLEAN DEFAULT 0 NOT NULL"))
        if 'finished_at' not in event_existing:
            db.session.execute(text("ALTER TABLE events ADD COLUMN finished_at DATETIME"))
        if 'finished_by_id' not in event_existing:
            db.session.execute(text("ALTER TABLE events ADD COLUMN finished_by_id INTEGER"))

        existing = {row[1] for row in db.session.execute(text("PRAGMA table_info(matches)"))}
        if 'pending_team1_score' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN pending_team1_score INTEGER"))
        if 'pending_team2_score' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN pending_team2_score INTEGER"))
        if 'pending_is_submitted' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN pending_is_submitted BOOLEAN DEFAULT 0 NOT NULL"))
        if 'pending_submitted_by_id' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN pending_submitted_by_id INTEGER"))
        if 'pending_submitted_at' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN pending_submitted_at DATETIME"))
        if 'team1_signature' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN team1_signature TEXT"))
        if 'team2_signature' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN team2_signature TEXT"))
        if 'scorer_signature' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN scorer_signature TEXT"))
        if 'score_ends' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN score_ends TEXT"))
        if 'scorecard_token' not in existing:
            db.session.execute(text("ALTER TABLE matches ADD COLUMN scorecard_token VARCHAR(80)"))
    db.session.commit()


def ensure_match_scorecard_token(match):
    """สร้าง token ลับสำหรับ QR สกอร์การ์ดของแมตช์ ถ้ายังไม่มี"""
    if getattr(match, "scorecard_token", None):
        return match.scorecard_token

    match.scorecard_token = secrets.token_urlsafe(24)
    return match.scorecard_token


def ensure_match_tokens(matches):
    changed = False
    for match in matches:
        if not getattr(match, "scorecard_token", None):
            ensure_match_scorecard_token(match)
            changed = True
    if changed:
        db.session.commit()
    return matches


def ensure_playoff_tables():
    """Create persistent playoff/bracket tables used after a Swiss standing is finished."""
    dialect = db.engine.dialect.name
    if dialect == 'postgresql':
        stmts = [
            """CREATE TABLE IF NOT EXISTS playoff_competitions (
                id SERIAL PRIMARY KEY, source_event_id INTEGER NOT NULL, title VARCHAR(255) NOT NULL,
                competition_type VARCHAR(40) NOT NULL, pairing_method VARCHAR(30) NOT NULL DEFAULT 'seed',
                status VARCHAR(30) NOT NULL DEFAULT 'active', report_note TEXT, config_json TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS playoff_rounds (
                id SERIAL PRIMARY KEY, playoff_id INTEGER NOT NULL, round_no INTEGER NOT NULL,
                round_name VARCHAR(255) NOT NULL, round_type VARCHAR(40) NOT NULL, round_meta_json TEXT, created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS playoff_slots (
                id SERIAL PRIMARY KEY, round_id INTEGER NOT NULL, group_no INTEGER NOT NULL, slot_no INTEGER NOT NULL,
                seed INTEGER, team_id INTEGER, team_name VARCHAR(255), is_bye BOOLEAN DEFAULT FALSE, court_name VARCHAR(80),
                entry_label VARCHAR(80), source_label VARCHAR(255),
                UNIQUE(round_id, group_no, slot_no)
            )""",
            """CREATE TABLE IF NOT EXISTS playoff_scores (
                id SERIAL PRIMARY KEY, round_id INTEGER NOT NULL, group_no INTEGER NOT NULL, slot_no INTEGER NOT NULL,
                stage_no INTEGER NOT NULL, score INTEGER, updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(round_id, group_no, slot_no, stage_no)
            )""",
            """CREATE TABLE IF NOT EXISTS playoff_manual_results (
                id SERIAL PRIMARY KEY, round_id INTEGER NOT NULL, group_no INTEGER NOT NULL,
                winner_slot_no INTEGER, second_slot_no INTEGER, UNIQUE(round_id, group_no)
            )""",
        ]
    else:
        stmts = [
            """CREATE TABLE IF NOT EXISTS playoff_competitions (
                id INTEGER PRIMARY KEY AUTOINCREMENT, source_event_id INTEGER NOT NULL, title TEXT NOT NULL,
                competition_type TEXT NOT NULL, pairing_method TEXT NOT NULL DEFAULT 'seed',
                status TEXT NOT NULL DEFAULT 'active', report_note TEXT, config_json TEXT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS playoff_rounds (
                id INTEGER PRIMARY KEY AUTOINCREMENT, playoff_id INTEGER NOT NULL, round_no INTEGER NOT NULL,
                round_name TEXT NOT NULL, round_type TEXT NOT NULL, round_meta_json TEXT, created_at DATETIME DEFAULT CURRENT_TIMESTAMP
            )""",
            """CREATE TABLE IF NOT EXISTS playoff_slots (
                id INTEGER PRIMARY KEY AUTOINCREMENT, round_id INTEGER NOT NULL, group_no INTEGER NOT NULL, slot_no INTEGER NOT NULL,
                seed INTEGER, team_id INTEGER, team_name TEXT, is_bye BOOLEAN DEFAULT 0, court_name TEXT,
                entry_label TEXT, source_label TEXT,
                UNIQUE(round_id, group_no, slot_no)
            )""",
            """CREATE TABLE IF NOT EXISTS playoff_scores (
                id INTEGER PRIMARY KEY AUTOINCREMENT, round_id INTEGER NOT NULL, group_no INTEGER NOT NULL, slot_no INTEGER NOT NULL,
                stage_no INTEGER NOT NULL, score INTEGER, updated_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(round_id, group_no, slot_no, stage_no)
            )""",
            """CREATE TABLE IF NOT EXISTS playoff_manual_results (
                id INTEGER PRIMARY KEY AUTOINCREMENT, round_id INTEGER NOT NULL, group_no INTEGER NOT NULL,
                winner_slot_no INTEGER, second_slot_no INTEGER, UNIQUE(round_id, group_no)
            )""",
        ]
    for stmt in stmts:
        db.session.execute(text(stmt))

    # อัปเกรดฐานข้อมูลเดิมโดยไม่ลบข้อมูลเก่า
    if dialect == 'postgresql':
        db.session.execute(text("ALTER TABLE playoff_competitions ADD COLUMN IF NOT EXISTS config_json TEXT"))
        db.session.execute(text("ALTER TABLE playoff_rounds ADD COLUMN IF NOT EXISTS round_meta_json TEXT"))
        db.session.execute(text("ALTER TABLE playoff_slots ADD COLUMN IF NOT EXISTS entry_label VARCHAR(80)"))
        db.session.execute(text("ALTER TABLE playoff_slots ADD COLUMN IF NOT EXISTS source_label VARCHAR(255)"))
    else:
        comp_cols = {row[1] for row in db.session.execute(text("PRAGMA table_info(playoff_competitions)"))}
        round_cols = {row[1] for row in db.session.execute(text("PRAGMA table_info(playoff_rounds)"))}
        slot_cols = {row[1] for row in db.session.execute(text("PRAGMA table_info(playoff_slots)"))}
        if 'config_json' not in comp_cols:
            db.session.execute(text("ALTER TABLE playoff_competitions ADD COLUMN config_json TEXT"))
        if 'round_meta_json' not in round_cols:
            db.session.execute(text("ALTER TABLE playoff_rounds ADD COLUMN round_meta_json TEXT"))
        if 'entry_label' not in slot_cols:
            db.session.execute(text("ALTER TABLE playoff_slots ADD COLUMN entry_label TEXT"))
        if 'source_label' not in slot_cols:
            db.session.execute(text("ALTER TABLE playoff_slots ADD COLUMN source_label TEXT"))
    db.session.commit()


with app.app_context():
    db.create_all()  # สร้างตารางตามโมเดล
    ensure_runtime_columns()  # อัปเกรดฐานข
    ensure_playoff_tables()  # ตารางระบบน็อคเอาท์/ดับเบิ้ลหลังจบ Standing
    #ensure_match_tokens(Match.query.all())  # สร้าง token QR ให้แมตช์เก่า้อมูลเดิมแบบไม่ลบข้อมูล

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

@login_manager.unauthorized_handler
def unauthorized():
    """ถ้ายังไม่ล็อกอินแล้วเปิดหน้าที่ต้องใช้สิทธิ์ ให้พาไปหน้า login แทน 403"""
    flash("กรุณาเข้าสู่ระบบก่อน", "warning")
    return redirect(url_for("login", next=request.path))

#------------------------เรียลไทม์ SocketIO-------------------------------------------------------------


def _round_live_state(event_id, round_no=None):
    """สร้าง fingerprint สำหรับการเปลี่ยนแปลงโครงคู่/สนามเท่านั้น

    สำคัญ: ไม่เอาคะแนน/pending/lock status เข้าไปคำนวณ version แล้ว
    เพราะเวลาคีย์คะแนนแบบ realtime จะได้ไม่ทำให้หน้าใบประกบ/ใบบันทึก reload ทั้งหน้า
    """
    query = Match.query.filter_by(event_id=event_id)
    if round_no:
        query = query.filter_by(round=round_no)

    matches = query.order_by(Match.round.asc(), Match.field.asc(), Match.id.asc()).all()
    team_ids = set()
    for match in matches:
        if match.team1_id:
            team_ids.add(match.team1_id)
        if match.team2_id:
            team_ids.add(match.team2_id)

    team_map = {team.id: team.name for team in Team.query.filter(Team.id.in_(team_ids)).all()} if team_ids else {}
    rows = []
    version_rows = []
    for match in matches:
        row = {
            'id': match.id,
            'round': match.round,
            'field': str(match.field) if match.field is not None else '',
            'team1_id': match.team1_id,
            'team1_name': team_map.get(match.team1_id, '-'),
            'team2_id': match.team2_id,
            'team2_name': team_map.get(match.team2_id, 'BYE') if match.team2_id else 'BYE',
            'team1_score': match.team1_score,
            'team2_score': match.team2_score,
            'pending_team1_score': match.pending_team1_score,
            'pending_team2_score': match.pending_team2_score,
            'pending_is_submitted': bool(match.pending_is_submitted),
            'is_locked': bool(match.is_locked),
            'is_manual': bool(match.is_manual),
        }
        rows.append(row)
        version_rows.append({
            'id': match.id,
            'round': match.round,
            'field': row['field'],
            'team1_id': match.team1_id,
            'team2_id': match.team2_id,
            'is_manual': bool(match.is_manual),
        })

    raw = json.dumps(version_rows, ensure_ascii=False, sort_keys=True, default=str)
    version = hashlib.sha1(raw.encode('utf-8')).hexdigest()
    return version, rows


def _emit_round_live_changed(event_id, round_no, force_reload=False, reason='pairing'):
    """ส่งสัญญาณให้หน้าประกบคู่/ใบบันทึก reload เฉพาะตอนโครงคู่/สนามเปลี่ยนจริง"""
    try:
        version, rows = _round_live_state(event_id, round_no)
        payload = {
            'event_id': event_id,
            'round_no': round_no,
            'version': version,
            'matches': rows,
            'force_reload': bool(force_reload),
            'reason': reason,
            'timestamp': datetime.utcnow().isoformat(),
        }
        socketio.emit('round_pairing_updated', payload, to=f'event_{event_id}_round_{round_no}')
        socketio.emit(
            'event_pairing_updated',
            {
                'event_id': event_id,
                'round_no': round_no,
                'version': version,
                'force_reload': bool(force_reload),
                'reason': reason,
                'timestamp': payload['timestamp'],
            },
            to=f'event_{event_id}_all'
        )
    except Exception as exc:
        print(f'emit round live changed failed: {exc}')


@app.route('/event/<int:event_id>/round/<int:round_no>/live-version')
def round_live_version(event_id, round_no):
    Event.query.get_or_404(event_id)
    version, rows = _round_live_state(event_id, round_no)
    return jsonify({
        'ok': True,
        'event_id': event_id,
        'round_no': round_no,
        'version': version,
        'matches': rows,
    })


@app.route('/event/<int:event_id>/live-version')
def event_live_version(event_id):
    Event.query.get_or_404(event_id)
    version, rows = _round_live_state(event_id, None)
    return jsonify({
        'ok': True,
        'event_id': event_id,
        'round_no': None,
        'version': version,
        'matches': rows,
    })


@socketio.on('join_round')
def handle_join_round(data):
    try:
        from flask_socketio import join_room
        event_id = int((data or {}).get('event_id', 0))
        round_no = (data or {}).get('round_no')
        if event_id:
            join_room(f'event_{event_id}_all')
        if event_id and round_no not in (None, '', 'null'):
            join_room(f'event_{event_id}_round_{int(round_no)}')
    except Exception:
        pass


@socketio.on('join_playoff')
def handle_join_playoff(data):
    try:
        playoff_id = int((data or {}).get('playoff_id', 0))
        if playoff_id:
            from flask_socketio import join_room
            join_room(f'playoff_{playoff_id}')
    except Exception:
        pass

@socketio.on('update_score')
@login_required
def handle_update_score(data):
    """อัปเดตคะแนนจริงจากหน้า round สำหรับ admin/superadmin เท่านั้น"""
    if current_user.role not in ['admin', 'superadmin']:
        emit('error_message', {'message': 'เฉพาะ admin เท่านั้นที่แก้คะแนนจริงจากหน้า round ได้'}, room=request.sid)
        return

    match_id = data.get('match_id')
    team_a_score = data.get('team_a_score')
    team_b_score = data.get('team_b_score')
    username = current_user.username

    with app.app_context():
        match = Match.query.get(match_id)
        if match:
            try:
                match.team1_score = int(team_a_score)
                match.team2_score = int(team_b_score)
                db.session.commit()

                emit('score_updated', {
                    'match_id': match.id,
                    'team_a_score': match.team1_score,
                    'team_b_score': match.team2_score,
                    'pending_team_a_score': match.pending_team1_score,
                    'pending_team_b_score': match.pending_team2_score,
                    'has_pending': match.pending_team1_score is not None or match.pending_team2_score is not None,
                    'pending_is_submitted': bool(match.pending_is_submitted),
                    'updated_by_username': username,
                    'timestamp': datetime.utcnow().isoformat()
                }, broadcast=True)
                # คะแนน realtime มี socket score_updated อยู่แล้ว ไม่ต้องสั่ง reload หน้า

            except ValueError:
                emit('error_message', {'message': 'Invalid score format'}, room=request.sid)
            except Exception as e:
                db.session.rollback()
                print(f"Database error updating score for Match {match_id}: {e}")
                emit('error_message', {'message': 'Server error updating score'}, room=request.sid)
        else:
            emit('error_message', {'message': 'Match not found'}, room=request.sid)



# ฟังก์ชันแปลงวันที่เป็นภาษาไทยแบบเต็ม
def thai_date_full(dt):
    days = ['วันจันทร์', 'วันอังคาร', 'วันพุธ', 'วันพฤหัสบดี', 'วันศุกร์', 'วันเสาร์', 'วันอาทิตย์']
    months = ['มกราคม', 'กุมภาพันธ์', 'มีนาคม', 'เมษายน', 'พฤษภาคม', 'มิถุนายน',
              'กรกฎาคม', 'สิงหาคม', 'กันยายน', 'ตุลาคม', 'พฤศจิกายน', 'ธันวาคม']
    
    day_name = days[dt.weekday()]  # weekday() จันทร์=0
    day = dt.day
    month_name = months[dt.month - 1]
    year = dt.year + 543  # ปี พ.ศ.
    
    return f"{day_name} ที่ {day} {month_name} {year}"

# ลงทะเบียนฟังก์ชันนี้ให้ Jinja template ใช้งานได้
app.jinja_env.globals.update(thai_date_full=thai_date_full)

def roles_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if not current_user.is_authenticated:
                flash("กรุณาเข้าสู่ระบบก่อน", "warning")
                return redirect(url_for('login'))

            role = current_user.role.lower()
            print(f"DEBUG current_user.role = {role}")

            if role == 'superadmin':
                return f(*args, **kwargs)

            if role in [r.lower() for r in roles]:
                return f(*args, **kwargs)

            flash("คุณไม่มีสิทธิ์ทำรายการนี้", "warning")
            return redirect(url_for('index'))
        return decorated_function
    return decorator

# ฟังก์ชันช่วยแปลงค่าเป็น int และคืน 0 ถ้าค่าเป็น None หรือไม่ถูกต้อง
def safe_int(value):
    try:
        return int(value)
    except (TypeError, ValueError):
        return 0




def _natural_alpha_key(name):
    """คีย์เรียงชื่อแบบ A-Z และเข้าใจตัวเลขท้ายชื่อ เช่น TEAM 2 มาก่อน TEAM 10"""
    text = (name or '').strip().upper()
    text = re.sub(r'\s+', ' ', text)
    return [int(part) if part.isdigit() else part for part in re.split(r'(\d+)', text)]


def _order_pair_by_alphabet(team1_id, team2_id, team_lookup=None):
    """สลับตำแหน่งในคู่ให้ชื่อที่มาก่อนตามอัลฟาเบตอยู่ฝั่งซ้าย โดยยังเป็นคู่เดิม
    - BYE/X: ทีมจริงต้องอยู่ฝั่งซ้ายเหมือนเดิม
    - ใช้กับ Swiss auto pairing เท่านั้น
    """
    if not team1_id or not team2_id:
        return team1_id, team2_id

    def _name(tid):
        if team_lookup is not None:
            obj = team_lookup.get(tid)
            if obj is None:
                return ''
            return obj.name if hasattr(obj, 'name') else str(obj)
        team = Team.query.get(tid)
        return team.name if team else ''

    name1 = _name(team1_id)
    name2 = _name(team2_id)
    if _natural_alpha_key(name2) < _natural_alpha_key(name1):
        return team2_id, team1_id
    return team1_id, team2_id


def swiss_pairing(event_id, round_no, separate_same_name=False):
    # ตรวจสอบว่ารอบก่อนหน้า (round_no - 1) ถูกล็อกผลหมดหรือยัง
    if round_no > 1:
        unlocked_matches = Match.query.filter_by(event_id=event_id, round=round_no - 1, is_locked=False).all()
        if unlocked_matches:
            return False, f"กรุณาล็อกผลการแข่งขันรอบที่ {round_no - 1} ก่อนจับคู่รอบถัดไป"

    # ลบแมตช์รอบนี้ถ้ามีอยู่ก่อน เพื่อจับคู่ใหม่
    Match.query.filter_by(event_id=event_id, round=round_no).delete()
    db.session.commit()

    # เรียกใช้ฟังก์ชัน generate_pairings เพื่อจับคู่
    pairing_results = generate_pairings(event_id, round_no, separate_same_name=separate_same_name)
    
    # ถ้าจับไม่ได้จริง ๆ มักเกิดจากกฎแข็ง เช่น ติ๊กห้ามทีมชื่อฐานเดียวกันเจอกัน
    # แต่จำนวนทีม/ชื่อฐานทำให้จัดครบไม่ได้ จึงส่งไปจัดด้วยมือ
    if pairing_results is None:
        return False, "ระบบไม่สามารถจัดคู่ตามเงื่อนไขที่เลือกได้ ต้องจัดการด้วยมือ"

    team_lookup = {team.id: team for team in Team.query.filter_by(event_id=event_id).all()}
    matches = []
    for p in pairing_results:
        team1_id, team2_id = _order_pair_by_alphabet(p[0], p[1], team_lookup)
        match = Match(
            round=round_no,
            team1_id=team1_id,
            team2_id=team2_id,
            event_id=event_id,
            is_locked=False,
        )
        if team2_id is None:
            match.team1_score = 1
            match.team2_score = 0
            match.is_locked = True
        matches.append(match)


    db.session.add_all(matches)
    db.session.commit()

    return True, "จับคู่สำเร็จ"


@app.route("/event/<int:event_id>/manual_pairing/<int:round_num>", methods=["GET", "POST"])
@login_required
@roles_required("admin")
def manual_pairing(event_id, round_num):
    event = Event.query.get_or_404(event_id)
     # ดึงข้อมูลทีมจากฐานข้อมูลหรือที่เก็บไว้ (ตัวอย่างสมมติ)
    
    # ✅ แปลง logo_filename เป็น list
    try:
        event.logo_list = json.loads(event.logo_filename) if event.logo_filename else []
    except Exception:
        event.logo_list = []

    if request.method == "POST":
        # รับข้อมูลคู่จากฟอร์ม เช่น pairs=["1,2", "3,0", ...]  (0=BYE)
        pairs_raw = request.form.getlist("pairs")
        pairs = []
        
        for p in pairs_raw:
            try:
                t1, t2 = p.split(",")
                t1_id = int(t1)
                t2_id = int(t2) if t2.isdigit() and int(t2) != 0 else None  # None = BYE
                pairs.append((t1_id, t2_id))
            except Exception:
                flash("ข้อมูลคู่แข่งขันไม่ถูกต้อง", "danger")
                return redirect(request.url)

        # ลบแมตช์เดิมในรอบนี้ทั้งหมด (ไม่แยก is_manual)
        Match.query.filter_by(event_id=event_id, round=round_num).delete()
        
        # บันทึกแมตช์คู่ใหม่ตามแมนนวลโดยตรง ไม่เช็คกฎใดๆ
        for t1_id, t2_id in pairs:
            match = Match(event_id=event_id, round=round_num,
                          team1_id=t1_id, team2_id=t2_id,
                          is_manual=True, is_locked=False)
            db.session.add(match)
        db.session.commit()

        flash(f"บันทึกคู่แข่งขันรอบที่ {round_num} แบบแมนนวลเรียบร้อยแล้ว", "success")
        return redirect(url_for("round_matches", event_id=event_id, round=round_num))

    else:
        # GET: แสดงฟอร์มให้แก้ไขคู่
        standings = calculate_standings(event_id)
        team_ids = [team['team_id'] for team in standings]
        
        

        # ดึงชื่อทีมจากฐานข้อมูล
        teams_query = Team.query.filter(Team.id.in_(team_ids)).all()
        teams = {team.id: team.name for team in teams_query}
         # ดึงข้อมูลทีมจากฐานข้อมูลหรือที่เก็บไว้ (ตัวอย่างสมมติ)
        suggested_pairings, unpaired = generate_manual_pairings(event_id, team_ids)

        # สร้างข้อความโน้ตจากคู่แนะนำ
        pairing_notes = []
        team_dict = {team['team_id']: team['team_name'] for team in standings}
        

        for pair in suggested_pairings:
            t1 = team_dict.get(pair['team1_id'], 'Unknown Team')
            t2 = team_dict.get(pair['team2_id'], 'BYE') if pair['team2_id'] != 0 else 'BYE'
            pairing_notes.append(f" {t1} VS {t2}")

        for bye_team_id in unpaired:
            t = team_dict.get(bye_team_id, 'Unknown Team')
            pairing_notes.append(f" {t} ควรได้ BYE")


        # สร้างคู่คร่าวๆ จากฟังก์ชัน generate_manual_pairings (ยังไม่บันทึก)
        pairings, unpaired = generate_manual_pairings(event_id, team_ids)

        return render_template("admin_manual_pairings.html",
                               standings=standings, 
                               event=event,
                               round_num=round_num,
                               pairings=pairings,
                               pairings_count=(len(standings) + 1) // 2,   # <== เพิ่มตรงนี้
                               teams=teams,  # <== สำคัญ
                               pairing_notes=pairing_notes,
                               unpaired=unpaired)
        

@app.route("/")
def index():
    """หน้าแรก: แยกอีเว้นท์ที่กำลังดำเนินการ และประวัติตามปี/เดือน"""
    all_events = Event.query.order_by(Event.date.desc(), Event.id.desc()).all()

    event_ids = [event.id for event in all_events]
    team_counts = {}
    match_counts = {}
    field_counts = {}

    if event_ids:
        team_counts = dict(
            db.session.query(Team.event_id, func.count(Team.id))
            .filter(Team.event_id.in_(event_ids))
            .group_by(Team.event_id)
            .all()
        )
        match_counts = dict(
            db.session.query(Match.event_id, func.count(Match.id))
            .filter(Match.event_id.in_(event_ids))
            .group_by(Match.event_id)
            .all()
        )
        field_counts = dict(
            db.session.query(Match.event_id, func.count(func.distinct(Match.field)))
            .filter(Match.event_id.in_(event_ids), Match.field.isnot(None))
            .group_by(Match.event_id)
            .all()
        )

    active_events = []
    finished_nested = defaultdict(lambda: defaultdict(list))

    for event in all_events:
        # ค่าเหล่านี้ใช้เฉพาะแสดง Dashboard หน้าแรก ไม่เปลี่ยนฐานข้อมูล
        event.team_count = int(team_counts.get(event.id, 0) or 0)
        event.match_count = int(match_counts.get(event.id, 0) or 0)
        event.field_count = int(field_counts.get(event.id, 0) or 0)

        if getattr(event, 'is_finished', False):
            year = event.date.year if event.date else 0
            month = event.date.month if event.date else 0
            finished_nested[year][month].append(event)
        else:
            active_events.append(event)

    # เรียงอีเว้นท์กำลังดำเนินการ: วันที่ใกล้สุดก่อน และรายการไม่มีวันที่ไว้ท้ายสุด
    active_events = sorted(
        active_events,
        key=lambda e: (e.date is None, e.date or date.max, (e.name or '').casefold(), e.id)
    )

    thai_months = {
        0: 'ไม่ระบุเดือน', 1: 'มกราคม', 2: 'กุมภาพันธ์', 3: 'มีนาคม',
        4: 'เมษายน', 5: 'พฤษภาคม', 6: 'มิถุนายน', 7: 'กรกฎาคม',
        8: 'สิงหาคม', 9: 'กันยายน', 10: 'ตุลาคม', 11: 'พฤศจิกายน', 12: 'ธันวาคม'
    }

    finished_events_by_year_month = {}
    for year in sorted(finished_nested.keys(), reverse=True):
        months = {}
        for month in sorted(finished_nested[year].keys(), reverse=True):
            month_events = sorted(
                finished_nested[year][month],
                key=lambda e: (e.date or date.min, e.id),
                reverse=True
            )
            months[month] = {
                'month_name': thai_months.get(month, 'ไม่ระบุเดือน'),
                'events': month_events,
                'event_count': len(month_events),
                'team_count': sum(e.team_count for e in month_events),
                'match_count': sum(e.match_count for e in month_events),
                'field_count': sum(e.field_count for e in month_events),
            }
        finished_events_by_year_month[year] = months

    category_options = sorted({(e.category or '').strip() for e in all_events if (e.category or '').strip()})
    sex_options = sorted({(e.sex or '').strip() for e in all_events if (e.sex or '').strip()})

    return render_template(
        "index.html",
        upcoming_events=active_events,
        finished_events_by_year_month=finished_events_by_year_month,
        category_options=category_options,
        sex_options=sex_options,
        events=all_events,
    )


@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("index"))

    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")
        user = User.query.filter_by(username=username).first()
        
        if user is None:
            flash("ไม่พบชื่อผู้ใช้", "danger")
        elif not user.check_password(password):
            flash("รหัสผ่านไม่ถูกต้อง", "danger")
        elif user.end_time and user.end_time <= datetime.utcnow():
            flash("บัญชีนี้หมดอายุการใช้งานแล้ว กรุณาติดต่อผู้ดูแลระบบ", "danger")
        else:
            login_user(user)
            next_page = request.args.get("next")
            # ป้องกันการ redirect ไป URL ภายนอก (security)
            if not next_page or not next_page.startswith('/'):
                next_page = url_for("index")
            return redirect(next_page)
            
    return render_template("login.html")


@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("ออกจากระบบแล้ว", "success")
    return redirect(url_for("index"))

#-------------------------------------------------------------------------------

@app.route('/event/<int:event_id>/score-sheet')
@login_required
@roles_required('admin', 'superadmin')  # admin/superadmin พิมพ์สกอร์ชีทได้
def score_sheet_all(event_id):
  
    event = Event.query.get_or_404(event_id)
    selected_round = request.args.get('round', type=int)

    # ✅ แปลง logo_filename เป็น list
    try:
        event.logo_list = json.loads(event.logo_filename) if event.logo_filename else []
    except Exception:
        event.logo_list = []

    query = Match.query.filter_by(event_id=event_id)
    if selected_round:
        query = query.filter_by(round=selected_round)

    matches = query.order_by(Match.round, Match.id).all()
    ensure_match_tokens(matches)
    teams = Team.query.filter_by(event_id=event_id).all()
    num_teams = len(teams)

    # คำนวณจำนวนใบต่อหน้า (1 ใบต่อ 1 แมตช์)
    matches_per_page = max(1, num_teams // 2)  # อย่างน้อย 1 เพื่อป้องกันหารศูนย์

    total_rounds = db.session.query(func.max(Match.round)).filter_by(event_id=event_id).scalar() or 1
    live_version, _ = _round_live_state(event_id, selected_round)
    print(event.logo_list)
    return render_template(
        'score_sheet.html',
        event=event,
        matches=matches,
        teams=teams,
        total_rounds=total_rounds,
        num_teams=num_teams,
        matches_per_page=matches_per_page,  # ✅ ส่งเข้า template
        selected_round=selected_round,
        live_version=live_version
        
    )
#-------------------------------------------------------------------------------

@app.route("/event/<int:event_id>")
#--------------------------สำหรับเข้าดูอีเว้น-------------------------------------------------------------อย่าไปยุ่งกับมัน
def event_detail(event_id):
    event = db.session.get(Event, event_id)
    teams = Team.query.filter_by(event_id=event_id).all()
    standings = calculate_standings(event_id)
    matches = Match.query.filter_by(event_id=event_id).all()

   

    # 🔧 แก้ตรงนี้: คำนวณ current_round
    current_round = (
        db.session.query(db.func.max(Match.round))
        .filter(Match.event_id == event_id)
        .scalar()
        or 0
    )

    # ถ้าคุณใช้ matches_round_1 ใน template ก็เพิ่มตรงนี้ด้วย
    matches_round_1 = Match.query.filter_by(event_id=event_id, round=1).all()

    # ✅ แปลง logo_filename เป็น list
    try:
        event.logo_list = json.loads(event.logo_filename) if event.logo_filename else []
    except Exception:
        event.logo_list = []

    active_playoffs = _active_playoffs_for_event(event_id)

    return render_template(
        "event.html",
        event=event,
        teams=teams,
        standings=standings,
        matches=matches,
        current_round=current_round,      # ✅ ส่งไปยัง template
        matches_round_1=matches_round_1,  # ✅ ส่งไปด้วยหากใช้
        active_playoffs=active_playoffs
    )


@app.route('/event/<int:event_id>/finish', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def finish_event(event_id):
    event = db.session.get(Event, event_id)
    if not event:
        flash('ไม่พบอีเว้นท์', 'warning')
        return redirect(url_for('index'))

    event.is_finished = True
    event.finished_at = datetime.now()
    event.finished_by_id = current_user.id
    db.session.commit()
    flash(f'จบอีเว้นท์แล้ว: {event.name}', 'success')
    return redirect(url_for('event_detail', event_id=event.id))


@app.route('/event/<int:event_id>/reopen', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def reopen_event(event_id):
    event = db.session.get(Event, event_id)
    if not event:
        flash('ไม่พบอีเว้นท์', 'warning')
        return redirect(url_for('index'))

    event.is_finished = False
    event.finished_at = None
    event.finished_by_id = None
    db.session.commit()
    flash(f'เปิดอีเว้นท์กลับมาใช้งานแล้ว: {event.name}', 'success')
    return redirect(url_for('event_detail', event_id=event.id))


from flask import flash

@app.route("/event/<int:event_id>/upload", methods=["POST"])
@login_required
@roles_required('admin')
def upload_teams(event_id):
    file = request.files.get("file")

    if not file or file.filename == '':
        flash("กรุณาเลือกไฟล์เพื่ออัปโหลด", "danger")
        return redirect(url_for("event_detail", event_id=event_id))

    if not file.filename.endswith(('.xls', '.xlsx')):
        flash("กรุณาอัปโหลดไฟล์ Excel (.xls หรือ .xlsx)", "danger")
        return redirect(url_for("event_detail", event_id=event_id))

    try:
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(filepath)

        df = pd.read_excel(filepath)

        if "team_name" not in df.columns:
            flash("ไฟล์ต้องมีคอลัมน์ชื่อ 'team_name'", "danger")
            return redirect(url_for("event_detail", event_id=event_id))

        latest_round = get_latest_round_for_event(event_id)
        late_pair_mode = request.form.get('late_pair_mode', 'ask')
        new_teams = 0
        late_matches = 0

        for raw_name in df["team_name"]:
            name = str(raw_name).strip() if raw_name is not None else ""
            if not name or name.lower() == "nan":
                continue

            if not Team.query.filter_by(name=name, event_id=event_id).first():
                team = Team(name=name, event_id=event_id)
                db.session.add(team)
                db.session.flush()
                new_teams += 1

                if latest_round:
                    created = create_late_entry_match(event_id, team, latest_round)
                    if created:
                        late_matches += 1

        paired = remaining = 0
        if latest_round and late_pair_mode == 'auto':
            paired, remaining = auto_pair_late_entries(event_id, latest_round)

        db.session.commit()

        if latest_round:
            if late_pair_mode == 'auto':
                flash(
                    f"เพิ่มทีม/นักกีฬาสำเร็จ {new_teams} รายการ และระบบประกบทีมใหม่ให้อัตโนมัติ {paired} คู่ เหลือรอคู่ {remaining} รายการ โดยไม่กระทบคู่เดิม",
                    "success"
                )
            else:
                flash(
                    f"เพิ่มทีม/นักกีฬาสำเร็จ {new_teams} รายการ และสร้างแถวทีมเพิ่มภายหลังในรอบที่ {latest_round} จำนวน {late_matches} แถว โดยไม่กระทบผลจับสลากเดิม",
                    "success"
                )
        else:
            flash(f"เพิ่มทีมสำเร็จ {new_teams} ทีม", "success")
    except Exception as e:
        flash(f"เกิดข้อผิดพลาดในการอัปโหลด: {str(e)}", "danger")

    return redirect(url_for("event_detail", event_id=event_id))
def generate_field_numbers(event):
    prefix = event.field_prefix or ''
    start = event.field_start or 1
    max_field = event.field_max or 16
    exclude = set(event.field_exclude.split(',')) if event.field_exclude else set()

    fields = []
    for i in range(start, start + max_field):
        field_name = f"{prefix}{i}"
        if field_name not in exclude:
            fields.append(field_name)
    return fields


def get_latest_round_for_event(event_id):
    """คืนค่ารอบล่าสุดที่มีการจับคู่แล้ว ถ้ายังไม่จับคู่คืน None"""
    return db.session.query(db.func.max(Match.round)).filter_by(event_id=event_id).scalar()


def create_late_entry_match(event_id, team, target_round=None):
    """
    สร้างแถวคะแนนเฉพาะทีม/นักกีฬาที่เพิ่มภายหลัง โดยไม่ลบหรือสุ่มคู่เดิม
    ใช้ is_manual=False เพื่อให้ template แยกสี/ป้ายว่าเป็นรายการเพิ่มภายหลังได้
    """
    if target_round is None:
        target_round = get_latest_round_for_event(event_id)

    if not target_round:
        return None

    existing = Match.query.filter(
        Match.event_id == event_id,
        Match.round == target_round,
        ((Match.team1_id == team.id) | (Match.team2_id == team.id))
    ).first()
    if existing:
        return existing

    late_match = Match(
        event_id=event_id,
        round=target_round,
        team1_id=team.id,
        team2_id=None,
        team1_score=None,
        team2_score=None,
        is_locked=False,
        is_manual=False
    )
    db.session.add(late_match)
    return late_match



def _late_entry_matches(event_id, target_round=None):
    """รายการแถวทีมเพิ่มภายหลังที่ยังไม่มีคู่จริงในรอบที่กำหนด"""
    if target_round is None:
        target_round = get_latest_round_for_event(event_id)
    if not target_round:
        return []
    return Match.query.filter_by(
        event_id=event_id,
        round=target_round,
        team2_id=None,
        is_manual=False
    ).order_by(Match.id.asc()).all()


def _next_available_field(event, round_num):
    """หาเลขสนามถัดไป โดยไม่ไปแก้เลขสนามเดิม"""
    prefix = event.field_prefix or ''
    start = event.field_start or 1
    max_field = event.field_max or 16
    exclude = set(x.strip() for x in (event.field_exclude or '').split(',') if x.strip())
    used = set()
    for m in Match.query.filter_by(event_id=event.id, round=round_num).all():
        if m.field is not None and str(m.field).strip():
            used.add(str(m.field).strip())
    n = start
    while n < start + max_field * 20:
        label = f"{prefix}{n}"
        if label not in exclude and label not in used:
            return label
        n += 1
    return None


def auto_pair_late_entries(event_id, target_round=None):
    """ประกบเฉพาะทีมที่เพิ่มภายหลัง ไม่แตะคู่เดิม"""
    event = Event.query.get_or_404(event_id)
    if target_round is None:
        target_round = get_latest_round_for_event(event_id)
    late = _late_entry_matches(event_id, target_round)
    paired = 0
    while len(late) >= 2:
        m1 = late.pop(0)
        m2 = late.pop(0)
        m1.team2_id = m2.team1_id
        m1.team1_score = None
        m1.team2_score = None
        m1.is_manual = True
        if event.auto_field_enabled and not m1.field:
            m1.field = _next_available_field(event, target_round)
        db.session.delete(m2)
        paired += 1
    return paired, len(late)



def finalize_single_late_entry_as_bye(event_id, target_round=None):
    """ถ้าเพิ่มทีมมาทีมเดียวหลังมีการจับคู่แล้ว และไม่มีคู่ให้ประกบ ให้ถือว่าได้ BYE ทันที
    ใช้เฉพาะรายการเพิ่มภายหลังแบบ team2 ว่าง ไม่แตะคู่แข่งขันจริงเดิม
    """
    if target_round is None:
        target_round = get_latest_round_for_event(event_id)
    if not target_round:
        return 0
    late = _late_entry_matches(event_id, target_round)
    if len(late) != 1:
        return 0
    late_match = late[0]
    # ถ้ายังมีคู่จริงที่ยังไม่ล็อก ให้ยังไม่ล็อก BYE เพื่อให้ผู้จัดการตรวจเองก่อน
    real_unlocked = Match.query.filter(
        Match.event_id == event_id,
        Match.round == target_round,
        Match.id != late_match.id,
        Match.team2_id != None,
        Match.is_locked == False,
    ).count()
    if real_unlocked > 0:
        return 0
    late_match.team1_score = 1
    late_match.team2_score = 0
    late_match.is_locked = True
    late_match.field = None
    return 1


@app.route('/event/<int:event_id>/add_team', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def add_team_route(event_id):
    Event.query.get_or_404(event_id)

    team_name = (request.form.get('team_name') or '').strip()
    if not team_name:
        flash('กรุณากรอกชื่อทีม/ชื่อนักกีฬา', 'danger')
        return redirect(url_for('event_detail', event_id=event_id))

    existing_team = Team.query.filter_by(name=team_name, event_id=event_id).first()
    if existing_team:
        flash('ชื่อทีม/ชื่อนักกีฬานี้มีอยู่แล้ว', 'warning')
        return redirect(url_for('event_detail', event_id=event_id))

    latest_round = get_latest_round_for_event(event_id)

    new_team = Team(name=team_name, event_id=event_id)
    db.session.add(new_team)
    db.session.flush()  # ให้ได้ id ก่อนสร้างแถวคะแนน

    if latest_round:
        create_late_entry_match(event_id, new_team, latest_round)
        late_pair_mode = request.form.get('late_pair_mode', 'ask')
        paired = remaining = 0
        bye_locked = 0
        if late_pair_mode == 'auto':
            paired, remaining = auto_pair_late_entries(event_id, latest_round)
            bye_locked = finalize_single_late_entry_as_bye(event_id, latest_round)
        else:
            # ถ้าเพิ่มมาแค่ทีมเดียวหลังคู่เดิมลงผล/ล็อกครบแล้ว ให้เป็น BYE ทันที ไม่ทำให้รอบค้าง
            bye_locked = finalize_single_late_entry_as_bye(event_id, latest_round)
        db.session.commit()
        if paired:
            flash(
                f'เพิ่ม {team_name} เรียบร้อยแล้ว และระบบประกบทีมเพิ่มใหม่ให้อัตโนมัติ {paired} คู่ โดยไม่กระทบคู่เดิม',
                'success'
            )
        elif bye_locked:
            flash(
                f'เพิ่ม {team_name} เรียบร้อยแล้ว และให้ BYE ในรอบที่ {latest_round} อัตโนมัติ เพราะไม่มีคู่ใหม่ให้ประกบ',
                'success'
            )
        else:
            flash(
                f'เพิ่ม {team_name} เรียบร้อยแล้ว และสร้างแถวทีมเพิ่มภายหลังในรอบที่ {latest_round} โดยไม่กระทบผลจับสลากเดิม',
                'success'
            )
        return redirect(url_for('round_matches', event_id=event_id, round=latest_round))

    db.session.commit()
    flash('เพิ่มทีม/นักกีฬาเรียบร้อยแล้ว', 'success')
    return redirect(url_for('event_detail', event_id=event_id))


@app.route("/event/add", methods=["POST"])
@login_required
@roles_required('admin')
def add_event_route():
    try:
        name = request.form["name"]
        rounds = int(request.form["rounds"])
        location = request.form.get("location")
        category = request.form["category"]
        sex = request.form["sex"]
        age_group = request.form["age_group"]
        date_str = request.form.get("date")
        event_date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else None

        field_prefix = request.form.get("field_prefix", "")
        field_start = request.form.get("field_start", type=int)
        field_max = request.form.get("field_max", type=int)
        field_exclude = request.form.get("field_exclude", "")

        logo_files = request.files.getlist("logo")
        logo_filenames = []

        upload_folder = os.path.join("static", "logos")
        os.makedirs(upload_folder, exist_ok=True)

        # ฟังก์ชันลบพื้นหลัง (copy จากที่คุณมี)
        def remove_background(image, threshold=200):
            image = image.convert("RGBA")
            datas = image.getdata()
            new_data = []
            for item in datas:
                if item[0] > threshold and item[1] > threshold and item[2] > threshold:
                    new_data.append((255, 255, 255, 0))
                else:
                    new_data.append(item)
            image.putdata(new_data)
            return image

        def resize_logo_fixed_height(image, fixed_height=60):
            width, height = image.size
            new_height = fixed_height
            new_width = int((new_height / height) * width)
            return image.resize((new_width, new_height), Image.Resampling.LANCZOS)

        # ประมวลผลไฟล์โลโก้ทีละไฟล์
        for logo_file in logo_files:
            if logo_file and logo_file.filename != "":
                filename = secure_filename(logo_file.filename)
                logo_path = os.path.join(upload_folder, filename)

                image = Image.open(logo_file)
                image = remove_background(image, threshold=200)
                image = resize_logo_fixed_height(image, fixed_height=60)
                image.save(logo_path, format="PNG")

                logo_filenames.append(filename)

        new_event = Event(
            name=name,
            rounds=rounds,
            location=location,
            sex=sex,
            category=category,
            age_group=age_group,
            date=event_date,
            field_prefix=field_prefix,
            field_start=field_start,
            field_max=field_max,
            field_exclude=field_exclude,
            creator_id=current_user.id,
            logo_filename=json.dumps(logo_filenames)
        )

        db.session.add(new_event)
        db.session.commit()
        flash("เพิ่มรายการแข่งขันเรียบร้อยแล้ว", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"เกิดข้อผิดพลาด: {str(e)}", "danger")

    return redirect(url_for("index"))





@app.route('/event/<int:event_id>/team/<int:team_id>/edit', methods=['GET', 'POST'])
@login_required
@roles_required('admin')  # เพิ่มบรรทัดนี้
def edit_teams_route(event_id, team_id):
   
    team = Team.query.filter_by(id=team_id, event_id=event_id).first_or_404()
    if request.method == 'POST':
        new_name = request.form.get('name')
        if new_name:
            team.name = new_name
            db.session.commit()
            flash('แก้ไขชื่อทีมเรียบร้อย', 'success')
            return redirect(url_for('event_detail', event_id=event_id))
        else:
            flash('กรุณากรอกชื่อทีม', 'warning')
    return render_template('edit_team.html', team=team, event_id=event_id)


@app.route("/event/<int:event_id>/pair_first_round", methods=['POST'])
@login_required
@roles_required('admin')
def pair_first_round(event_id):
    import re
    import random
    from collections import defaultdict

    def extract_base_name(name):
        # ลบรหัสตัวเลข หรือเครื่องหมายด้านหลังชื่อ เช่น "ขอนแก่น 1", "ขอนแก่น-2" → "ขอนแก่น"
        base = re.split(r'[\s\-]*\d+$', name.strip())[0]
        return base

    event = Event.query.get(event_id)
    if event is None:
        flash("ไม่พบรายการแข่งขันนี้", "warning")
        return redirect(url_for("event_detail", event_id=event_id))

    existing_matches = Match.query.filter_by(event_id=event_id, round=1).count()
    if existing_matches > 0:
        flash("มีการจับคู่รอบแรกแล้ว", "warning")
        return redirect(url_for("event_detail", event_id=event_id))

    teams = Team.query.filter_by(event_id=event_id).all()
    if len(teams) < 2:
        flash("ต้องมีทีมอย่างน้อย 2 ทีมในการจับคู่", "warning")
        return redirect(url_for("event_detail", event_id=event_id))

    separate_same_name = request.form.get('separate_same_name') == 'on'

    # สุ่มลำดับทีม
    teams = teams[:]  # copy
    random.shuffle(teams)

    pairings = []
    used_ids = set()

    for i, team1 in enumerate(teams):
        if team1.id in used_ids:
            continue
        for j in range(i + 1, len(teams)):
            team2 = teams[j]
            if team2.id in used_ids:
                continue
            if separate_same_name and extract_base_name(team1.name) == extract_base_name(team2.name):
                continue  # ข้ามถ้ารากชื่อเหมือนกัน เช่น "ขอนแก่น 1" vs "ขอนแก่น 2"
            # จับคู่
            pairings.append((team1, team2))
            used_ids.add(team1.id)
            used_ids.add(team2.id)
            break

    # BYE หากยังเหลือทีมที่จับไม่ได้
    remaining = [t for t in teams if t.id not in used_ids]
    if remaining:
        pairings.append((remaining[0], None))

    team_lookup = {team.id: team for team in teams}
    for team1, team2 in pairings:
        team1_id, team2_id = _order_pair_by_alphabet(team1.id, team2.id if team2 else None, team_lookup)
        match = Match(
            event_id=event_id,
            round=1,
            team1_id=team1_id,
            team2_id=team2_id,
            is_locked=False
        )
        db.session.add(match)

    db.session.commit()
    flash("จับคู่รอบแรกเสร็จสิ้น", "success")
    return redirect(url_for("round_matches", event_id=event_id, round=1))



@app.route('/event/<int:event_id>/team/<int:team_id>/delete', methods=['POST'])
@login_required
@roles_required('admin')  # เพิ่มบรรทัดนี้
def delete_team_route(event_id, team_id):
    team = Team.query.get_or_404(team_id)

    # ตรวจสอบว่าเป็นทีมใน event นี้จริง
    if team.event_id != event_id:
        flash('ไม่พบทีมในรายการนี้', 'danger')
        return redirect(url_for('event_detail', event_id=event_id))

    # เช็คว่ามีแมตช์ที่ประกบคู่ทีมนี้แล้วหรือยัง
    existing_matches = Match.query.filter(
        ((Match.team1_id == team_id) | (Match.team2_id == team_id)) &
        (Match.event_id == event_id)
    ).count()

    

    # ถ้ายังไม่มีแมตช์ที่ประกบคู่ทีมนี้ ลบได้เลย
    Match.query.filter(
        ((Match.team1_id == team_id) | (Match.team2_id == team_id)) &
        (Match.event_id == event_id)
    ).delete()

    db.session.delete(team)
    db.session.commit()
    flash('ลบทีมเรียบร้อยแล้ว', 'success')
    return redirect(url_for('event_detail', event_id=event_id))

@app.route("/event/<int:event_id>/round/<int:round>/delete_all", methods=["POST"])
@login_required
@roles_required('admin', 'superadmin')
def delete_round_pairings(event_id, round):
    deleted = Match.query.filter_by(event_id=event_id, round=round).delete()
    db.session.commit()

    if deleted:
        flash(f"ลบคู่ประกบทั้งหมดของรอบที่ {round} เรียบร้อยแล้ว", "success")
    else:
        flash(f"ไม่พบคู่ประกบของรอบที่ {round}", "warning")

    return redirect(url_for("event_detail", event_id=event_id))

@app.route('/event/<int:event_id>/team/<int:team_id>/edit', methods=['POST'])
@login_required
@roles_required('admin')  # เพิ่มบรรทัดนี้
def edit_team(event_id, team_id):
    
    new_name = request.form.get('new_name')
    if not new_name:
        flash('กรุณากรอกชื่อทีมใหม่', 'danger')
        return redirect(url_for('event_detail', event_id=event_id))

    team = Team.query.get_or_404(team_id)
    if team.event_id != event_id:
        flash('ไม่พบทีมในรายการนี้', 'danger')
        return redirect(url_for('event_detail', event_id=event_id))

    team.name = new_name.strip()
    db.session.commit()
    flash('แก้ไขชื่อทีมเรียบร้อยแล้ว', 'success')
    return redirect(url_for('event_detail', event_id=event_id))


@app.route("/event/<int:event_id>/pair_next_round", methods=['POST'])
@login_required
@roles_required('admin')
def pair_next_round(event_id):
    import re
    import random
    from collections import defaultdict

    def extract_base_name(name):
        # ลบรหัสตัวเลข หรือเครื่องหมายด้านหลังชื่อ เช่น "ขอนแก่น 1", "ขอนแก่น-2" → "ขอนแก่น"
        base = re.split(r'[\s\-]*\d+$', name.strip())[0]
        return base
    
    max_round = db.session.query(db.func.max(Match.round)).filter_by(event_id=event_id).scalar()
    next_round = (max_round or 0) + 1

    if max_round is None:
        flash("ยังไม่มีการจับคู่รอบแรก กรุณาจับคู่รอบแรกก่อน", "warning")
        return redirect(url_for("event_detail", event_id=event_id))

    # ตรวจสอบว่าแมตช์รอบก่อนหน้าล็อกผลหมดหรือยัง
    unlocked_matches = Match.query.filter_by(event_id=event_id, round=max_round, is_locked=False).count()
    if unlocked_matches > 0:
        flash(f"กรุณาล็อกผลการแข่งขันรอบที่ {max_round} ก่อนจับคู่รอบถัดไป", "warning")
        return redirect(url_for("event_detail", event_id=event_id))

    event = Event.query.get(event_id)
    if not event:
        flash("ไม่พบรายการแข่งขันนี้", "danger")
        return redirect(url_for("index"))

    if next_round > event.rounds:
        flash("ครบจำนวนรอบการแข่งขันแล้ว ไม่สามารถจับคู่รอบใหม่ได้", "info")
        return redirect(url_for("event_detail", event_id=event_id))

    # อ่านค่า checkbox: ถ้าไม่ติ๊ก ทีมชื่อฐานเดียวกันสามารถเจอกันได้ตามปกติ
    separate_same_name = request.form.get('separate_same_name') == 'on'

    # 🔁 เรียก swiss_pairing แล้วตรวจสอบผลลัพธ์
    success, message = swiss_pairing(event_id, next_round, separate_same_name=separate_same_name)

    if not success:
        flash(message, "warning")

        # ถ้าระบบจัดตามเงื่อนไขไม่ได้ → ส่งไป manual pairing
        if "จัดการด้วยมือ" in message or "จัดคู่ตามเงื่อนไข" in message:
            return redirect(url_for("manual_pairing", event_id=event_id, round_num=next_round))

        # กรณีอื่นๆ กลับหน้า event_detail
        return redirect(url_for("event_detail", event_id=event_id))

    flash(f"จับคู่รอบที่ {next_round} เรียบร้อยแล้ว", "success")
    return redirect(url_for("round_matches", event_id=event_id, round=next_round))





@app.route("/event/<int:event_id>/match/<int:match_id>", methods=["POST"])
@login_required
@roles_required('admin')  # เพิ่มบรรทัดนี้
def submit_score(event_id, match_id):
 
    match = Match.query.get_or_404(match_id)
    match.team1_score = int(request.form.get("team1_score", 0))
    match.team2_score = int(request.form.get("team2_score", 0))
    db.session.commit()
    return redirect(url_for("event_detail", event_id=event_id))




# ---------- BETA: bulk event import / edit event ----------
def _clean_excel_value(value, default=""):
    """อ่านค่าจาก Excel ให้เป็นข้อความสะอาด ๆ ไม่ให้ nan หลุดไปแสดงในระบบ"""
    if value is None:
        return default
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass
    text_value = str(value).strip()
    if text_value.lower() == "nan":
        return default
    return text_value


def _pick_excel_value(row, *names, default=""):
    """รองรับชื่อคอลัมน์ทั้งอังกฤษและไทย เพื่อให้ทำไฟล์นำเข้าได้ง่าย"""
    for name in names:
        if name in row.index:
            value = _clean_excel_value(row.get(name), default="")
            if value != "":
                return value
    return default


def _parse_event_import_date(value):
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    text_value = str(value).strip()
    if not text_value or text_value.lower() == "nan":
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text_value, fmt).date()
        except ValueError:
            pass
    parsed = pd.to_datetime(text_value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.date()


def _parse_event_import_int(value, default=1):
    text_value = _clean_excel_value(value, "")
    if not text_value:
        return default
    try:
        return int(float(text_value))
    except Exception:
        return default


def _save_uploaded_logos(logo_files):
    logo_filenames = []
    upload_folder = os.path.join("static", "logos")
    os.makedirs(upload_folder, exist_ok=True)

    def remove_background(image, threshold=200):
        image = image.convert("RGBA")
        datas = image.getdata()
        new_data = []
        for item in datas:
            if item[0] > threshold and item[1] > threshold and item[2] > threshold:
                new_data.append((255, 255, 255, 0))
            else:
                new_data.append(item)
        image.putdata(new_data)
        return image

    def resize_logo_fixed_height(image, fixed_height=60):
        width, height = image.size
        if not height:
            return image
        new_height = fixed_height
        new_width = int((new_height / height) * width)
        return image.resize((new_width, new_height), Image.Resampling.LANCZOS)

    for logo_file in logo_files or []:
        if logo_file and logo_file.filename:
            filename = secure_filename(logo_file.filename)
            logo_path = os.path.join(upload_folder, filename)
            image = Image.open(logo_file)
            image = remove_background(image, threshold=200)
            image = resize_logo_fixed_height(image, fixed_height=60)
            image.save(logo_path, format="PNG")
            logo_filenames.append(filename)
    return logo_filenames


@app.route("/event/import-template")
@login_required
@roles_required('superadmin')
def download_event_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "events_import"
    headers = [
        "event_key", "event_name", "rounds", "date", "location", "category", "sex", "age_group",
        "field_prefix", "field_start", "field_max", "field_exclude", "team_name"
    ]
    ws.append(headers)
    sample_rows = [
        ["U18M_TEAM", "อบจ.ขอนแก่น U18 ทีมชาย", 3, "2026-07-08", "สนามเปตอง", "ทีม", "ชาย", "อายุไม่เกิน 18 ปี", "", 1, 27, "", "โรงเรียน ก"],
        ["U18M_TEAM", "อบจ.ขอนแก่น U18 ทีมชาย", 3, "2026-07-08", "สนามเปตอง", "ทีม", "ชาย", "อายุไม่เกิน 18 ปี", "", 1, 27, "", "โรงเรียน ข"],
        ["U16W_PAIR", "อบจ.ขอนแก่น U16 คู่หญิง", 3, "2026-07-09", "สนามเปตอง", "คู่", "หญิง", "อายุไม่เกิน 16 ปี", "", 1, 27, "", "โรงเรียน ค"],
    ]
    for row in sample_rows:
        ws.append(row)
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F81BD")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(idx)].width = max(14, len(header) + 4)
    ws.freeze_panes = "A2"
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        as_attachment=True,
        download_name="event_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/event/import", methods=["POST"])
@login_required
@roles_required('superadmin')
def import_events_from_excel():
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("กรุณาเลือกไฟล์ Excel สำหรับสร้างหลายอีเว้นท์", "danger")
        return redirect(url_for("index"))
    if not file.filename.lower().endswith((".xls", ".xlsx")):
        flash("กรุณาอัปโหลดไฟล์ Excel (.xls หรือ .xlsx)", "danger")
        return redirect(url_for("index"))

    try:
        df = pd.read_excel(file)
        if df.empty:
            flash("ไฟล์ Excel ไม่มีข้อมูล", "warning")
            return redirect(url_for("index"))

        required_any_name = {"event_name", "name", "ชื่อรายการ", "รายการ"}
        required_any_team = {"team_name", "team", "ชื่อทีม", "ชื่อนักกีฬา"}
        columns = set(str(c).strip() for c in df.columns)
        if not columns.intersection(required_any_name):
            flash("ไฟล์ต้องมีคอลัมน์ event_name หรือ ชื่อรายการ", "danger")
            return redirect(url_for("index"))
        if not columns.intersection(required_any_team):
            flash("ไฟล์ต้องมีคอลัมน์ team_name หรือ ชื่อทีม/ชื่อนักกีฬา", "danger")
            return redirect(url_for("index"))

        created_events = 0
        created_teams = 0
        skipped_duplicate_teams = 0
        groups = defaultdict(list)

        for _, row in df.iterrows():
            event_name = _pick_excel_value(row, "event_name", "name", "ชื่อรายการ", "รายการ")
            if not event_name:
                continue
            event_key = _pick_excel_value(row, "event_key", "รหัสอีเว้นท์", "รหัสรายการ")
            event_date = _parse_event_import_date(row.get("date") if "date" in row.index else row.get("วันที่") if "วันที่" in row.index else None)
            group_key = event_key or "|".join([
                event_name,
                str(event_date or ""),
                _pick_excel_value(row, "category", "ประเภท"),
                _pick_excel_value(row, "sex", "เพศ"),
                _pick_excel_value(row, "age_group", "รุ่น", "รุ่นอายุ"),
            ])
            groups[group_key].append(row)

        for _, rows in groups.items():
            first = rows[0]
            event_name = _pick_excel_value(first, "event_name", "name", "ชื่อรายการ", "รายการ")
            event_date = _parse_event_import_date(first.get("date") if "date" in first.index else first.get("วันที่") if "วันที่" in first.index else None)
            rounds = _parse_event_import_int(first.get("rounds") if "rounds" in first.index else first.get("จำนวนรอบ") if "จำนวนรอบ" in first.index else None, 3)
            field_start = _parse_event_import_int(first.get("field_start") if "field_start" in first.index else first.get("สนามเริ่ม") if "สนามเริ่ม" in first.index else None, 1)
            field_max = _parse_event_import_int(first.get("field_max") if "field_max" in first.index else first.get("จำนวนสนาม") if "จำนวนสนาม" in first.index else None, 16)

            new_event = Event(
                name=event_name,
                rounds=rounds,
                date=event_date,
                location=_pick_excel_value(first, "location", "สถานที่"),
                category=_pick_excel_value(first, "category", "ประเภท"),
                sex=_pick_excel_value(first, "sex", "เพศ"),
                age_group=_pick_excel_value(first, "age_group", "รุ่น", "รุ่นอายุ"),
                field_prefix=_pick_excel_value(first, "field_prefix", "คำนำหน้าสนาม"),
                field_start=field_start,
                field_max=field_max,
                field_exclude=_pick_excel_value(first, "field_exclude", "สนามที่เว้น"),
                creator_id=current_user.id,
                logo_filename=json.dumps([]),
            )
            db.session.add(new_event)
            db.session.flush()
            created_events += 1

            seen_team_names = set()
            for row in rows:
                team_name = _pick_excel_value(row, "team_name", "team", "ชื่อทีม", "ชื่อนักกีฬา")
                if not team_name:
                    continue
                team_key = team_name.strip().casefold()
                if team_key in seen_team_names:
                    skipped_duplicate_teams += 1
                    continue
                seen_team_names.add(team_key)
                db.session.add(Team(name=team_name, event_id=new_event.id))
                created_teams += 1

        db.session.commit()
        flash(f"นำเข้าเรียบร้อย: สร้างอีเว้นท์ {created_events} รายการ / เพิ่มทีม {created_teams} ทีม / ข้ามชื่อซ้ำ {skipped_duplicate_teams} รายการ", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"นำเข้าไม่สำเร็จ: {str(e)}", "danger")
    return redirect(url_for("index"))


@app.route("/event/<int:event_id>/edit", methods=["GET", "POST"])
@login_required
@roles_required('admin')
def edit_event(event_id):
    event = Event.query.get_or_404(event_id)
    if request.method == "POST":
        try:
            event.name = request.form.get("name", event.name).strip() or event.name
            event.rounds = int(request.form.get("rounds", event.rounds or 1))
            event.location = request.form.get("location", "")
            event.category = request.form.get("category", "")
            event.sex = request.form.get("sex", "")
            event.age_group = request.form.get("age_group", "")
            date_str = request.form.get("date")
            event.date = datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else None
            event.field_prefix = request.form.get("field_prefix", "")
            event.field_start = int(request.form.get("field_start") or 1)
            event.field_max = int(request.form.get("field_max") or 16)
            event.field_exclude = request.form.get("field_exclude", "")

            existing_logos = []
            try:
                existing_logos = json.loads(event.logo_filename) if event.logo_filename else []
            except Exception:
                existing_logos = []
            new_logos = _save_uploaded_logos(request.files.getlist("logo"))
            if new_logos:
                event.logo_filename = json.dumps(existing_logos + new_logos)

            db.session.commit()
            flash("แก้ไขอีเว้นท์เรียบร้อยแล้ว", "success")
            return redirect(url_for("index"))
        except Exception as e:
            db.session.rollback()
            flash(f"แก้ไขอีเว้นท์ไม่สำเร็จ: {str(e)}", "danger")

    try:
        event.logo_list = json.loads(event.logo_filename) if event.logo_filename else []
    except Exception:
        event.logo_list = []
    return render_template("edit_event.html", event=event)
# ---------- end beta bulk event import / edit event ----------

# ---------- BETA: multi-event score center ----------
def _selected_event_ids_from_request():
    """อ่านรายการ event id จาก checkbox/query string แล้วคืนเป็น list[int] แบบไม่ซ้ำ"""
    raw_ids = []
    raw_ids.extend(request.form.getlist("event_ids"))
    raw_ids.extend(request.args.getlist("event_ids"))
    comma_ids = request.args.get("events") or request.form.get("events") or ""
    if comma_ids:
        raw_ids.extend(comma_ids.split(","))

    event_ids = []
    seen = set()
    for raw in raw_ids:
        try:
            event_id = int(str(raw).strip())
        except Exception:
            continue
        if event_id not in seen:
            event_ids.append(event_id)
            seen.add(event_id)
    return event_ids


def _event_has_unlocked_matches(event_id):
    return Match.query.filter_by(event_id=event_id, is_locked=False).count() > 0


@app.route("/multi-score", methods=["GET", "POST"])
@login_required
@roles_required('superadmin')
def multi_event_score_center():
    """หน้าเดียวสำหรับเลือกหลายอีเว้นท์ แล้วคีย์คะแนนตามครั้ง/รอบรวมกัน"""
    today = date.today()
    selected_round = request.values.get("round", type=int) or 1
    status_filter = request.values.get("status", "all")
    field_query = (request.values.get("field") or "").strip()
    selected_event_ids = _selected_event_ids_from_request()

    events_query = Event.query.order_by(Event.date.desc().nullslast(), Event.id.desc())
    all_events = events_query.all()

    if not selected_event_ids:
        # ค่าเริ่มต้น: เลือกรายการวันนี้ที่ยังไม่จบก่อน เพื่อลดงานในวันแข่ง
        today_events = [e.id for e in all_events if e.date == today and _event_has_unlocked_matches(e.id)]
        selected_event_ids = today_events[:12]

    selected_events = [e for e in all_events if e.id in set(selected_event_ids)]
    selected_events.sort(key=lambda e: (e.date or date.min, e.name or "", e.id))
    # เลือกรอบแยกต่ออีเว้นท์ได้ เช่น U12 คีย์ครั้งที่ 2 แต่ U14 คีย์ครั้งที่ 3 ในหน้าเดียว
    event_round_map = _beta_event_rounds_map_from_request(selected_events, selected_round)
    event_rounds_param = _beta_event_rounds_param(event_round_map)

    if request.method == "POST" and request.form.get("action") in {"save_scores", "save_and_lock"}:
        action = request.form.get("action")
        selected_matches = (
            Match.query
            .filter(Match.id.in_([int(x) for x in request.form.getlist("match_ids") if str(x).isdigit()]))
            .all()
        )
        match_map = {m.id: m for m in selected_matches}
        updated = 0
        locked = 0
        draw_matches = []
        touched_rounds = set()

        for match_id, match in match_map.items():
            if match.is_locked:
                continue
            score1_raw = (request.form.get(f"score_{match_id}_1") or "").strip()
            score2_raw = (request.form.get(f"score_{match_id}_2") or "").strip()
            orig1_raw = (request.form.get(f"orig_{match_id}_1") or "").strip()
            orig2_raw = (request.form.get(f"orig_{match_id}_2") or "").strip()

            # ไม่กรอก = ข้ามแถวนั้น เพื่อให้คีย์แบบทยอยได้
            if score1_raw == "" and score2_raw == "":
                continue

            def _norm_score(value):
                return "" if value is None else str(value)

            current1_raw = _norm_score(match.team1_score)
            current2_raw = _norm_score(match.team2_score)
            submitted_changed = (score1_raw != orig1_raw) or (score2_raw != orig2_raw)
            db_changed_after_page_load = (current1_raw != orig1_raw) or (current2_raw != orig2_raw)
            submitted_differs_from_db = (score1_raw != current1_raw) or (score2_raw != current2_raw)
            if submitted_changed and db_changed_after_page_load and submitted_differs_from_db:
                team1 = match.team1.name if match.team1 else "-"
                team2 = match.team2.name if match.team2 else "BYE"
                flash(f"กันคะแนนทับกัน: {match.event.name} สนาม {match.field or '-'} ({team1} - {team2}) มีการแก้จากอีกหน้าก่อนแล้ว กรุณารีเฟรชตรวจอีกครั้ง", "danger")
                return redirect(url_for("multi_event_score_center", events=",".join(map(str, selected_event_ids)), round=selected_round, status=status_filter, field=field_query, event_rounds=event_rounds_param))

            if match.team2_id is None:
                # BYE/รายชื่อเดี่ยวในระบบเดิม ใช้คะแนนฝั่งแรก ฝั่งสองเป็น 0
                if score1_raw == "":
                    flash("มีรายการที่กรอกคะแนนไม่ครบ กรุณาตรวจช่องคะแนน", "danger")
                    return redirect(url_for("multi_event_score_center", events=",".join(map(str, selected_event_ids)), round=selected_round, status=status_filter, field=field_query, event_rounds=event_rounds_param))
                try:
                    match.team1_score = int(score1_raw)
                    match.team2_score = 0
                except Exception:
                    flash("คะแนนต้องเป็นตัวเลขเท่านั้น", "danger")
                    return redirect(url_for("multi_event_score_center", events=",".join(map(str, selected_event_ids)), round=selected_round, status=status_filter, field=field_query, event_rounds=event_rounds_param))
            else:
                if score1_raw == "" or score2_raw == "":
                    flash("มีรายการที่กรอกคะแนนไม่ครบ กรุณาตรวจช่องคะแนน", "danger")
                    return redirect(url_for("multi_event_score_center", events=",".join(map(str, selected_event_ids)), round=selected_round, status=status_filter, field=field_query, event_rounds=event_rounds_param))
                try:
                    match.team1_score = int(score1_raw)
                    match.team2_score = int(score2_raw)
                except Exception:
                    flash("คะแนนต้องเป็นตัวเลขเท่านั้น", "danger")
                    return redirect(url_for("multi_event_score_center", events=",".join(map(str, selected_event_ids)), round=selected_round, status=status_filter, field=field_query, event_rounds=event_rounds_param))

                if match.team1_score == match.team2_score:
                    team1 = match.team1.name if match.team1 else "-"
                    team2 = match.team2.name if match.team2 else "-"
                    draw_matches.append(f"{match.event.name} | สนาม {match.field or '-'} | {team1} {match.team1_score}-{match.team2_score} {team2}")

            updated += 1
            touched_rounds.add((match.event_id, match.round))

            if action == "save_and_lock":
                match.is_locked = True
                locked += 1

        db.session.commit()

        for event_id, round_no in touched_rounds:
            _emit_round_live_changed(event_id, round_no, force_reload=(action == "save_and_lock"), reason="multi_score_center")
            try:
                for m in Match.query.filter_by(event_id=event_id, round=round_no).all():
                    socketio.emit('score_updated', {
                        'match_id': m.id,
                        'team_a_score': m.team1_score,
                        'team_b_score': m.team2_score,
                        'pending_team_a_score': m.pending_team1_score,
                        'pending_team_b_score': m.pending_team2_score,
                        'has_pending': m.pending_team1_score is not None or m.pending_team2_score is not None,
                        'pending_is_submitted': bool(m.pending_is_submitted),
                        'updated_by_username': current_user.username,
                        'timestamp': datetime.utcnow().isoformat(),
                    })
            except Exception:
                pass

        if draw_matches:
            flash("บันทึกแล้ว แต่พบคู่เสมอ {} คู่: {}".format(len(draw_matches), " | ".join(draw_matches[:8])), "draw-warning")
        elif action == "save_and_lock":
            flash(f"บันทึกและล็อกผลแล้ว {locked} คู่", "success")
        else:
            flash(f"บันทึกคะแนนแล้ว {updated} คู่", "success")

        return redirect(url_for("multi_event_score_center", events=",".join(map(str, selected_event_ids)), round=selected_round, status=status_filter, field=field_query, event_rounds=event_rounds_param))

    matches = []
    if selected_event_ids:
        # ดึงเป็นกลุ่มต่ออีเว้นท์ เพื่อให้คีย์หลายอีเว้นท์แบบละเอียดและไม่กระโดดข้ามรายการ
        event_lookup = {e.id: e for e in selected_events}
        for eid in selected_event_ids:
            event = event_lookup.get(eid)
            if not event:
                continue
            round_no = event_round_map.get(eid, selected_round)
            event_matches = (
                Match.query
                .filter_by(event_id=eid, round=round_no)
                .order_by(Match.field.asc(), Match.id.asc())
                .all()
            )
            matches.extend(event_matches)

    rows = []
    for m in matches:
        score_ready = (m.team1_score is not None and (m.team2_id is None or m.team2_score is not None))
        is_draw = bool(m.team2_id is not None and m.team1_score is not None and m.team2_score is not None and m.team1_score == m.team2_score)
        row_status = "locked" if m.is_locked else ("draw" if is_draw else ("scored" if score_ready else "waiting"))
        if status_filter == "waiting" and row_status != "waiting":
            continue
        if status_filter == "scored" and row_status not in {"scored", "locked", "draw"}:
            continue
        if status_filter == "draw" and not is_draw:
            continue
        if field_query and field_query not in str(m.field or ""):
            continue
        rows.append({
            "match": m,
            "event": m.event,
            "team1": m.team1.name if m.team1 else "-",
            "team2": m.team2.name if m.team2 else "BYE",
            "status": row_status,
            "is_draw": is_draw,
            "score_ready": score_ready,
        })

    summary = {
        "events": len(selected_events),
        "matches": len(rows),
        "waiting": sum(1 for r in rows if r["status"] == "waiting"),
        "scored": sum(1 for r in rows if r["status"] in {"scored", "locked", "draw"}),
        "locked": sum(1 for r in rows if r["status"] == "locked"),
        "draw": sum(1 for r in rows if r["is_draw"]),
    }

    return render_template(
        "multi_score_center.html",
        all_events=all_events,
        selected_events=selected_events,
        selected_event_ids=selected_event_ids,
        selected_round=selected_round,
        event_round_map=event_round_map,
        event_rounds_param=event_rounds_param,
        status_filter=status_filter,
        field_query=field_query,
        rows=rows,
        summary=summary,
        today=today,
    )
# ---------- end beta multi-event score center ----------



@app.route("/event/<int:event_id>/lock", methods=["POST"])
@login_required
@roles_required('admin')  # เพิ่มบรรทัดนี้
def lock_round(event_id):
   
    max_round = db.session.query(db.func.max(Match.round)).filter_by(event_id=event_id).scalar()
    if max_round:
        matches = Match.query.filter_by(event_id=event_id, round=max_round).all()
        for m in matches:
            if m.team2_id is not None and (m.team1_score is None or m.team2_score is None):
                flash("กรุณากรอกคะแนนให้ครบก่อนล็อกผล")
                return redirect(url_for("event_detail", event_id=event_id))
        for m in matches:
            m.is_locked = True
        db.session.commit()
    return redirect(url_for("event_detail", event_id=event_id))


@app.route("/event/<int:event_id>/delete")
@login_required
@roles_required('admin')  # เพิ่มบรรทัดนี้
def delete_event(event_id):
    
    event = Event.query.get(event_id)
    if event is None:
        flash("ไม่พบรายการแข่งขันนี้", "warning")
        return redirect(url_for("index"))
    db.session.delete(event)
    db.session.commit()
    flash("ลบรายการแข่งขันเรียบร้อยแล้ว", "success")
    return redirect(url_for("index"))

@app.route("/event/<int:event_id>/clear", methods=["POST"])
@login_required
@roles_required('admin')
def clear_teams_route(event_id):
    # เช็คว่ามีแมตช์จับคู่ในอีเวนท์นี้หรือยัง
    existing_match = Match.query.filter_by(event_id=event_id).first()
    if existing_match:
        flash("ไม่สามารถลบทีมทั้งหมดได้ เนื่องจากมีการจับคู่แมตช์แล้ว", "danger")
        return redirect(url_for("event_detail", event_id=event_id))

    # ถ้ายังไม่มีแมตช์จึงลบทีมทั้งหมด
    Team.query.filter_by(event_id=event_id).delete()
    db.session.commit()
    flash("ลบรายชื่อทีมทั้งหมดเรียบร้อยแล้ว", "success")
    return redirect(url_for("event_detail", event_id=event_id))



def normalize_score_ends(raw_ends):
    """ตรวจและคำนวณคะแนนแบบ end-by-end
    raw_ends: list ของ {team: 1/2, points: 1-6}
    return (clean_ends, team1_total, team2_total)
    """
    if raw_ends in (None, ''):
        raw_ends = []
    if isinstance(raw_ends, str):
        raw_ends = json.loads(raw_ends or '[]')
    if not isinstance(raw_ends, list):
        raise ValueError('รูปแบบรายการ ends ไม่ถูกต้อง')

    clean = []
    team1_total = 0
    team2_total = 0
    for idx, item in enumerate(raw_ends, start=1):
        if not isinstance(item, dict):
            raise ValueError('รูปแบบ end ไม่ถูกต้อง')
        team = int(item.get('team'))
        points = int(item.get('points'))
        if team not in (1, 2):
            raise ValueError('ต้องเลือกทีมที่ได้คะแนน')
        if points < 1 or points > 6:
            raise ValueError('คะแนนต่อ end ต้องอยู่ระหว่าง 1-6')
        if team == 1:
            if team1_total + points > 13:
                raise ValueError(f'End ที่ {idx} ทำให้คะแนนทีม 1 เกิน 13')
            team1_total += points
        else:
            if team2_total + points > 13:
                raise ValueError(f'End ที่ {idx} ทำให้คะแนนทีม 2 เกิน 13')
            team2_total += points
        clean.append({
            'end': idx,
            'team': team,
            'points': points,
            'team1_total': team1_total,
            'team2_total': team2_total,
        })
    return clean, team1_total, team2_total


@app.route('/event/<int:event_id>/match/<int:match_id>/scorecard', methods=['GET'])
@login_required
@roles_required('user', 'admin', 'superadmin')
def online_scorecard(event_id, match_id):
    event = Event.query.get_or_404(event_id)
    match = Match.query.get_or_404(match_id)
    if match.event_id != event.id:
        flash("คู่แข่งขันไม่ตรงกับรายการนี้", "danger")
        return redirect(url_for("event_detail", event_id=event.id))

    ensure_match_scorecard_token(match)
    db.session.commit()

    teams = {team.id: team.name for team in Team.query.filter_by(event_id=event.id).all()}
    try:
        score_ends, _, _ = normalize_score_ends(match.score_ends or '[]')
    except Exception:
        score_ends = []
    return render_template(
        'online_scorecard.html',
        event=event,
        match=match,
        teams=teams,
        score_ends=score_ends,
        is_public_scorecard=False,
        scorecard_public_url=url_for('public_online_scorecard', token=match.scorecard_token, _external=True),
    )


@app.route('/scorecard/<token>', methods=['GET'])
def public_online_scorecard(token):
    match = Match.query.filter_by(scorecard_token=token).first_or_404()
    event = Event.query.get_or_404(match.event_id)
    teams = {team.id: team.name for team in Team.query.filter_by(event_id=event.id).all()}
    try:
        score_ends, _, _ = normalize_score_ends(match.score_ends or '[]')
    except Exception:
        score_ends = []
    return render_template(
        'online_scorecard.html',
        event=event,
        match=match,
        teams=teams,
        score_ends=score_ends,
        is_public_scorecard=True,
        scorecard_public_url=url_for('public_online_scorecard', token=match.scorecard_token, _external=True),
    )


@app.route('/scorecard/<token>/qr.png')
def public_scorecard_qr(token):
    match = Match.query.filter_by(scorecard_token=token).first_or_404()
    url = url_for('public_online_scorecard', token=match.scorecard_token, _external=True)
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png', max_age=0)


def save_online_scorecard_payload(match, data, submitted_user=None):
    if match.is_locked:
        return {'ok': False, 'message': 'คู่นี้ล็อกผลแล้ว'}, 400

    # ถ้าเคยกดสิ้นสุดแล้ว ไม่ให้ autosave กลับมาแก้คะแนนรอยืนยัน ต้องให้ admin ยกเลิกก่อน
    if match.pending_is_submitted:
        return {'ok': False, 'message': 'คู่นี้สิ้นสุดการแข่งขันแล้ว หากต้องแก้ไขให้ admin ยกเลิกคะแนนรอยืนยันก่อน'}, 400

    try:
        raw_ends = data.get('score_ends')
        if raw_ends is not None:
            score_ends, score1, score2 = normalize_score_ends(raw_ends)
        else:
            score1 = int(data.get('team1_score'))
            score2 = int(data.get('team2_score')) if match.team2_id else 0
            score_ends = []
    except (TypeError, ValueError, json.JSONDecodeError) as exc:
        return {'ok': False, 'message': str(exc) or 'กรุณากรอกคะแนนให้ถูกต้อง'}, 400

    if not match.team2_id:
        score2 = 0
    if score1 < 0 or score2 < 0:
        return {'ok': False, 'message': 'คะแนนต้องไม่น้อยกว่า 0'}, 400
    if score1 > 13 or score2 > 13:
        return {'ok': False, 'message': 'คะแนนต้องไม่เกิน 13'}, 400

    match.pending_team1_score = score1
    match.pending_team2_score = score2
    match.score_ends = json.dumps(score_ends, ensure_ascii=False)
    match.pending_submitted_by_id = submitted_user.id if submitted_user and submitted_user.is_authenticated else None
    match.pending_submitted_at = None
    db.session.commit()

    payload = {
        'match_id': match.id,
        'pending_team_a_score': match.pending_team1_score,
        'pending_team_b_score': match.pending_team2_score,
        'submitted_by': submitted_user.username if submitted_user and submitted_user.is_authenticated else 'QR Scorecard',
        'submitted_at': None,
        'pending_is_submitted': False,
        'has_pending': True,
        'score_ends': json.loads(match.score_ends or '[]'),
    }
    socketio.emit('pending_score_updated', payload)
    return {'ok': True, **payload}, 200


@app.route('/event/<int:event_id>/match/<int:match_id>/scorecard/autosave', methods=['POST'])
@login_required
@roles_required('user', 'admin', 'superadmin')
def autosave_online_scorecard(event_id, match_id):
    match = Match.query.get_or_404(match_id)
    if match.event_id != event_id:
        return jsonify({'ok': False, 'message': 'คู่แข่งขันไม่ตรงกับรายการนี้'}), 400
    payload, status = save_online_scorecard_payload(match, request.get_json(silent=True) or {}, current_user)
    return jsonify(payload), status


@app.route('/scorecard/<token>/autosave', methods=['POST'])
def public_autosave_online_scorecard(token):
    match = Match.query.filter_by(scorecard_token=token).first_or_404()
    payload, status = save_online_scorecard_payload(match, request.get_json(silent=True) or {}, None)
    return jsonify(payload), status



def finish_online_scorecard_payload(match, form_data, submitted_user=None):
    if match.is_locked:
        return "คู่นี้ล็อกผลแล้ว", "warning", False

    try:
        raw_ends = form_data.get('score_ends') or '[]'
        score_ends, score1, score2 = normalize_score_ends(raw_ends)
        # ถ้าเลือกโหมดกรอกคะแนนรวมเลย จะไม่มีประวัติ End
        if not score_ends:
            score1 = int(form_data.get('team1_score', ''))
            score2 = int(form_data.get('team2_score', '')) if match.team2_id else 0
    except (ValueError, json.JSONDecodeError) as exc:
        return str(exc) or "กรุณากรอกคะแนนให้ถูกต้อง", "danger", False

    sig1 = (form_data.get('team1_signature') or '').strip()
    sig2 = (form_data.get('team2_signature') or '').strip()

    if score1 < 0 or score2 < 0:
        return "คะแนนต้องไม่น้อยกว่า 0", "danger", False
    if score1 > 13 or score2 > 13:
        return "คะแนนต้องไม่เกิน 13", "danger", False
    # อนุญาตให้ส่งผลได้แม้คะแนนยังไม่ถึง 13
    # ใช้สำหรับกรณีจบเกมก่อน, แข่งตามเวลา, ผู้จัดตัดสินให้จบ, หรือกรอกผลย้อนหลัง
    if not sig1 or not sig2:
        return "ต้องมีลายเซ็นนักกีฬาทั้งสองทีมก่อนสิ้นสุดการแข่งขัน", "danger", False

    match.pending_team1_score = score1
    match.pending_team2_score = score2
    match.score_ends = json.dumps(score_ends, ensure_ascii=False)
    match.team1_signature = sig1
    match.team2_signature = sig2
    match.scorer_signature = None
    match.pending_is_submitted = True
    match.pending_submitted_by_id = submitted_user.id if submitted_user and submitted_user.is_authenticated else None
    match.pending_submitted_at = datetime.utcnow()
    db.session.commit()

    socketio.emit('pending_score_updated', {
        'match_id': match.id,
        'pending_team_a_score': match.pending_team1_score,
        'pending_team_b_score': match.pending_team2_score,
        'submitted_by': submitted_user.username if submitted_user and submitted_user.is_authenticated else 'QR Scorecard',
        'submitted_at': match.pending_submitted_at.isoformat(),
        'pending_is_submitted': True,
        'has_pending': True,
    })

    return "สิ้นสุดการแข่งขันแล้ว คะแนนจะขึ้นให้ admin/superadmin ยืนยันก่อนนำไปใช้งานจริง", "success", True


@app.route('/event/<int:event_id>/match/<int:match_id>/scorecard/finish', methods=['POST'])
@login_required
@roles_required('user', 'admin', 'superadmin')
def finish_online_scorecard(event_id, match_id):
    match = Match.query.get_or_404(match_id)
    if match.event_id != event_id:
        flash("คู่แข่งขันไม่ตรงกับรายการนี้", "danger")
        return redirect(url_for("event_detail", event_id=event_id))

    message, category, ok = finish_online_scorecard_payload(match, request.form, current_user)
    flash(message, category)
    return redirect(url_for('online_scorecard', event_id=event_id, match_id=match.id))


@app.route('/scorecard/<token>/finish', methods=['POST'])
def public_finish_online_scorecard(token):
    match = Match.query.filter_by(scorecard_token=token).first_or_404()
    message, category, ok = finish_online_scorecard_payload(match, request.form, None)
    flash(message, category)
    return redirect(url_for('public_online_scorecard', token=token))


    try:
        raw_ends = request.form.get('score_ends') or '[]'
        score_ends, score1, score2 = normalize_score_ends(raw_ends)
        # เผื่อกรอกคะแนนแบบรวมโดยไม่ใช้ ends ยังให้ระบบเดิมทำงานได้
        if not score_ends:
            score1 = int(request.form.get('team1_score', ''))
            score2 = int(request.form.get('team2_score', '')) if match.team2_id else 0
    except (ValueError, json.JSONDecodeError) as exc:
        flash(str(exc) or "กรุณากรอกคะแนนให้ถูกต้อง", "danger")
        return redirect(url_for('online_scorecard', event_id=event_id, match_id=match.id))

    sig1 = (request.form.get('team1_signature') or '').strip()
    sig2 = (request.form.get('team2_signature') or '').strip()
    sig3 = (request.form.get('scorer_signature') or '').strip()

    if score1 < 0 or score2 < 0:
        flash("คะแนนต้องไม่น้อยกว่า 0", "danger")
        return redirect(url_for('online_scorecard', event_id=event_id, match_id=match.id))

    if not sig1 or not sig2 or not sig3:
        flash("ต้องมีลายเซ็นนักกีฬาทั้งสองทีม และลายเซ็นผู้กรอก ก่อนสิ้นสุดการแข่งขัน", "danger")
        return redirect(url_for('online_scorecard', event_id=event_id, match_id=match.id))

    match.pending_team1_score = score1
    match.pending_team2_score = score2
    match.score_ends = json.dumps(score_ends, ensure_ascii=False)
    match.team1_signature = sig1
    match.team2_signature = sig2
    match.scorer_signature = sig3
    match.pending_is_submitted = True
    match.pending_submitted_by_id = current_user.id
    match.pending_submitted_at = datetime.utcnow()
    db.session.commit()

    socketio.emit('pending_score_updated', {
        'match_id': match.id,
        'pending_team_a_score': match.pending_team1_score,
        'pending_team_b_score': match.pending_team2_score,
        'submitted_by': current_user.username,
        'submitted_at': match.pending_submitted_at.isoformat(),
        'pending_is_submitted': True,
        'has_pending': True,
    })

    flash("สิ้นสุดการแข่งขันแล้ว คะแนนจะขึ้นให้ admin/superadmin ยืนยันก่อนนำไปใช้งานจริง", "success")
    return redirect(url_for('online_scorecard', event_id=event_id, match_id=match.id))


@app.route('/event/<int:event_id>/match/<int:match_id>/approve-pending-score', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def approve_pending_score(event_id, match_id):
    match = Match.query.get_or_404(match_id)
    if match.event_id != event_id:
        flash("คู่แข่งขันไม่ตรงกับรายการนี้", "danger")
        return redirect(url_for("event_detail", event_id=event_id))

    if (not match.pending_is_submitted) or match.pending_team1_score is None or (match.team2_id is not None and match.pending_team2_score is None):
        flash("ยังไม่มีคะแนนที่สิ้นสุดการแข่งขันและรอยืนยัน", "warning")
        return redirect(url_for("round_matches", event_id=event_id, round=match.round))

    match.team1_score = match.pending_team1_score
    match.team2_score = match.pending_team2_score if match.team2_id is not None else 0
    match.pending_team1_score = None
    match.pending_team2_score = None
    match.pending_is_submitted = False
    match.pending_submitted_by_id = None
    match.pending_submitted_at = None
    # เก็บลายเซ็นไว้หลัง admin ยืนยัน เพื่อให้เปิดหน้าสกอร์การ์ดย้อนกลับมาตรวจสอบได้
    match.is_locked = True
    db.session.commit()

    socketio.emit('score_updated', {
        'match_id': match.id,
        'team_a_score': match.team1_score,
        'team_b_score': match.team2_score,
        'pending_team_a_score': None,
        'pending_team_b_score': None,
        'has_pending': False,
        'pending_is_submitted': False,
        'locked': True,
        'timestamp': datetime.utcnow().isoformat()
    })

    flash("ยืนยันคะแนนออนไลน์และล็อกผลเรียบร้อยแล้ว", "success")
    return redirect(url_for("round_matches", event_id=event_id, round=match.round))


@app.route('/event/<int:event_id>/match/<int:match_id>/reject-pending-score', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def reject_pending_score(event_id, match_id):
    match = Match.query.get_or_404(match_id)
    if match.event_id != event_id:
        flash("คู่แข่งขันไม่ตรงกับรายการนี้", "danger")
        return redirect(url_for("event_detail", event_id=event_id))

    match.pending_team1_score = None
    match.pending_team2_score = None
    match.pending_is_submitted = False
    match.pending_submitted_by_id = None
    match.pending_submitted_at = None
    match.team1_signature = None
    match.team2_signature = None
    match.scorer_signature = None
    match.score_ends = None
    db.session.commit()

    socketio.emit('pending_score_updated', {
        'match_id': match.id,
        'pending_team_a_score': None,
        'pending_team_b_score': None,
        'submitted_by': None,
        'submitted_at': None,
        'pending_is_submitted': False
    })

    flash("ยกเลิกคะแนนที่รอยืนยันแล้ว", "success")
    return redirect(url_for("round_matches", event_id=event_id, round=match.round))


@app.route('/event/<int:event_id>/round/<int:round>', methods=['GET', 'POST'])
def round_matches(event_id, round):
    event = Event.query.get(event_id)
    if event is None:
        flash("ไม่พบรายการแข่งขันนี้", "warning")
        return redirect(url_for("index"))
    
    # ✅ แปลง logo_filename เป็น list
    try:
        event.logo_list = json.loads(event.logo_filename) if event.logo_filename else []
    except Exception:
        event.logo_list = []

    matches = Match.query.filter_by(event_id=event_id, round=round).order_by(Match.field.asc(), Match.id.asc()).all()
    teams = {team.id: team.name for team in Team.query.filter_by(event_id=event_id).all()}

    auto_assign_field = event.auto_field_enabled
     # ดึงรายการรอบทั้งหมดเพื่อทำปุ่มเปลี่ยนรอบ
    rounds = db.session.query(Match.round).filter_by(event_id=event_id).distinct().order_by(Match.round).all()
    rounds = [r[0] for r in rounds]

    def generate_field_numbers(event, count):
        prefix = event.field_prefix or ''
        start = event.field_start or 1
        max_field = event.field_max or 16
        exclude_str = event.field_exclude or ''
        exclude = set(x.strip() for x in exclude_str.split(',') if x.strip())

        fields = []
        num = start
        while len(fields) < count and num < start + max_field * 10:
            field_name = f"{prefix}{num}"
            if field_name not in exclude:
                fields.append(field_name)
            num += 1
        return fields

    # กำหนดค่าพวกนี้ก่อน render เสมอ
    standings = calculate_standings(event_id)
    total_rounds = event.rounds if event.rounds else db.session.query(db.func.max(Match.round)).filter(Match.event_id == event_id).scalar() or 1
    auto_fields = generate_field_numbers(event, len(matches)) if auto_assign_field else []
    all_current_round_locked = bool(matches) and all(m.is_locked for m in matches)
    is_last_configured_round = bool(total_rounds) and round >= total_rounds

    if request.method == "POST":
        action = request.form.get("action")

        # ** แก้ไขตรงนี้: toggle auto_assign_field โดยใช้ชื่อปุ่มเฉพาะ (เช่น toggle_auto_assign)**
        # วิธีที่ถูกต้อง: ตรวจสอบว่าปุ่ม toggle_auto_assign ถูกกดหรือไม่ก่อนอัปเดตสถานะ
        if "toggle_auto_assign" in request.form:
            event.auto_field_enabled = not event.auto_field_enabled  # สลับสถานะ
            auto_assign_field = event.auto_field_enabled

            # อัปเดตค่าที่รับจาก form ด้วย
            try:
                event.field_start = int(request.form.get("field_start", event.field_start or 1))
            except ValueError:
                event.field_start = 1
            event.field_prefix = request.form.get("field_prefix", event.field_prefix or "")
            event.field_exclude = request.form.get("field_exclude", event.field_exclude or "")

            db.session.commit()

            # คำนวณเลขสนามถ้าเปิด auto_assign_field
            auto_fields = generate_field_numbers(event, len(matches)) if auto_assign_field else []

            # กำหนดเลขสนามอัตโนมัติหากเปิดใช้งาน
            if auto_assign_field:
                for i, match in enumerate(matches):
                    match.field = auto_fields[i] if i < len(auto_fields) else None
            else:
                # ถ้าปิด auto ให้ล้างเลขสนามออก
                for match in matches:
                    match.field = None

            db.session.commit()
            _emit_round_live_changed(event_id, round)
            flash(f"{'เปิด' if auto_assign_field else 'ปิด'} การกำหนดเลขสนามอัตโนมัติเรียบร้อยแล้ว", "success")
            return redirect(url_for("round_matches", event_id=event_id, round=round))

        # อ่านค่าจาก form ที่เหลือ (เมื่อไม่ใช่ toggle)
        # อ่านสถานะ checkbox auto_assign_field จริงๆ จากชื่อในฟอร์ม ไม่ใช่ "auto_assign_field" in request.form
        # **แก้ไข**: ดึงค่า checkbox ให้ถูกต้อง
        auto_assign_field = True if request.form.get("auto_assign_field") == "on" else False
        event.auto_field_enabled = auto_assign_field

        # อัปเดต config ค่าอื่นๆ
        try:
            event.field_start = int(request.form.get("field_start", event.field_start or 1))
        except ValueError:
            event.field_start = 1
        event.field_prefix = request.form.get("field_prefix", event.field_prefix or "")
        event.field_exclude = request.form.get("field_exclude", event.field_exclude or "")

        db.session.commit()  # commit ก่อน เพื่อใช้ config ล่าสุด

        auto_fields = generate_field_numbers(event, len(matches)) if auto_assign_field else []

        if action == "save_fields":
            # กรณี auto assign
            
            if auto_assign_field:
                for i, match in enumerate(matches):
                    match.field = auto_fields[i] if i < len(auto_fields) else None
            else:
                # กรณีกรอกเลขสนามแมนนวล (ใน form)
                for match in matches:
                    field_key = f"field_{match.id}"
                    field_value = request.form.get(field_key)
                    print(f"{field_key} = {field_value}")  # debug
                    if field_value:
                        match.field = field_value.strip()
                    else:
                        match.field = None  # ถ้าไม่กรอก ให้ล้างเลขสนาม

            db.session.commit()
            _emit_round_live_changed(event_id, round)
            flash("บันทึกเลขสนามเรียบร้อยแล้ว", "success")
            return redirect(url_for("round_matches", event_id=event_id, round=round))

        # 🔴 บันทึกคะแนน + ล็อกผล
        elif action == "lock_scores":
            draw_matches = []

            for match in matches:
                score1 = request.form.get(f"score_{match.id}_1")
                score2 = request.form.get(f"score_{match.id}_2")

                if score1 is not None:
                    try:
                        match.team1_score = int(score1)
                    except ValueError:
                        flash(f"คะแนนทีม {teams.get(match.team1_id, '')} ต้องเป็นตัวเลข", "danger")
                        return redirect(url_for("round_matches", event_id=event_id, round=round))

                # รายชื่อที่เพิ่มภายหลังจะไม่มีคู่แข่ง ให้คีย์เฉพาะคะแนนฝั่งนี้ และเก็บฝั่งคู่แข่งเป็น 0
                if match.team2_id is None and match.is_manual == False:
                    if score1 is None or str(score1).strip() == "":
                        flash(f"กรุณากรอกคะแนนของ {teams.get(match.team1_id, '')}", "danger")
                        return redirect(url_for("round_matches", event_id=event_id, round=round))
                    match.team2_score = 0
                elif match.team2_id is not None and score2 is not None:
                    try:
                        match.team2_score = int(score2)
                    except ValueError:
                        flash(f"คะแนนทีม {teams.get(match.team2_id, '')} ต้องเป็นตัวเลข", "danger")
                        return redirect(url_for("round_matches", event_id=event_id, round=round))

                if match.team1_score is None:
                    flash(f"กรุณากรอกคะแนนของ {teams.get(match.team1_id, '')}", "danger")
                    return redirect(url_for("round_matches", event_id=event_id, round=round))
                if match.team2_id is not None and match.team2_score is None:
                    flash(f"กรุณากรอกคะแนนของ {teams.get(match.team2_id, '')}", "danger")
                    return redirect(url_for("round_matches", event_id=event_id, round=round))

                if match.team2_id is not None and match.team1_score == match.team2_score:
                    field_label = match.field if match.field else '-'
                    team1_name = teams.get(match.team1_id, '-')
                    team2_name = teams.get(match.team2_id, '-')
                    draw_matches.append(f"สนาม {field_label}: {team1_name} {match.team1_score}-{match.team2_score} {team2_name}")

                match.is_locked = True

            db.session.commit()
            _emit_round_live_changed(event_id, round, force_reload=True, reason='lock_scores')
            if draw_matches:
                flash("บันทึกและล็อกผลเรียบร้อยแล้ว แต่มีคู่เสมอ {} คู่\n{}".format(len(draw_matches), "\n".join(draw_matches)), "draw-warning")
            else:
                flash("บันทึกคะแนนและล็อกผลเรียบร้อยแล้ว", "success")
            return redirect(url_for("round_matches", event_id=event_id, round=round))

        # กรณี POST ที่ไม่ได้กด save_fields หรือ lock_scores (เช่น แค่ติ๊ก checkbox)
        standings = calculate_standings(event_id)
        total_rounds = event.rounds if event.rounds else db.session.query(db.func.max(Match.round)).filter(Match.event_id == event_id).scalar() or 1
        auto_fields = generate_field_numbers(event, len(matches)) if auto_assign_field else []

    live_version, _ = _round_live_state(event_id, round)
    return render_template(
        "round_matches.html",
        event=event,
        matches=matches,
        teams=teams,
        round=round,
        standings=standings,
        total_rounds=total_rounds,
        auto_assign_field=auto_assign_field,
        auto_fields=auto_fields,
        error_message=None,
        selected_round=round,  # ✅ ส่งตัวแปรนี้ไปให้ HTML
        all_current_round_locked=all_current_round_locked,
        is_last_configured_round=is_last_configured_round,
        live_version=live_version,
    )


@app.route('/event/<int:event_id>/swith/<int:round_no>', methods=['GET', 'POST'])
@login_required
@roles_required('superadmin')
def swith_round_positions(event_id, round_no):
    """หน้าลับสำหรับ superadmin สลับตำแหน่งทีมหลังประกบคู่แล้ว

    วิธีใช้: เปลี่ยน URL จาก /event/<id>/round/<round> เป็น /event/<id>/swith/<round>
    เลือกทีม 2 ทีม แล้วกดสลับตำแหน่ง ระบบจะสลับ team1_id/team2_id ในรอบนั้นแบบเงียบ ๆ
    โดยไม่ flash ข้อความแจ้งเตือนใด ๆ ตามที่ต้องการ
    """
    event = Event.query.get_or_404(event_id)

    try:
        event.logo_list = json.loads(event.logo_filename) if event.logo_filename else []
    except Exception:
        event.logo_list = []

    matches = Match.query.filter_by(event_id=event_id, round=round_no).order_by(Match.field.asc(), Match.id.asc()).all()
    teams = {team.id: team.name for team in Team.query.filter_by(event_id=event_id).all()}

    if request.method == 'POST':
        selected_ids = []
        for raw_id in request.form.getlist('team_ids'):
            try:
                team_id = int(raw_id)
            except (TypeError, ValueError):
                continue
            if team_id not in selected_ids:
                selected_ids.append(team_id)

        # ไม่ต้องแจ้งข้อความ: ถ้าเลือกไม่ครบ 2 ทีม ให้กลับหน้าเดิมเงียบ ๆ
        if len(selected_ids) == 2:
            selected_set = set(selected_ids)
            found_slots = {}

            for match in matches:
                if match.team1_id in selected_set and match.team1_id not in found_slots:
                    found_slots[match.team1_id] = (match, 'team1_id')
                if match.team2_id in selected_set and match.team2_id not in found_slots:
                    found_slots[match.team2_id] = (match, 'team2_id')

            # ต้องเจอทีมที่เลือกทั้ง 2 ทีมในรอบนี้ จึงสลับ
            if all(team_id in found_slots for team_id in selected_ids):
                first_team_id, second_team_id = selected_ids
                first_match, first_side = found_slots[first_team_id]
                second_match, second_side = found_slots[second_team_id]

                setattr(first_match, first_side, second_team_id)
                setattr(second_match, second_side, first_team_id)
                db.session.commit()
                _emit_round_live_changed(event_id, round_no)

        return redirect(url_for('swith_round_positions', event_id=event_id, round_no=round_no))

    return render_template(
        'swith.html',
        event=event,
        matches=matches,
        teams=teams,
        round_no=round_no,
    )


@app.route('/event/<int:event_id>/late_pairing', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def late_pairing_route(event_id):
    event = Event.query.get_or_404(event_id)
    target_round = request.form.get('round', type=int) or get_latest_round_for_event(event_id)
    mode = request.form.get('mode', 'auto')
    late = _late_entry_matches(event_id, target_round)
    if not late:
        flash('ไม่มีทีมเพิ่มภายหลังที่รอประกบคู่', 'info')
        return redirect(url_for('round_matches', event_id=event_id, round=target_round or 1))

    if mode == 'manual':
        team1_id = request.form.get('team1_id', type=int)
        team2_id = request.form.get('team2_id', type=int)
        if not team1_id or not team2_id or team1_id == team2_id:
            flash('กรุณาเลือกทีมใหม่ 2 ทีมที่ไม่ซ้ำกัน', 'warning')
            return redirect(url_for('round_matches', event_id=event_id, round=target_round))
        by_team = {m.team1_id: m for m in late}
        m1 = by_team.get(team1_id)
        m2 = by_team.get(team2_id)
        if not m1 or not m2:
            flash('เลือกได้เฉพาะทีมที่เพิ่มใหม่และยังไม่ได้ประกบคู่เท่านั้น', 'warning')
            return redirect(url_for('round_matches', event_id=event_id, round=target_round))
        m1.team2_id = m2.team1_id
        m1.team1_score = None
        m1.team2_score = None
        m1.is_manual = True
        if event.auto_field_enabled and not m1.field:
            m1.field = _next_available_field(event, target_round)
        db.session.delete(m2)
        db.session.commit()
        _emit_round_live_changed(event_id, target_round)
        flash('ประกบคู่ทีมเพิ่มภายหลังแบบแมนนวลเรียบร้อยแล้ว โดยไม่แก้คู่เดิม', 'success')
    else:
        paired, remaining = auto_pair_late_entries(event_id, target_round)
        db.session.commit()
        _emit_round_live_changed(event_id, target_round)
        flash(f'ระบบประกบทีมเพิ่มภายหลังให้อัตโนมัติ {paired} คู่ เหลือรอคู่ {remaining} รายการ โดยไม่แก้คู่เดิม', 'success')

    return redirect(url_for('round_matches', event_id=event_id, round=target_round))


@app.route('/api/abbr_lookup')
@login_required
def abbr_lookup_api():
    """ค้นคำอ่าน/ความหมายคำย่อจากอินเทอร์เน็ตแบบ optional ถ้าค้นไม่ได้ให้ frontend ถามผู้ใช้ต่อ"""
    q = (request.args.get('q') or '').strip()
    lang = (request.args.get('lang') or 'th').strip().lower()
    wiki_lang = 'th' if lang.startswith('th') else 'en'
    if not q or len(q) > 60:
        return jsonify({'ok': False, 'reading': '', 'source': ''})
    local = {
        'มข': 'มหาวิทยาลัยขอนแก่น', 'รร': 'โรงเรียน', 'ร.ร.': 'โรงเรียน',
        'อบจ': 'องค์การบริหารส่วนจังหวัด', 'ทม': 'เทศบาลเมือง', 'ทต': 'เทศบาลตำบล',
        'อบต': 'องค์การบริหารส่วนตำบล', 'กกท': 'การกีฬาแห่งประเทศไทย',
        'TH': 'Thailand', 'KKU': 'Khon Kaen University'
    }
    if q in local:
        return jsonify({'ok': True, 'reading': local[q], 'source': 'local'})
    try:
        url = f'https://{wiki_lang}.wikipedia.org/w/api.php?action=query&list=search&srsearch={quote(q)}&format=json&srlimit=1'
        req = Request(url, headers={'User-Agent': 'TournoiPetanqueVoice/1.0'})
        with urlopen(req, timeout=4) as res:
            data = json.loads(res.read().decode('utf-8'))
        hits = data.get('query', {}).get('search', [])
        if hits:
            title = hits[0].get('title') or ''
            if title:
                return jsonify({'ok': True, 'reading': title, 'source': 'wikipedia'})
    except Exception:
        pass
    return jsonify({'ok': False, 'reading': '', 'source': ''})


@app.route("/users")
def user_list():
    users = User.query.all()
    return render_template("user_list.html", users=users)

@app.route("/users/add", methods=["POST"])
def add_user():
    username = request.form["username"]
    password = request.form["password"]
    role = request.form["role"]
    start_time = request.form.get("start_time")
    end_time = request.form.get("end_time")

    if User.query.filter_by(username=username).first():
        flash("Username already exists")
        return redirect(url_for("user.user_list"))

    user = User(
        username=username,
        password_hash=generate_password_hash(password),
        role=role,
        start_time=datetime.fromisoformat(start_time) if start_time else None,
        end_time=datetime.fromisoformat(end_time) if end_time else None
    )
    db.session.add(user)
    db.session.commit()
    flash("User added successfully")
    return redirect(url_for("user.user_list"))

#--------------------------------------------------------------------------------------------

@app.route('/admin/users/add', methods=['GET', 'POST'])
@login_required
@roles_required('superadmin')
def admin_add_user():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        role = request.form['role']
        duration = request.form.get('duration', '1m')  # default 1 เดือน

        if role not in ['user', 'admin', 'superadmin']:
            flash("สิทธิ์ผู้ใช้ไม่ถูกต้อง", "danger")
            return redirect(url_for('admin_add_user'))

        if User.query.filter_by(username=username).first():
            flash("ชื่อผู้ใช้นี้มีอยู่แล้ว", "danger")
            return redirect(url_for('admin_add_user'))

        start_time = datetime.utcnow()
        end_time = calculate_end_time(duration)

        new_user = User(username=username, role=role, start_time=start_time, end_time=end_time)
        new_user.set_password(password)
        db.session.add(new_user)
        db.session.commit()
        flash("เพิ่มผู้ใช้เรียบร้อยแล้ว", "success")
        return redirect(url_for('admin_users'))  # ไปหน้ารายชื่อผู้ใช้

    return render_template('admin_add_user.html')
  


@app.route('/admin/users')
@login_required
@roles_required('admin', 'superadmin')
def admin_users():
    users = User.query.all()
    for u in users:
        if u.end_time:
            u.end_time_str = u.end_time.strftime('%Y-%m-%d %H:%M:%S')
        else:
            u.end_time_str = "ถาวร"
    return render_template('admin_users.html', users=users, now=datetime.utcnow())


@app.route('/admin/users/delete/<int:user_id>', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def delete_user(user_id):
    user = User.query.get_or_404(user_id)

    if user.id == current_user.id:
        flash("ไม่สามารถลบบัญชีที่กำลังใช้งานอยู่ได้", "danger")
        return redirect(url_for('admin_users'))

    # ห้าม admin ลบ superadmin
    if current_user.role == 'admin' and user.role == 'superadmin':
        flash("คุณไม่มีสิทธิ์ลบ superadmin", "danger")
        return redirect(url_for('admin_users'))

    # ไม่ลบอีเว้นท์ของ user คนนั้น: โอน owner ของอีเว้นท์มาให้ผู้ที่กดลบแทน
    Event.query.filter_by(creator_id=user.id).update({Event.creator_id: current_user.id})

    # เคลียร์ผู้ส่งคะแนน pending เพื่อไม่ให้ FK ค้างกับ user ที่ถูกลบ
    Match.query.filter_by(pending_submitted_by_id=user.id).update({
        Match.pending_submitted_by_id: None,
        Match.pending_is_submitted: False,
        Match.team1_signature: None,
        Match.team2_signature: None,
        Match.scorer_signature: None,
    })

    db.session.delete(user)
    db.session.commit()
    flash("ลบผู้ใช้เรียบร้อยแล้ว และเก็บอีเว้นท์เดิมไว้โดยโอนเจ้าของให้บัญชีของคุณ", "success")
    return redirect(url_for('admin_users'))


@app.route('/admin/users/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@roles_required('superadmin')
def edit_user(user_id):
    user = User.query.get_or_404(user_id)

    if request.method == 'POST':
        new_username = (request.form.get('username') or '').strip()
        new_password = request.form.get('password') or ''
        new_role = request.form.get('role')
        duration = request.form.get('duration')  # '' = ไม่เปลี่ยนเวลาใช้งาน

        if not new_username:
            flash('กรุณากรอกชื่อผู้ใช้', 'danger')
            return redirect(url_for('edit_user', user_id=user.id))

        duplicate = User.query.filter(User.username == new_username, User.id != user.id).first()
        if duplicate:
            flash('ชื่อผู้ใช้นี้มีอยู่แล้ว', 'danger')
            return redirect(url_for('edit_user', user_id=user.id))

        if new_role not in ['user', 'admin', 'superadmin']:
            flash('สิทธิ์ผู้ใช้ไม่ถูกต้อง', 'danger')
            return redirect(url_for('edit_user', user_id=user.id))

        user.username = new_username
        user.role = new_role

        if new_password.strip():
            user.set_password(new_password.strip())

        # สำคัญ: ถ้าไม่ได้เลือกเปลี่ยนระยะเวลา ให้คงวันหมดอายุเดิมไว้
        # ป้องกันปัญหาแก้ข้อมูลนิดเดียวแล้วกลายเป็น "ถาวร"
        if duration:
            user.start_time = datetime.utcnow()
            user.end_time = calculate_end_time(duration)

        db.session.commit()
        flash('แก้ไขผู้ใช้สำเร็จ', 'success')
        return redirect(url_for('admin_users'))

    return render_template('admin_edit_user.html', user=user)

@app.route('/match/event/<int:event_id>/match_pairs')
def match_pairs(event_id):
    
    event = Event.query.get_or_404(event_id)
    selected_round = request.args.get('round', None, type=int)

    
     # แปลง logo_filename เป็น list หรือ [] ถ้า error
    try:
        event.logo_list = json.loads(event.logo_filename) if event.logo_filename else []
    except Exception:
        event.logo_list = []   
    
    # ถ้าไม่ระบุรอบ ให้เลือกรอบล่าสุดที่มีใน event นั้น
    if not selected_round:
        latest_match = db.session.query(Match.round)\
            .filter_by(event_id=event_id)\
            .order_by(Match.round.desc())\
            .first()
        selected_round = latest_match.round if latest_match else None

    # ดึงแมตช์ของ event และรอบที่เลือก (หรือรอบล่าสุด)
    query = Match.query.filter_by(event_id=event_id)
    if selected_round:
        query = query.filter_by(round=selected_round)
    matches = query.order_by(Match.field.asc()).all()

    # เตรียมชื่อทีม
    teams = {team.id: team.name for team in Team.query.filter_by(event_id=event_id).all()}
    
    matches_by_round = defaultdict(list)
    for m in matches:
        matches_by_round[m.round].append(m)

    live_version, _ = _round_live_state(event_id, selected_round)
    return render_template(
        'match_pairs.html',
        event=event,
        matches=matches,
        matches_by_round=matches_by_round,
        teams=teams,
        selected_round=selected_round,
        live_version=live_version
    )

#--------------------------------------------------------------------------------------

# -----------------------------------------------------------------------------
# Finish round / next competition helpers

def _as_int(value, default=0):
    try:
        return int(value)
    except Exception:
        return default


def _team_rank_map(event_id):
    """Return standings rows plus rank map from the current Swiss standings."""
    rows = calculate_standings(event_id)
    rank_map = {}
    for idx, row in enumerate(rows, start=1):
        rank_map[_as_int(row.get("team_id"))] = idx
    return rows, rank_map


def _selected_rows_from_standings(event_id, selected_ids):
    standings_rows, rank_map = _team_rank_map(event_id)
    by_id = {_as_int(row.get("team_id")): row for row in standings_rows}
    selected = []
    for tid in selected_ids:
        row = by_id.get(tid)
        if row:
            row = dict(row)
            row["rank"] = rank_map.get(tid, 999999)
            selected.append(row)
    selected.sort(key=lambda r: r.get("rank", 999999))
    return selected


def _make_next_event_from_selected(source_event, selected_rows, stage_name, rounds=3):
    """Create a fresh Swiss event from selected teams. Existing scores/matches are not copied."""
    owner_id = getattr(current_user, "id", None) or source_event.creator_id
    new_event = Event(
        name=f"{source_event.name} - {stage_name}" if stage_name else f"{source_event.name} - รอบใหม่",
        location=source_event.location,
        category=source_event.category,
        sex=source_event.sex,
        age_group=source_event.age_group,
        rounds=rounds or 3,
        current_round=1,
        date=source_event.date,
        auto_field_enabled=source_event.auto_field_enabled,
        field_prefix=source_event.field_prefix,
        field_start=source_event.field_start,
        field_max=source_event.field_max,
        field_exclude=source_event.field_exclude,
        logo_filename=source_event.logo_filename,
        creator_id=owner_id,
    )
    db.session.add(new_event)
    db.session.flush()
    for row in selected_rows:
        db.session.add(Team(name=row.get("team_name", ""), event_id=new_event.id))
    db.session.commit()
    return new_event


def _next_power_of_two(n):
    if n <= 1:
        return 1
    return 1 << (n - 1).bit_length()


def _seed_spread_order(n):
    """Fishbone/bracket seed order: 1 and 2 stay on opposite sides."""
    if n <= 1:
        return [1]
    size = _next_power_of_two(n)
    order = [1, 2]
    while len(order) < size:
        next_size = len(order) * 2
        order = [seed for old in order for seed in (old, next_size + 1 - old)]
    return [seed for seed in order if seed <= n]


def _seed_spread_order_with_byes(n):
    """ก้างปลาแบบแยก seed 1 กับ seed 2 ให้อยู่คนละครึ่งตาราง"""
    size = _next_power_of_two(n)
    fixed = {
        1: [1],
        2: [1, 2],
        4: [1, 4, 2, 3],
        8: [1, 8, 4, 5, 3, 6, 2, 7],
        16: [1, 16, 8, 9, 4, 13, 5, 12, 3, 14, 6, 11, 7, 10, 2, 15],
        32: [1, 32, 16, 17, 8, 25, 9, 24, 4, 29, 13, 20, 5, 28, 12, 21, 3, 30, 14, 19, 6, 27, 11, 22, 7, 26, 10, 23, 2, 31, 15, 18],
    }
    if size in fixed:
        return fixed[size]
    seeds = list(range(1, size + 1))
    return seeds[:1] + seeds[2:] + [2]


def _knockout_pairs(selected_rows):
    """Create fishbone knockout pairs: 1 is far from 2 and can meet 2 only in final."""
    teams = list(selected_rows)
    by_seed = {idx: row for idx, row in enumerate(teams, start=1)}
    slots = _seed_spread_order_with_byes(len(teams))
    pairs = []
    for i in range(0, len(slots), 2):
        a_seed = slots[i]
        b_seed = slots[i + 1] if i + 1 < len(slots) else None
        team_a = by_seed.get(a_seed)
        team_b = by_seed.get(b_seed) if b_seed else None
        if team_a:
            team_a = dict(team_a)
            team_a["bracket_seed"] = a_seed
        if team_b:
            team_b = dict(team_b)
            team_b["bracket_seed"] = b_seed
        pairs.append({
            "match_no": len(pairs) + 1,
            "slot_a_seed": a_seed,
            "slot_b_seed": b_seed,
            "team1": team_a,
            "team2": team_b,
        })
    return pairs


def _calculate_group_sizes_3_4(team_count):
    """Return group sizes for double knockout: each group must contain 3 or 4 real teams.

    Normal/random mode keeps 4-team groups first and 3-team groups last.
    If no valid 3-4 split exists (for example 1, 2, or 5 teams), return an empty list.
    """
    if team_count < 3:
        return []
    # Prefer as many 4-team groups as possible. The number of 3-team groups (b)
    # is the smallest b where team_count - 3b is divisible by 4.
    for b in range(0, 4):
        remaining = team_count - (3 * b)
        if remaining < 0:
            continue
        if remaining % 4 == 0:
            a = remaining // 4
            sizes = ([4] * a) + ([3] * b)
            if sum(sizes) == team_count and sizes:
                return sizes
    return []


def _double_seed_flat_order(team_count):
    """Seed order for normal double groups without the special X formula.

    It starts from real seed pairs: best vs worst, second vs second worst, etc.,
    then orders those pairs so seed 1 stays far from seed 2 in the display.
    """
    pair_count = team_count // 2
    raw_pairs = {idx: (idx, team_count + 1 - idx) for idx in range(1, pair_count + 1)}
    flat = []
    for pair_no in _pair_order_for_bracket(pair_count):
        pair = raw_pairs.get(pair_no)
        if pair:
            flat.extend(pair)
    if team_count % 2 == 1:
        middle_seed = pair_count + 1
        if middle_seed not in flat:
            flat.append(middle_seed)
    return flat


def _slots_from_double_chunk(chunk, mode='normal'):
    """Create 4 slots for one double group.

    Seed-special mode uses Team, X, Team, Team.
    Normal/random mode uses Team, Team, Team, X for 3-team groups.
    Four-team groups are Team, Team, Team, Team.
    """
    chunk = list(chunk)
    if len(chunk) >= 4:
        return chunk[:4]
    if len(chunk) == 3:
        return [chunk[0], None, chunk[1], chunk[2]] if mode == 'seed_special' else [chunk[0], chunk[1], chunk[2], None]
    return chunk + [None] * (4 - len(chunk))


def _arrange_random_double_group(rows):
    """Random/normal double group: place random teams one by one.

    If the group has 3 real teams, X/empty slot is always at slot 4, and
    groups with X must be placed after complete 4-team groups by the caller.
    """
    return _slots_from_double_chunk(list(rows), mode='normal')


def _double_top_seed_order(k):
    # กฎที่ตกลง: 12 ทีม(k=4), 24 ทีม(k=8), 48 ทีม(k=16)
    fixed = {
        1: [1],
        2: [1, 2],
        4: [1, 4, 3, 2],
        8: [1, 8, 5, 4, 3, 6, 7, 2],
        16: [1, 16, 9, 8, 5, 12, 13, 4, 3, 14, 11, 6, 7, 10, 15, 2],
    }
    if k in fixed:
        return fixed[k]
    return _seed_spread_order(k)


def _double_knockout_groups(selected_rows):
    """
    Double knockout setup:
    - If selected team count is divisible by 3, use the agreed 3-layer seed rule.
    - If not divisible by 3, randomize into 3-4 team groups using the petanque_tournament style.
    """
    rows = [dict(row) for row in selected_rows]
    total = len(rows)
    if total <= 0:
        return [], "ยังไม่ได้เลือกทีม"

    by_rank = {idx: row for idx, row in enumerate(rows, start=1)}
    groups = []

    if total >= 3 and total % 3 == 0:
        k = total // 3
        for group_no, top_seed in enumerate(_double_top_seed_order(k), start=1):
            middle_seed = (2 * k) + 1 - top_seed
            lower_seed = (3 * k) + 1 - top_seed
            groups.append({
                "group_no": group_no,
                "mode": "seeded",
                "slots": [
                    {"seed": top_seed, "team": by_rank.get(top_seed), "is_bye": False},
                    {"seed": "-", "team": None, "is_bye": True},
                    {"seed": lower_seed, "team": by_rank.get(lower_seed), "is_bye": False},
                    {"seed": middle_seed, "team": by_rank.get(middle_seed), "is_bye": False},
                ],
            })
        return groups, "เข้ากฎหาร 3 ลงตัว: ใช้กฎ seed บนเจอ X และกลุ่มกลางประกบกลุ่มล่าง"

    random_rows = list(rows)
    random.shuffle(random_rows)
    sizes = _calculate_group_sizes_3_4(total)
    cursor = 0
    for group_no, size in enumerate(sizes, start=1):
        chunk = random_rows[cursor:cursor + size]
        cursor += size
        arranged = _arrange_random_double_group(chunk)
        while len(arranged) < 4:
            arranged.append(None)
        slots = []
        for item in arranged[:4]:
            if item is None:
                slots.append({"seed": "-", "team": None, "is_bye": True})
            else:
                slots.append({"seed": item.get("rank", ""), "team": item, "is_bye": False})
        groups.append({"group_no": group_no, "mode": "random", "slots": slots})
    return groups, "จำนวนทีมไม่หาร 3 ลงตัว: ระบบสุ่มจัดสาย 3–4 ทีมตาม logic ตัวอย่าง"

def _round_robin_pairs(selected_rows):
    pairs = []
    match_no = 1
    for i in range(len(selected_rows)):
        for j in range(i + 1, len(selected_rows)):
            pairs.append({"match_no": match_no, "team1": selected_rows[i], "team2": selected_rows[j]})
            match_no += 1
    return pairs


def _round_robin_round_groups(selected_rows):
    """สร้างพบกันหมดแบบเป็นรอบจริง: รอบหนึ่งมีหลายคู่, ทีมคี่พัก BYE แต่ไม่สร้างคู่ BYE"""
    teams = [_row_to_seed_payload(row, idx) for idx, row in enumerate((selected_rows or []), start=1)]
    if len(teams) < 2:
        return []
    if len(teams) % 2 == 1:
        teams.append(None)
    n = len(teams)
    rounds = []
    arr = teams[:]
    for round_no in range(1, n):
        pairs = []
        for i in range(n // 2):
            a = arr[i]
            b = arr[n - 1 - i]
            if a is None or b is None:
                continue
            pairs.append([a, b])
        rounds.append(pairs)
        arr = [arr[0]] + [arr[-1]] + arr[1:-1]
    return rounds


def _create_round_robin_competition(source_event, selected_rows, title, direct_rows=None):
    """สร้างระบบ Round Robin บน playoff tables เดิม เพื่อให้ใช้หน้าคีย์/รายงานร่วมกันได้"""
    ensure_playoff_tables()
    config = {
        'direct_qualified': [
            {
                'rank': idx,
                'team_id': row.get('team_id'),
                'team_name': row.get('team_name'),
                'seed': row.get('seed') or row.get('rank') or idx,
            }
            for idx, row in enumerate((direct_rows or []), start=1)
        ],
        'round_robin': True,
    }
    res = db.session.execute(text("""
        INSERT INTO playoff_competitions (source_event_id, title, competition_type, pairing_method, config_json)
        VALUES (:source_event_id, :title, :competition_type, :pairing_method, :config_json)
    """), {
        'source_event_id': source_event.id if source_event else None,
        'title': title or 'Round Robin',
        'competition_type': 'round_robin',
        'pairing_method': 'round_robin',
        'config_json': json.dumps(config, ensure_ascii=False),
    })
    db.session.flush()
    if db.engine.dialect.name == 'postgresql':
        playoff_id = db.session.execute(text("SELECT currval(pg_get_serial_sequence('playoff_competitions','id')) AS id")).mappings().first()['id']
    else:
        playoff_id = res.lastrowid

    rr_rounds = _round_robin_round_groups(selected_rows)
    if not rr_rounds:
        raise ValueError('Round Robin ต้องมีทีมอย่างน้อย 2 ทีม')
    for round_no, pairs in enumerate(rr_rounds, start=1):
        _create_playoff_round(playoff_id, round_no, f'Round Robin ครั้งที่ {round_no}', 'round_robin', pairs)
    db.session.commit()
    return playoff_id


def _active_playoffs_for_event(event_id):
    # รายการเพลย์ออฟที่สร้างจาก Event นี้ เพื่อให้กลับเข้าไปทำงานต่อได้ถ้าเผลอออกจากหน้า
    try:
        ensure_playoff_tables()
        rows = db.session.execute(text("""
            SELECT pc.id, pc.title, pc.competition_type, pc.pairing_method,
                   COUNT(pr.id) AS round_count,
                   MAX(pr.round_no) AS latest_round
            FROM playoff_competitions pc
            LEFT JOIN playoff_rounds pr ON pr.playoff_id = pc.id
            WHERE pc.source_event_id = :event_id
            GROUP BY pc.id, pc.title, pc.competition_type, pc.pairing_method
            ORDER BY pc.id DESC
        """), {'event_id': event_id}).mappings().all()
        return [dict(r) for r in rows]
    except Exception:
        return []

# Swiss System Logic & Standings Display - เริ่มปรับปรุงที่นี่

# ฟังก์ชันคำนวณ Buchholz (BHN) และ Fine Buchholz (fBHN) ที่นำมาจาก standings.py
# (ถ้า standings.py คำนวณให้แล้ว ไม่ต้องทำซ้ำตรงนี้)
# จากที่ดู standings.py ของคุณ calculate_standings จะคืนค่า bhn และ fbhn มาให้แล้ว

@app.route('/event/<int:event_id>/standings')
@login_required
def event_standings(event_id):
    current_event = Event.query.get_or_404(event_id)
    standings_data = calculate_standings(event_id)

    # โหมดคัดหลัง Swiss แบบ 3 กลุ่ม:
    # 1) อันดับดีที่สุด เข้ารอบอัตโนมัติ
    # 2) อันดับถัดมา นำไปจัดแข่งต่อ
    # 3) ลำดับล่าง ตกรอบ
    total_rows = len(standings_data)
    direct_advance_count = request.args.get('direct_advance_count', type=int)
    continue_count = request.args.get('continue_count', type=int)

    # รองรับ URL เดิมที่เคยใช้ num_qualified_teams = จำนวนที่เลือกไปแข่งต่อ/เข้ารอบรวม
    legacy_qualified = request.args.get('num_qualified_teams', type=int, default=min(8, total_rows or 8))
    if direct_advance_count is None:
        direct_advance_count = 0
    if continue_count is None:
        continue_count = legacy_qualified

    direct_advance_count = max(0, min(total_rows, int(direct_advance_count or 0)))
    continue_count = max(0, min(total_rows - direct_advance_count, int(continue_count or 0)))
    num_qualified_teams = direct_advance_count + continue_count

    direct_rows = standings_data[:direct_advance_count]
    qualified_rows = standings_data[direct_advance_count:num_qualified_teams]
    eliminated_rows = standings_data[num_qualified_teams:]
    ab_a_team_count = request.args.get('ab_a_team_count', type=int)
    if ab_a_team_count is None:
        ab_a_team_count = (continue_count // 2) if continue_count > 1 else min(1, continue_count)
    ab_a_team_count = max(0, min(len(qualified_rows), int(ab_a_team_count or 0)))

    return render_template(
        'event_standings.html',
        event=current_event,
        standings=standings_data,
        direct_rows=direct_rows,
        qualified_rows=qualified_rows,
        eliminated_rows=eliminated_rows,
        direct_advance_count=direct_advance_count,
        continue_count=continue_count,
        num_qualified_teams=num_qualified_teams,
        ab_a_team_count=ab_a_team_count,
        active_playoffs=_active_playoffs_for_event(event_id),
    )

# เส้นทางใหม่สำหรับดาวน์โหลด Excel

@app.route('/event/<int:event_id>/download_standings_excel')
@login_required
def download_standings_excel(event_id):
    current_event = Event.query.get_or_404(event_id)
    
    # รับค่าแบ่งกลุ่มหลัง Swiss: direct / แข่งต่อ / ตกรอบ
    legacy_qualified = request.args.get('num_qualified_teams', type=int, default=8)
    direct_advance_count = request.args.get('direct_advance_count', type=int, default=0)
    continue_count = request.args.get('continue_count', type=int, default=legacy_qualified)
    ab_a_team_count = request.args.get('ab_a_team_count', type=int)
    competition_type = request.args.get('competition_type', '')
    show_ab_status = competition_type in ('ab_ladder', 'regional64_ladder')
    
    # ดึงข้อมูล Round ที่อาจจะส่งมา (ถ้ามี)
    # ถ้า 'round' ไม่ได้ถูกส่งมา, อาจจะถือว่าเป็นการดาวน์โหลดตารางคะแนนรวม
    requested_round = request.args.get('round', type=int) 
    
    standings_data = calculate_standings(event_id) # เรียกใช้ฟังก์ชันเดิม
    total_rows = len(standings_data)
    direct_advance_count = max(0, min(total_rows, int(direct_advance_count or 0)))
    continue_count = max(0, min(total_rows - direct_advance_count, int(continue_count or 0)))
    num_qualified_teams = direct_advance_count + continue_count
    if ab_a_team_count is None:
        ab_a_team_count = (continue_count // 2) if continue_count > 1 else min(1, continue_count)
    ab_a_team_count = max(0, min(continue_count, int(ab_a_team_count or 0)))
    
    # ถ้ามีการระบุ round ให้กรองข้อมูล (สำหรับตอนนี้ ยังคง pass ไว้)
    if requested_round:
        # คุณอาจจะต้องปรับ calculate_standings ในไฟล์ standings.py ให้รับ round_number
        # หรือกรองข้อมูล standings_data ที่นี่ หากข้อมูลที่ได้กลับมามีรายละเอียดรอบอยู่
        pass 

    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Standings"

    # --- Header Information (ชื่อกิจกรรม, ประเภท, รุ่น) ---
    ws.merge_cells('A1:G1') # รวมเซลล์สำหรับหัวข้อหลัก
    ws['A1'] = f"รายงานผลจัดลำดับ: {current_event.name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('A2:G2') # รวมเซลล์สำหรับประเภทและรุ่น
    # ใช้ฟิลด์ category และ age_group ตามที่เราได้ตกลงกันไว้ใน models.py
    event_type_str = current_event.category if hasattr(current_event, 'category') and current_event.category else 'ไม่ระบุประเภท'
    event_category_str = current_event.age_group if hasattr(current_event, 'age_group') and current_event.age_group else 'ไม่ระบุรุ่น'
    event_sex_str = current_event.sex if hasattr(current_event, 'sex') and current_event.sex else 'ไม่ระบุเพศ'
    ws['A2'] = f"ประเภท: {event_type_str}{event_sex_str} | รุ่น: {event_category_str}"
    ws['A2'].font = Font(size=12)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # ถ้ามีข้อมูลรอบที่เฉพาะเจาะจงที่ดาวน์โหลด
    if requested_round:
        ws.merge_cells('A3:G3') # รวมเซลล์สำหรับรอบที่
        ws['A3'] = f"ครั้งการแข่งขัน: รอบที่ {requested_round}"
        ws['A3'].font = Font(bold=True)
        ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
        header_row_start = 5 # ถ้ามีรอบ จะเริ่มหัวตารางที่แถว 5
    else:
        header_row_start = 4 # ไม่มีรอบ จะเริ่มหัวตารางที่แถว 4
    
    ws.row_dimensions[1].height = 25 # กำหนดความสูงของแถว
    ws.row_dimensions[2].height = 20
    if requested_round: ws.row_dimensions[3].height = 20

    # --- Table Headers (หัวตาราง) ---
    # ******************************** สำคัญ: ลบ "Point For" และ "Point Against" ออกจาก headers ********************************
    headers = ["อันดับ", "ชื่อทีม", "คะแนน", "BHN", "fBHN", "Point Diff", "หมายเหตุ"] 
    col_start = 1 # เริ่มต้นคอลัมน์ A (Excel)
    for col_num, header in enumerate(headers, col_start):
        cell = ws.cell(row=header_row_start, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF") # กำหนดฟอนต์ตัวหนาและสีข้อความเป็นสีขาว
        cell.alignment = Alignment(horizontal='center', vertical='center') # จัดกึ่งกลาง
        cell.fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid") # สีพื้นหลังหัวตารางเป็นสีเขียวเข้ม
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')) # ขอบเซลล์

    # --- Data Rows (ข้อมูลทีมและการกำหนดสีแถว) ---
    row_num = header_row_start + 1 # เริ่มต้นแถวข้อมูลหลังจากหัวตาราง
    for i, team in enumerate(standings_data, 1): # i คืออันดับทีม
        # กำหนดสไตล์เริ่มต้น (ไม่มีพื้นหลัง)
        row_fill = None
        row_font = Font() # ฟอนต์ปกติ
        row_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        if i <= direct_advance_count:
            row_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
            status_text = f"ตัวแทนคนที่ {i}"
        elif i <= num_qualified_teams:
            ab_order = i - direct_advance_count
            ab_zone = 'A' if ab_order <= ab_a_team_count else 'B'
            if show_ab_status:
                if ab_zone == 'A':
                    row_fill = PatternFill(start_color="D7ECFF", end_color="D7ECFF", fill_type="solid")
                else:
                    row_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                status_text = f"เข้ารอบ กลุ่ม {ab_zone}"
            else:
                row_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
                status_text = "เข้ารอบ"
        else:
            row_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            status_text = "ตกรอบ"
        row_font = Font(color="000000")


        # ใส่ข้อมูลและกำหนดสไตล์ให้กับแต่ละเซลล์ในแถวนี้
        ws.cell(row=row_num, column=1, value=i).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=team['team_name'])
        ws.cell(row=row_num, column=3, value=team['score']).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=4, value=team['buchholz']).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=5, value=team['final_buchholz']).alignment = Alignment(horizontal='center')
        # ******************************** สำคัญ: ลบ team['point_for'] และ team['point_against'] ออก ********************************
        # ******************************** และปรับ column ของ 'point_diff' เป็น 6 ********************************
        ws.cell(row=row_num, column=6, value=f"{team['point_for']}:{team['point_against']}").alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=7, value=status_text).alignment = Alignment(horizontal='center')

        # วนลูปเพื่อกำหนดสีพื้นหลัง, ฟอนต์ และขอบให้กับทุกเซลล์ในแถวปัจจุบัน
        # len(headers) จะเป็น 6 (ตามจำนวน headers ใหม่) ทำให้ลูปวนแค่ 6 คอลัมน์
        for col_idx in range(1, len(headers) + 1): 
            cell = ws.cell(row=row_num, column=col_idx)
            if row_fill:
                cell.fill = row_fill
            if row_font:
                cell.font = row_font
            cell.border = row_border # ใช้ border ที่กำหนดไว้
        
        row_num += 1

    # --- Auto-size Columns --- 
    for col_cells in ws.columns: # 'col_cells' ตอนนี้คือ tuple ของเซลล์ในคอลัมน์นั้นๆ แล้ว
        max_length = 0
        # เราจะวนลูปโดยตรงผ่าน tuple ของเซลล์ในคอลัมน์นั้นๆ
        for cell in col_cells:
            try:
                if cell.value is not None: # ตรวจสอบว่ามีค่าในเซลล์ก่อน
                    cell_value_str = str(cell.value)
                    if len(cell_value_str) > max_length:
                        max_length = len(cell_value_str)
            except:
                pass
        
        # ปรับความกว้างของคอลัมน์
        adjusted_width = (max_length + 2)
        if col_cells: # ตรวจสอบว่า col_cells ไม่ว่างเปล่าก่อนเข้าถึง [0] เพื่อหา column letter
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = adjusted_width

    wb.save(output)
    output.seek(0)

    # --- ตั้งชื่อไฟล์ Excel สำหรับดาวน์โหลด ---
    # ทำให้ชื่อไฟล์สะอาดสำหรับ URL
    event_name_safe = current_event.name.replace(" ", "_").replace("/", "-")
    # ใช้ค่าที่ได้จาก current_event.category และ current_event.age_group
    event_type_safe = event_type_str.replace(" ", "_").replace("/", "-")
    event_category_safe = event_category_str.replace(" ", "_").replace("/", "-")
    
    filename_parts = [
        event_name_safe,
        event_type_safe,
        event_category_safe
    ]
    if requested_round:
        filename_parts.append(f"Round_{requested_round}")

    filename = "_".join(part for part in filename_parts if part) + "_Standings.xlsx"

    return send_file(output, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

#--------------------------------------------------------------------------------------
# เส้นทางใหม่สำหรับนำข้อมูลทีมที่ได้ไปจัดทำตารางรอบน็อคเอาท์
@app.route('/event/<int:event_id>/create_bracket')
@login_required
def create_bracket(event_id):
    current_event = Event.query.get_or_404(event_id)
    standings_data = calculate_standings(event_id)

    num_qualified_teams = request.args.get('num_qualified_teams', type=int, default=8)
    if num_qualified_teams > len(standings_data):
        num_qualified_teams = len(standings_data)
    if num_qualified_teams < 0:
        num_qualified_teams = 0

    qualified_teams = standings_data[:num_qualified_teams]

    # ส่งข้อมูลทีมที่เข้ารอบไปยัง template ใหม่สำหรับสร้างตารางรอบน็อคเอาท์
    return render_template(
        'bracket_setup.html', # คุณจะต้องสร้างไฟล์นี้ขึ้นมา
        event=current_event,
        qualified_teams=qualified_teams
    )

# --- เพิ่ม Route สำหรับบันทึกการจับคู่ ---
@app.route('/event/<int:event_id>/save_bracket_pairings', methods=['POST'])
@login_required # ถ้าต้องการให้ต้อง Login ก่อน
def save_bracket_pairings(event_id):
    """
    ฟังก์ชันนี้จะรับข้อมูลการจับคู่รอบน็อคเอาท์จากฟอร์มใน bracket_setup.html
    และทำการบันทึกข้อมูลนั้นลงในฐานข้อมูล
    """
    try:
        # รับข้อมูลการจับคู่จากฟอร์ม
        # วิธีการรับข้อมูลขึ้นอยู่กับว่าคุณสร้างฟอร์มใน bracket_setup.html อย่างไร
        # ตัวอย่าง: หากคุณใช้ JavaScript ส่งข้อมูลเป็น JSON
        # pairings_data = request.get_json()

        # ตัวอย่าง: หากคุณใช้ input hidden field ชื่อ 'pairings_json' เก็บ JSON string
        pairings_json = request.form.get('pairings_data') # สมมติว่ามี input field ชื่อ 'pairings_data'
        if pairings_json:
            pairings_list = json.loads(pairings_json)
            
            # === ส่วนนี้คือที่คุณจะเขียนโค้ดเพื่อบันทึกข้อมูลลงฐานข้อมูล ===
            # วนลูปผ่าน pairings_list และสร้าง/อัปเดต Match objects
            # for pairing in pairings_list:
            #     team1_id = pairing['team1_id']
            #     team2_id = pairing['team2_id']
            #     round_num = pairing.get('round', 1) # กำหนดรอบเริ่มต้น
            #     
            #     # ตรวจสอบว่าคู่นี้มีอยู่แล้วหรือไม่ หรือสร้างใหม่
            #     # existing_match = Match.query.filter_by(event_id=event_id, round=round_num, team1_id=team1_id, team2_id=team2_id).first()
            #     # if not existing_match:
            #     #    new_match = Match(event_id=event_id, team1_id=team1_id, team2_id=team2_id, round=round_num, is_locked=False)
            #     #    db.session.add(new_match)
            # db.session.commit() # บันทึกการเปลี่ยนแปลงทั้งหมดในคราวเดียว
            # === จบส่วนบันทึกข้อมูล ===
            
            flash('บันทึกการจับคู่รอบน็อคเอาท์สำเร็จ!', 'success')
        else:
            flash('ไม่พบข้อมูลการจับคู่ที่จะบันทึก', 'warning')

        # หลังจากบันทึกเสร็จแล้ว Redirect ไปยังหน้าที่แสดง Bracket ที่สร้างเสร็จ
        # หรือกลับไปหน้า create_bracket อีกครั้ง
        return redirect(url_for('create_bracket', event_id=event_id))

    except Exception as e:
        # db.session.rollback() # หากมีการใช้ db.session.add/commit และเกิดข้อผิดพลาด
        flash(f'เกิดข้อผิดพลาดในการบันทึกการจับคู่: {str(e)}', 'danger')
        return redirect(url_for('create_bracket', event_id=event_id))




@app.route('/event/<int:event_id>/create-next-competition', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def create_next_competition(event_id):
    source_event = Event.query.get_or_404(event_id)

    selected_ids = []
    for raw in request.form.getlist('team_ids'):
        tid = _as_int(raw)
        if tid and tid not in selected_ids:
            selected_ids.append(tid)

    direct_ids = []
    for raw in request.form.getlist('direct_team_ids'):
        tid = _as_int(raw)
        if tid and tid not in direct_ids:
            direct_ids.append(tid)

    competition_type = (request.form.get('competition_type') or 'knockout').strip()
    next_stage_name = (request.form.get('next_stage_name') or '').strip() or 'รอบถัดไป'
    swiss_rounds = max(1, _as_int(request.form.get('swiss_rounds'), 3))
    pairing_method = (request.form.get('pairing_method') or 'seed').strip()
    if pairing_method not in {'seed', 'random', 'manual', 'bracket', 'national_qualifier'}:
        pairing_method = 'seed'
    add_bye = request.form.get('add_bye') == '1'

    selected_rows = _selected_rows_from_standings(event_id, selected_ids)
    direct_rows = _selected_rows_from_standings(event_id, direct_ids) if direct_ids else []

    # ไม่ติ๊ก: direct_rows = ตัวแทนคนที่ 1, 2 ... และไม่แข่ง playoff
    # ติ๊ก: รวม direct_rows เข้า playoff แล้วไม่ล็อกเป็นตัวแทนอัตโนมัติ
    include_direct_in_playoff = request.form.get('include_direct_in_playoff') == '1'
    if include_direct_in_playoff and direct_rows:
        merged_rows, seen = [], set()
        for row in list(direct_rows) + list(selected_rows):
            key = row.get('team_id') or row.get('team_name')
            if key in seen:
                continue
            seen.add(key)
            merged_rows.append(row)
        selected_rows = merged_rows
        direct_rows_for_report = []
    else:
        direct_rows_for_report = direct_rows

    if not selected_rows:
        flash('กรุณาเลือกทีมอย่างน้อย 1 ทีมก่อนสร้างรอบถัดไป', 'warning')
        return redirect(url_for('event_standings', event_id=event_id,
                                direct_advance_count=request.form.get('direct_advance_count', 0),
                                continue_count=request.form.get('continue_count', request.form.get('num_qualified_teams', 8))))

    payload = {
        'source_event_id': event_id,
        'competition_type': competition_type,
        'next_stage_name': next_stage_name,
        'team_ids': [row['team_id'] for row in selected_rows],
        'direct_team_ids': [row['team_id'] for row in direct_rows_for_report],
        'include_direct_in_playoff': include_direct_in_playoff,
        'pairing_method': pairing_method,
        'add_bye': add_bye,
        'created_at': datetime.utcnow().isoformat(),
    }
    session[f'next_competition_{event_id}'] = payload

    if competition_type == 'swiss':
        new_event = _make_next_event_from_selected(source_event, selected_rows, next_stage_name, rounds=swiss_rounds)
        flash('สร้างอีเว้นท์ Swiss ใหม่แล้ว ข้อมูลคะแนน/แมตช์เดิมไม่ถูกนำมาด้วย ให้เริ่มจับคู่รอบแรกใหม่', 'success')
        return redirect(url_for('event_detail', event_id=new_event.id))

    national_qualifier_mode = (competition_type == 'regional64_ladder' or pairing_method == 'national_qualifier')
    if competition_type in {'ab_ladder', 'regional64_ladder'} or national_qualifier_mode:
        try:
            if national_qualifier_mode:
                if len(selected_rows) not in (48, 64):
                    raise ValueError('ประกบคู่คัดตัวแทนทีมชาติ ต้องเลือกทีมเข้าเพลย์ออฟ 64 ทีม หรือ 48 ทีมเท่านั้น')
                a_team_count = 32
                if len(selected_rows) == 64:
                    advance_a = 8
                    advance_b = 12
                else:
                    advance_a = 8
                    advance_b = 10
                # ใช้ seed เป็นฐานในการวางรอบแรก แต่ล็อกโหมดพิเศษด้วย config_json
                playoff_id = _create_ab_ladder_competition(source_event, selected_rows, next_stage_name, 'national_qualifier',
                                                           a_team_count, advance_a, advance_b,
                                                           direct_rows=direct_rows_for_report,
                                                           special_mode='regional64_ladder')
            else:
                a_team_count = _as_int(request.form.get('ab_a_team_count'), max(1, len(selected_rows)//2))
                advance_a = _as_int(request.form.get('ab_advance_a'), 0)
                advance_b = _as_int(request.form.get('ab_advance_b'), 0)
                playoff_id = _create_ab_ladder_competition(source_event, selected_rows, next_stage_name, pairing_method,
                                                           a_team_count, advance_a, advance_b,
                                                           direct_rows=direct_rows_for_report)
        except ValueError as exc:
            flash(str(exc), 'warning')
            return redirect(url_for('event_standings', event_id=event_id,
                                    direct_advance_count=request.form.get('direct_advance_count', 0),
                                    continue_count=request.form.get('continue_count', 8)))
        flash('สร้างการแข่งขันแบบแบ่งกลุ่ม A/B ต่อจาก Swiss แล้ว: A แพ้ตกลง B, B แพ้ตกรอบ', 'success')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))

    if competition_type == 'round_robin':
        try:
            playoff_id = _create_round_robin_competition(source_event, selected_rows, next_stage_name, direct_rows=direct_rows_for_report)
        except ValueError as exc:
            flash(str(exc), 'warning')
            return redirect(url_for('event_standings', event_id=event_id))
        flash('สร้าง Round Robin พบกันหมดแล้ว สามารถคีย์คะแนนและสร้างรอบต่อไปได้', 'success')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))

    if competition_type in {'knockout', 'double_knockout'}:
        if pairing_method == 'manual':
            flash('เลือก MANUAL แล้ว กรุณาจิ้มทีมลงคู่/ลงสายเองก่อนสร้างรอบ', 'info')
            return redirect(url_for('playoff_manual_pairing_from_event', event_id=event_id))
        try:
            playoff_id = _create_playoff_competition(source_event, selected_rows, next_stage_name,
                                                     competition_type, pairing_method,
                                                     add_bye=add_bye,
                                                     direct_rows=direct_rows_for_report)
        except ValueError as exc:
            flash(str(exc), 'warning')
            return redirect(url_for('event_standings', event_id=event_id,
                                    direct_advance_count=request.form.get('direct_advance_count', 0),
                                    continue_count=request.form.get('continue_count', 8)))
        flash('สร้างระบบแข่งขันต่อแล้ว สามารถกรอกผลและสร้างรอบต่อไปได้จากหน้านี้', 'success')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))

    return redirect(url_for('knockout_setup', event_id=event_id))


def _get_next_competition_payload(event_id):
    payload = session.get(f'next_competition_{event_id}')
    if not payload:
        return None, []
    selected_ids = [_as_int(x) for x in payload.get('team_ids', [])]
    return payload, _selected_rows_from_standings(event_id, selected_ids)




def _manual_pairing_plan(selected_rows, competition_type):
    """Return a plan for the manual pairing page.

    knockout = one pair per group, 2 slots per pair.
    double_knockout = one group per 3-4 teams, 4 slots per group.
    """
    team_count = len(selected_rows)
    if competition_type == 'double_knockout':
        sizes = _calculate_group_sizes_3_4(team_count)
        if not sizes:
            raise ValueError('Double knockout แบบ MANUAL ต้องจัดทีมเป็นสายละ 3-4 ทีม และต้องมีอย่างน้อย 3 ทีม')
        return {
            'slot_count': 4,
            'groups': [{'group_no': idx + 1, 'required_count': size} for idx, size in enumerate(sizes)],
            'mode_label': 'Double knockout: เลือกทีมลงตามสาย / ลำดับ 1-4',
        }
    group_count = (team_count + 1) // 2
    return {
        'slot_count': 2,
        'groups': [{'group_no': idx + 1, 'required_count': 2} for idx in range(group_count)],
        'mode_label': 'Knockout: เลือกทีมลงคู่ ทีม A / ทีม B',
    }


def _manual_groups_from_request(selected_rows, competition_type):
    teams = [_row_to_seed_payload(row, idx) for idx, row in enumerate(selected_rows, start=1)]
    by_id = {str(t['team_id']): t for t in teams if t.get('team_id')}
    plan = _manual_pairing_plan(selected_rows, competition_type)
    used = []
    groups = []
    for group in plan['groups']:
        group_no = group['group_no']
        slots = []
        for slot_no in range(1, plan['slot_count'] + 1):
            raw = (request.form.get(f'team_{group_no}_{slot_no}') or '').strip()
            if not raw:
                slots.append(None)
                continue
            if raw not in by_id:
                raise ValueError('พบทีมที่ไม่มีอยู่ในรายการ กรุณาเลือกใหม่')
            if raw in used:
                raise ValueError('มีทีมถูกเลือกซ้ำ กรุณาเลือกทีมแต่ละทีมได้ครั้งเดียว')
            used.append(raw)
            slots.append(dict(by_id[raw]))
        groups.append(slots)

    all_ids = set(by_id.keys())
    used_ids = set(used)
    missing = [by_id[x]['team_name'] for x in all_ids - used_ids]
    if missing:
        raise ValueError('ยังมีทีมที่ยังไม่ได้ลงคู่/ลงสาย: ' + ', '.join(missing))
    if len(used_ids) < 2:
        raise ValueError('ต้องเลือกทีมอย่างน้อย 2 ทีม')
    if competition_type == 'double_knockout':
        for idx, slots in enumerate(groups, start=1):
            real_count = sum(1 for x in slots if x is not None)
            if real_count not in (3, 4):
                raise ValueError(f'สาย {idx} ต้องมีทีมจริง 3 หรือ 4 ทีม')
    return groups


def _create_playoff_competition_with_groups(source_event, title, competition_type, pairing_method, groups, direct_rows=None):
    config = {
        'direct_qualified': [
            {
                'rank': idx,
                'team_id': row.get('team_id'),
                'team_name': row.get('team_name'),
                'seed': row.get('seed') or row.get('rank') or idx,
            }
            for idx, row in enumerate((direct_rows or []), start=1)
        ]
    }
    res = db.session.execute(text("""
        INSERT INTO playoff_competitions (source_event_id, title, competition_type, pairing_method, config_json)
        VALUES (:source_event_id, :title, :competition_type, :pairing_method, :config_json)
    """), {
        'source_event_id': source_event.id if source_event else None,
        'title': title,
        'competition_type': competition_type,
        'pairing_method': pairing_method,
        'config_json': json.dumps(config, ensure_ascii=False),
    })
    db.session.flush()
    if db.engine.dialect.name == 'postgresql':
        playoff_id = db.session.execute(text("SELECT currval(pg_get_serial_sequence('playoff_competitions','id')) AS id")).mappings().first()['id']
    else:
        playoff_id = res.lastrowid
    _create_playoff_round(playoff_id, 1, title or 'รอบที่ 1', competition_type, groups)
    db.session.commit()
    return playoff_id


@app.route('/event/<int:event_id>/playoff/manual-pairing', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'superadmin')
def playoff_manual_pairing_from_event(event_id):
    source_event = Event.query.get_or_404(event_id)
    payload, selected_rows = _get_next_competition_payload(event_id)
    if not payload or not selected_rows:
        flash('กรุณาเลือกทีมจากหน้า Standing ก่อน', 'warning')
        return redirect(url_for('event_standings', event_id=event_id))
    competition_type = payload.get('competition_type') or 'knockout'
    if competition_type not in {'knockout', 'double_knockout'}:
        flash('MANUAL ใช้กับ Knockout หรือ Double knockout เท่านั้น', 'warning')
        return redirect(url_for('event_standings', event_id=event_id))
    try:
        plan = _manual_pairing_plan(selected_rows, competition_type)
    except ValueError as exc:
        flash(str(exc), 'warning')
        return redirect(url_for('event_standings', event_id=event_id))
    teams = [_row_to_seed_payload(row, idx) for idx, row in enumerate(selected_rows, start=1)]
    if request.method == 'POST':
        try:
            groups = _manual_groups_from_request(selected_rows, competition_type)
            playoff_id = _create_playoff_competition_with_groups(
                source_event,
                payload.get('next_stage_name') or 'รอบถัดไป',
                competition_type,
                'manual',
                groups,
                direct_rows=_selected_rows_from_standings(event_id, payload.get('direct_team_ids', [])),
            )
        except ValueError as exc:
            flash(str(exc), 'warning')
            return render_template('playoff_manual_pairing.html', mode='event', source_event=source_event, view=None, payload=payload, teams=teams, plan=plan)
        session.pop(f'next_competition_{event_id}', None)
        flash('สร้างรอบ MANUAL ตามที่แอดมินจิ้มคู่แล้ว', 'success')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    return render_template('playoff_manual_pairing.html', mode='event', source_event=source_event, view=None, payload=payload, teams=teams, plan=plan)

@app.route('/event/<int:event_id>/knockout/setup')
@login_required
def knockout_setup(event_id):
    event = Event.query.get_or_404(event_id)
    payload, selected_rows = _get_next_competition_payload(event_id)
    if not payload or not selected_rows:
        flash('กรุณาเลือกทีมจากหน้า Standing ก่อน', 'warning')
        return redirect(url_for('event_standings', event_id=event_id))
    return render_template(
        'next_knockout_setup.html',
        event=event,
        next_stage_name=payload.get('next_stage_name', 'รอบถัดไป'),
        selected_rows=selected_rows,
        pairs=_knockout_pairs(selected_rows),
    )


@app.route('/event/<int:event_id>/double-knockout/setup')
@login_required
def double_knockout_setup(event_id):
    event = Event.query.get_or_404(event_id)
    payload, selected_rows = _get_next_competition_payload(event_id)
    if not payload or not selected_rows:
        flash('กรุณาเลือกทีมจากหน้า Standing ก่อน', 'warning')
        return redirect(url_for('event_standings', event_id=event_id))
    try:
        raw_groups = _playoff_double_groups(
            selected_rows,
            payload.get('pairing_method', 'seed'),
            add_bye=bool(payload.get('add_bye'))
        )
        groups = []
        for group_no, slots in enumerate(raw_groups, start=1):
            groups.append({
                'group_no': group_no,
                'mode': payload.get('pairing_method', 'seed'),
                'slots': [
                    {'seed': '-', 'team': None, 'is_bye': True} if slot is None else {'seed': slot.get('seed'), 'team': slot, 'is_bye': False}
                    for slot in slots
                ]
            })
        draw_message = 'จัดสายดับเบิ้ลตามวิธีประกบคู่ที่เลือกไว้'
        error_message = None
    except ValueError as exc:
        groups = []
        draw_message = ''
        error_message = str(exc)
    return render_template(
        'next_double_knockout_setup.html',
        event=event,
        next_stage_name=payload.get('next_stage_name', 'รอบถัดไป'),
        selected_rows=selected_rows,
        groups=groups,
        draw_message=draw_message,
        error_message=error_message,
    )


@app.route('/event/<int:event_id>/round-robin/setup')
@login_required
def round_robin_setup(event_id):
    event = Event.query.get_or_404(event_id)
    payload, selected_rows = _get_next_competition_payload(event_id)
    if not payload or not selected_rows:
        flash('กรุณาเลือกทีมจากหน้า Standing ก่อน', 'warning')
        return redirect(url_for('event_standings', event_id=event_id))
    return render_template(
        'next_round_robin_setup.html',
        event=event,
        next_stage_name=payload.get('next_stage_name', 'รอบถัดไป'),
        selected_rows=selected_rows,
        pairs=_round_robin_pairs(selected_rows),
    )


@app.route('/event/<int:event_id>/clear-next-competition')
@login_required
@roles_required('admin', 'superadmin')
def clear_next_competition(event_id):
    session.pop(f'next_competition_{event_id}', None)
    flash('ล้างชุดทีมที่เลือกสำหรับรอบถัดไปแล้ว', 'info')
    return redirect(url_for('event_standings', event_id=event_id))


# ------------------------- Playoff / bracket engine after Standing -------------------------
def _row_to_seed_payload(row, seed=None):
    return {
        'team_id': _as_int(row.get('team_id')),
        'team_name': row.get('team_name') or row.get('name') or '',
        'seed': seed if seed is not None else _as_int(row.get('rank'), 0),
        'rank': _as_int(row.get('rank'), 0),
        # ป้ายทางเดินทีม: ใช้บอกว่าทีมนี้มาจากชนะ/แพ้สายใดในรอบก่อนหน้า
        'entry_label': row.get('entry_label') or row.get('national_label') or row.get('regional64_label'),
        'source_label': row.get('source_label'),
    }


def _rows_for_pairing(selected_rows, pairing_method):
    # ถ้า row มี seed อยู่แล้ว (เช่น A/B ที่ B ต้องเริ่มต่อจาก A) ให้รักษา seed เดิมไว้
    # ถ้าไม่มี seed ให้เริ่มนับ 1..N ตามลำดับที่ส่งเข้ามา
    rows = []
    for idx, row in enumerate(selected_rows, start=1):
        existing_seed = _as_int(row.get('seed'), 0) if isinstance(row, dict) else 0
        rows.append(_row_to_seed_payload(row, existing_seed or idx))

    if pairing_method == 'random':
        # สุ่มตำแหน่ง แต่ไม่รีเซ็ตเลข seed เพื่อให้เลขลำดับบนใบพิมพ์ยังต่อเนื่อง
        random.shuffle(rows)
    return rows


def _pair_order_for_bracket(pair_count):
    """Order pair numbers so pair 1 and pair 2 are far apart in the display."""
    fixed = {
        1: [1],
        2: [1, 2],
        3: [1, 3, 2],
        4: [1, 4, 3, 2],
        6: [1, 6, 5, 4, 3, 2],
        8: [1, 8, 5, 4, 3, 6, 7, 2],
        12: [1, 12, 8, 5, 4, 9, 3, 10, 6, 7, 11, 2],
        16: [1, 16, 9, 8, 5, 12, 13, 4, 3, 14, 11, 6, 7, 10, 15, 2],
    }
    if pair_count in fixed:
        return fixed[pair_count]
    order = [x for x in _seed_spread_order_with_byes(pair_count) if 1 <= x <= pair_count]
    if 2 in order:
        order = [x for x in order if x != 2] + [2]
    return order


def _normal_seed_pairs(rows):
    """Pair real teams only: best vs worst, second best vs second worst, then display fishbone."""
    n = len(rows)
    raw_pairs = []
    for i in range(n // 2):
        raw_pairs.append([rows[i], rows[n - 1 - i]])
    order = _pair_order_for_bracket(len(raw_pairs))
    by_no = {idx + 1: pair for idx, pair in enumerate(raw_pairs)}
    return [by_no[i] for i in order if i in by_no]


def _adjacent_bracket_pairs(rows):
    """Bracket mode: keep the existing bracket/slot order and pair adjacent teams.

    This is for playoff next rounds: do not reseed winners again.
    Example: winner of pair 1 meets winner of pair 2, pair 3 meets pair 4, etc.
    """
    rows = list(rows)
    pairs = []
    for i in range(0, len(rows), 2):
        pairs.append([rows[i], rows[i + 1] if i + 1 < len(rows) else None])
    return pairs


def _random_pairs(rows):
    """True random: shuffle and pair adjacent teams. No seed formula."""
    rows = list(rows)
    random.shuffle(rows)
    pairs = []
    for i in range(0, len(rows), 2):
        pairs.append([rows[i], rows[i + 1] if i + 1 < len(rows) else None])
    return pairs


def _playoff_knockout_groups(selected_rows, pairing_method='seed', add_bye=False):
    rows = _rows_for_pairing(selected_rows, pairing_method)
    if len(rows) % 2 == 1 and not add_bye:
        raise ValueError('จำนวนทีมเป็นเลขคี่ ถ้าไม่ต้องการเพิ่ม X/BYE กรุณาเลือกทีมให้เป็นจำนวนคู่ หรือเปิดตัวเลือกเพิ่ม X/BYE')

    if pairing_method == 'random':
        if add_bye and len(rows) % 2 == 1:
            rows.append(None)
        return _random_pairs(rows)

    if pairing_method == 'bracket':
        # โหมด bracket = คงลำดับสายเดิม แล้วจับคู่ที่ติดกัน
        # ใช้สำหรับรอบต่อไปหลังจากรอบ 8/16 จัดตาม seed แล้ว ไม่เอาผู้ชนะมา seed ใหม่
        if add_bye and len(rows) % 2 == 1:
            rows.append(None)
        return _adjacent_bracket_pairs(rows)

    if add_bye:
        # เพิ่ม X/BYE เฉพาะเมื่อผู้ใช้ติ๊กเท่านั้น
        by_seed = {idx: _row_to_seed_payload(row, idx) for idx, row in enumerate(selected_rows, start=1)}
        slots = _seed_spread_order_with_byes(len(selected_rows))
        groups = []
        for i in range(0, len(slots), 2):
            groups.append([by_seed.get(slots[i]), by_seed.get(slots[i + 1]) if i + 1 < len(slots) else None])
        return groups
    # ค่าเริ่มต้น seed: ไม่เพิ่ม X จับทีมจริง ดีสุดเจอท้ายสุด และวาง 1 ไกล 2
    return _normal_seed_pairs(rows)


def _playoff_double_groups(selected_rows, pairing_method='seed', add_bye=False):
    """Create double-knockout groups.

    Rules locked with the user:
    - Double knockout is NOT pair-by-pair like knockout. It is 3-4 teams per group.
    - Random = shuffle all teams, then place teams into groups one by one.
    - Normal/random 3-team groups use slots: Team, Team, Team, X and stay at the end.
    - Seed + special X uses the divisible-by-3 formula: Team, X, Team, Team.
    - Seed + no special X uses real teams in 3-4 team groups; no seed-special X.
    """
    n = len(selected_rows)
    if n < 3:
        raise ValueError('ดับเบิ้ลน็อคเอาท์ต้องมีอย่างน้อย 3 ทีมต่อการจัดสาย')

    # Seed + special X formula: K groups, each group = 3 real teams + X in slot 2.
    if pairing_method != 'random' and add_bye and n % 3 == 0:
        rows = [_row_to_seed_payload(row, idx) for idx, row in enumerate(selected_rows, start=1)]
        by_rank = {idx: row for idx, row in enumerate(rows, start=1)}
        k = n // 3
        groups = []
        for top_seed in _double_top_seed_order(k):
            middle_seed = (2 * k) + 1 - top_seed
            lower_seed = (3 * k) + 1 - top_seed
            groups.append([
                by_rank.get(top_seed),
                None,
                by_rank.get(lower_seed),
                by_rank.get(middle_seed),
            ])
        return groups

    sizes = _calculate_group_sizes_3_4(n)
    if not sizes:
        raise ValueError('จำนวนทีมนี้จัดดับเบิ้ลแบบสายละ 3–4 ทีมไม่ได้ กรุณาเลือกจำนวนทีมใหม่ หรือเปลี่ยนระบบแข่งขัน')

    if pairing_method == 'random':
        rows = [_row_to_seed_payload(row, idx) for idx, row in enumerate(selected_rows, start=1)]
        random.shuffle(rows)
    else:
        seed_rows = [_row_to_seed_payload(row, idx) for idx, row in enumerate(selected_rows, start=1)]
        by_seed = {row['seed']: row for row in seed_rows}
        rows = [by_seed[s] for s in _double_seed_flat_order(n) if s in by_seed]

    # sizes already keeps 4-team groups first and 3-team groups last.
    groups, cursor = [], 0
    for size in sizes:
        chunk = rows[cursor:cursor + size]
        cursor += size
        groups.append(_slots_from_double_chunk(chunk, mode='normal'))
    return groups

def _create_playoff_round(playoff_id, round_no, round_name, round_type, grouped_slots, round_meta=None):
    res = db.session.execute(text("""
        INSERT INTO playoff_rounds (playoff_id, round_no, round_name, round_type, round_meta_json)
        VALUES (:playoff_id, :round_no, :round_name, :round_type, :round_meta_json)
    """), {
        'playoff_id': playoff_id, 'round_no': round_no, 'round_name': round_name, 'round_type': round_type,
        'round_meta_json': json.dumps(round_meta or {}, ensure_ascii=False),
    })
    db.session.flush()
    if db.engine.dialect.name == 'postgresql':
        round_id = db.session.execute(text("SELECT currval(pg_get_serial_sequence('playoff_rounds','id')) AS id")).mappings().first()['id']
    else:
        round_id = res.lastrowid
    for group_no, slots in enumerate(grouped_slots, start=1):
        for slot_no, item in enumerate(slots, start=1):
            is_bye = item is None
            db.session.execute(text("""
                INSERT INTO playoff_slots (round_id, group_no, slot_no, seed, team_id, team_name, is_bye, entry_label, source_label)
                VALUES (:round_id, :group_no, :slot_no, :seed, :team_id, :team_name, :is_bye, :entry_label, :source_label)
            """), {
                'round_id': round_id, 'group_no': group_no, 'slot_no': slot_no,
                'seed': None if is_bye else item.get('seed'),
                'team_id': None if is_bye else item.get('team_id'),
                'team_name': 'X' if is_bye else item.get('team_name'),
                'is_bye': bool(is_bye),
                'entry_label': None if is_bye else (item.get('entry_label') or item.get('national_label') or item.get('regional64_label')),
                'source_label': None if is_bye else item.get('source_label'),
            })
    return round_id


def _safe_json_loads(raw, default=None):
    if not raw:
        return default if default is not None else {}
    try:
        return json.loads(raw)
    except Exception:
        return default if default is not None else {}


def _ab_pair_groups(rows, pairing_method='seed'):
    """จับคู่ในกลุ่ม A หรือ B โดยไม่รีเซ็ตเลข seed.

    สำคัญสำหรับ A/B หลัง Swiss:
    - A ใช้เลข 1..N
    - B ต้องใช้เลขต่อจาก A เช่น 9..16, 17..32
    ฟังก์ชันเดิมไปเรียก _playoff_knockout_groups() แล้วถูกนับใหม่เป็น 1..N
    จึงทำให้ B ออกเป็น 1 vs 8 ซ้ำกับ A; ฟังก์ชันนี้จะรักษา seed เดิมไว้เสมอ
    """
    rows = [dict(r) for r in (rows or [])]
    if not rows:
        return []

    method = pairing_method if pairing_method in {'seed', 'random', 'bracket'} else 'seed'
    if method == 'random':
        random.shuffle(rows)
    else:
        rows.sort(key=lambda r: (_as_int(r.get('seed'), 0), _as_int(r.get('rank'), 0), r.get('team_name') or ''))

    if len(rows) % 2 == 1:
        rows.append(None)

    if method == 'bracket' or method == 'random':
        return _adjacent_bracket_pairs(rows)

    # seed: ดีสุดพบลำดับท้ายสุดภายในกลุ่ม โดยเลข seed ยังเป็นเลขต่อเนื่องจริง
    raw_pairs = []
    n = len(rows)
    for i in range(n // 2):
        raw_pairs.append([rows[i], rows[n - 1 - i]])

    order = _pair_order_for_bracket(len(raw_pairs))
    by_no = {idx + 1: pair for idx, pair in enumerate(raw_pairs)}
    return [by_no[i] for i in order if i in by_no]


def _ab_make_round_groups(a_rows, b_rows, pairing_method='seed'):
    a_groups = _ab_pair_groups(a_rows, pairing_method)
    b_groups = _ab_pair_groups(b_rows, pairing_method)
    return a_groups + b_groups, {'a_pair_count': len(a_groups), 'b_pair_count': len(b_groups)}


def _regional64_enabled(view):
    """True เฉพาะเมื่อผู้ใช้เลือกโหมด/วิธีประกบคู่คัดทีมชาติเท่านั้น

    สำคัญ: ไม่เดาจากจำนวนทีม 64 หรือจำนวนเข้ารอบเอง
    เพื่อไม่ให้รายการอื่นที่จำนวนทีมบังเอิญเท่ากันโดนลอจิกพิเศษนี้
    """
    if not view or (view.get('competition') or {}).get('competition_type') != 'ab_ladder':
        return False
    config = _safe_json_loads((view.get('competition') or {}).get('config_json'), {})
    return (config.get('mode') == 'regional64_ladder'
            or config.get('regional64_ladder') is True
            or config.get('regional64') is True
            or config.get('pairing_method') == 'national_qualifier')


def _playoff_initial_team_count(view):
    # นับจำนวนทีมจริงในรอบเพลย์ออฟแรก ใช้กันกรณีผู้ใช้มาเลือกคัดตัวแทนทีมชาติในหน้าสร้างรอบถัดไป
    if not view or not view.get('round_views'):
        return 0
    first = view['round_views'][0]
    count = 0
    seen = set()
    for g in first.get('group_views', []):
        for slot in g.get('slots', []):
            if slot.get('is_bye'):
                continue
            key = slot.get('team_id') if slot.get('team_id') is not None else slot.get('team_name')
            if key in seen:
                continue
            seen.add(key)
            count += 1
    return count


def _national_qualifier_size(view, force=False):
    # คืนจำนวนทีมเพลย์ออฟเริ่มต้นของโหมดคัดตัวแทนทีมชาติ: 64 หรือ 48
    # force=True ใช้ตอนผู้ใช้เลือก "ประกบคู่คัดตัวแทนทีมชาติ" ในหน้าสร้างรอบถัดไป
    if not force and not _regional64_enabled(view):
        return 0
    config = _safe_json_loads((view.get('competition') or {}).get('config_json'), {}) if view else {}
    try:
        total = int(config.get('initial_total') or 0)
    except Exception:
        total = 0
    if total in (48, 64):
        return total
    total = _playoff_initial_team_count(view)
    return total if total in (48, 64) else 0


def _regional64_round4_created(view):
    """กันไม่ให้สร้างลอจิกรอบที่ 4 แบบตัวแทนทีมชาติ ซ้ำ"""
    for rv in (view or {}).get('round_views', []):
        meta = (rv.get('round') or {}).get('round_meta') or _safe_json_loads((rv.get('round') or {}).get('round_meta_json'), {})
        if meta.get('regional64_round4'):
            return True
    return False


def _national48_round4_created(view):
    """กันไม่ให้สร้างลอจิกรอบที่ 4 แบบคัดตัวแทนทีมชาติ 48 ทีม ซ้ำ"""
    for rv in (view or {}).get('round_views', []):
        meta = (rv.get('round') or {}).get('round_meta') or _safe_json_loads((rv.get('round') or {}).get('round_meta_json'), {})
        if meta.get('national48_round4'):
            return True
    return False



def _make_national48_round3_groups(latest_round_view):
    """สร้างรอบที่ 3 สำหรับคัดตัวแทนทีมชาติแบบ 48 ทีมหลัง Swiss

    หลังจบรอบ 2:
    - กลุ่ม A: ผู้ชนะรอบ 2 แข่งต่อกันตาม bracket เดิม 8 คู่
    - กลุ่ม B: A1:A2, A3:A4, A5:B1 ... A12:B8, A13:A14, A15:A16
    """
    if not latest_round_view:
        raise ValueError('ยังไม่มีรอบก่อนหน้าให้สร้างรอบที่ 3')
    a_winners, a_losers, b_winners = [], [], []
    for g in sorted(latest_round_view.get('group_views', []), key=lambda x: int(x.get('group_no') or 0)):
        zone = _ab_group_zone(latest_round_view, g.get('group_no'))
        res = g.get('result') or {}
        winner = res.get('winner')
        slots = [st for st in g.get('slots', []) if not st.get('is_bye')]
        loser = None
        if winner:
            for st in slots:
                if st.get('team_id') != winner.get('team_id') or st.get('team_name') != winner.get('team_name'):
                    loser = _slot_payload(st)
                    break
        if zone == 'A':
            if winner:
                a_winners.append(winner)
            if loser:
                a_losers.append(loser)
        else:
            if winner:
                b_winners.append(winner)

    if len(a_winners) < 16:
        raise ValueError(f'รอบที่ 3 แบบคัดตัวแทนทีมชาติ 48 ทีม ต้องมีผู้ชนะกลุ่ม A 16 ทีม แต่พบ {len(a_winners)} ทีม')
    if len(a_losers) < 16:
        raise ValueError(f'รอบที่ 3 แบบคัดตัวแทนทีมชาติ 48 ทีม ต้องมีผู้แพ้กลุ่ม A 16 ทีม แต่พบ {len(a_losers)} ทีม')
    if len(b_winners) < 8:
        raise ValueError(f'รอบที่ 3 แบบคัดตัวแทนทีมชาติ 48 ทีม ต้องมีผู้ชนะกลุ่ม B 8 ทีม แต่พบ {len(b_winners)} ทีม')

    a_rows = []
    for idx, p in enumerate(a_winners[:16], start=1):
        row = dict(p)
        row['rank'] = idx
        row['seed'] = idx
        row['entry_label'] = f'W-A{idx}'
        row['source_label'] = f'ชนะรอบ 2 กลุ่ม A สาย {idx}'
        a_rows.append(row)
    a_groups = _adjacent_bracket_pairs(a_rows)

    al = []
    for idx, p in enumerate(a_losers[:16], start=1):
        row = dict(p)
        row['rank'] = idx
        row['seed'] = idx
        row['national_label'] = f'L-A{idx}'
        row['entry_label'] = f'L-A{idx}'
        row['source_label'] = f'แพ้รอบ 2 กลุ่ม A สาย {idx}'
        al.append(row)

    bw = []
    for idx, p in enumerate(b_winners[:8], start=1):
        row = dict(p)
        row['rank'] = idx
        row['seed'] = 100 + idx
        row['national_label'] = f'W-B{idx}'
        row['entry_label'] = f'W-B{idx}'
        row['source_label'] = f'ชนะรอบ 2 กลุ่ม B สาย {idx}'
        bw.append(row)

    b_groups = [
        [al[0], al[1]],    # A1 : A2
        [al[2], al[3]],    # A3 : A4
        [al[4], bw[0]],    # A5 : B1
        [al[5], bw[1]],    # A6 : B2
        [al[6], bw[2]],    # A7 : B3
        [al[7], bw[3]],    # A8 : B4
        [al[8], bw[4]],    # A9 : B5
        [al[9], bw[5]],    # A10 : B6
        [al[10], bw[6]],   # A11 : B7
        [al[11], bw[7]],   # A12 : B8
        [al[12], al[13]],  # A13 : A14
        [al[14], al[15]],  # A15 : A16
    ]

    meta = {
        'a_pair_count': len(a_groups),
        'b_pair_count': len(b_groups),
        'round_kind': 'national48_round3',
        'national48_round3': True,
        'regional64_round3': True,
        'note': 'รอบที่ 3 คัดตัวแทนทีมชาติ 48 ทีม: กลุ่ม A bracket เดิม, กลุ่ม B A1:A2, A3:A4, A5:B1 ... A12:B8, A13:A14, A15:A16',
    }
    return a_groups + b_groups, meta



def _regional64_round3_created(view):
    """กันไม่ให้สร้างรอบ 3 แบบ A1-B1 ซ้ำ"""
    for rv in (view or {}).get('round_views', []):
        meta = (rv.get('round') or {}).get('round_meta') or _safe_json_loads((rv.get('round') or {}).get('round_meta_json'), {})
        if meta.get('regional64_round3'):
            return True
    return False


def _make_regional64_round3_groups(latest_round_view):
    """สร้างรอบที่ 3 ตามผังคัดตัวแทนทีมชาติ 64 ทีม

    หลังจบรอบ 2:
    - กลุ่ม A: ผู้ชนะรอบ 2 แข่งต่อกันตาม bracket เดิม 8 คู่
    - กลุ่ม B: ผู้แพ้รอบ 2 กลุ่ม A (A1-A16) เจอผู้ชนะรอบ 2 กลุ่ม B (B1-B16) แบบตรงตัว
    """
    if not latest_round_view:
        raise ValueError('ยังไม่มีรอบก่อนหน้าให้สร้างรอบที่ 3')
    a_winners, a_losers, b_winners = [], [], []
    for g in sorted(latest_round_view.get('group_views', []), key=lambda x: int(x.get('group_no') or 0)):
        zone = _ab_group_zone(latest_round_view, g.get('group_no'))
        res = g.get('result') or {}
        winner = res.get('winner')
        slots = [s for s in g.get('slots', []) if not s.get('is_bye')]
        loser = None
        if winner:
            for st in slots:
                if st.get('team_id') != winner.get('team_id') or st.get('team_name') != winner.get('team_name'):
                    loser = _slot_payload(st)
                    break
        if zone == 'A':
            if winner:
                a_winners.append(winner)
            if loser:
                a_losers.append(loser)
        else:
            if winner:
                b_winners.append(winner)

    if len(a_winners) < 16:
        raise ValueError(f'รอบที่ 3 แบบตัวแทนทีมชาติ ต้องมีผู้ชนะกลุ่ม A 16 ทีม แต่พบ {len(a_winners)} ทีม')
    if len(a_losers) < 16:
        raise ValueError(f'รอบที่ 3 แบบตัวแทนทีมชาติ ต้องมีผู้แพ้กลุ่ม A 16 ทีม แต่พบ {len(a_losers)} ทีม')
    if len(b_winners) < 16:
        raise ValueError(f'รอบที่ 3 แบบตัวแทนทีมชาติ ต้องมีผู้ชนะกลุ่ม B 16 ทีม แต่พบ {len(b_winners)} ทีม')

    # กลุ่ม A รอบ 3: ผู้ชนะรอบ 2 แข่งต่อแบบ bracket ติดกัน
    a_rows = []
    for idx, p in enumerate(a_winners[:16], start=1):
        row = dict(p)
        row['rank'] = idx
        row['seed'] = idx
        row['entry_label'] = f'W-A{idx}'
        row['source_label'] = f'ชนะรอบ 2 กลุ่ม A สาย {idx}'
        a_rows.append(row)
    a_groups = _adjacent_bracket_pairs(a_rows)

    # กลุ่ม B รอบ 3: A1-B1 ถึง A16-B16
    b_groups = []
    for idx in range(16):
        a_loss = dict(a_losers[idx])
        b_win = dict(b_winners[idx])
        a_loss['rank'] = idx + 1
        a_loss['seed'] = idx + 1
        a_loss['regional64_label'] = f'L-A{idx + 1}'
        a_loss['entry_label'] = f'L-A{idx + 1}'
        a_loss['source_label'] = f'แพ้รอบ 2 กลุ่ม A สาย {idx + 1}'
        b_win['rank'] = idx + 1
        b_win['seed'] = 100 + idx + 1
        b_win['regional64_label'] = f'W-B{idx + 1}'
        b_win['entry_label'] = f'W-B{idx + 1}'
        b_win['source_label'] = f'ชนะรอบ 2 กลุ่ม B สาย {idx + 1}'
        b_groups.append([a_loss, b_win])

    meta = {
        'a_pair_count': len(a_groups),
        'b_pair_count': len(b_groups),
        'round_kind': 'regional64_round3',
        'regional64_round3': True,
        'note': 'รอบที่ 3 ตามผังคัดตัวแทนทีมชาติ 64 ทีม: กลุ่ม A ผู้ชนะรอบ 2 ต่อ bracket, กลุ่ม B A1-B1 ถึง A16-B16',
    }
    return a_groups + b_groups, meta



def _make_national48_round4_groups(latest_round_view):
    """สร้างรอบที่ 4 สำหรับคัดตัวแทนทีมชาติแบบ 48 ทีม

    ทำแบบรอบ 4 ของ 64 ทีม แต่ 48 ทีมมี B1-B12 และ I A - VIII A:
      B1:B2,
      B3:I A, B4:II A, B5:III A, B6:IV A,
      B7:V A, B8:VI A, B9:VII A, B10:VIII A,
      B11:B12
    """
    if not latest_round_view:
        raise ValueError('ยังไม่มีรอบก่อนหน้าให้สร้างรอบที่ 4')
    a_losers, b_winners = [], []
    for g in sorted(latest_round_view.get('group_views', []), key=lambda x: int(x.get('group_no') or 0)):
        zone = _ab_group_zone(latest_round_view, g.get('group_no'))
        res = g.get('result') or {}
        winner = res.get('winner')
        slots = [s for s in g.get('slots', []) if not s.get('is_bye')]
        loser = None
        if winner:
            for s in slots:
                if s.get('team_id') != winner.get('team_id') or s.get('team_name') != winner.get('team_name'):
                    loser = _slot_payload(s)
                    break
        if zone == 'A':
            if loser:
                a_losers.append(loser)
        else:
            if winner:
                b_winners.append(winner)

    if len(b_winners) < 12:
        raise ValueError(f'รอบที่ 4 แบบคัดตัวแทนทีมชาติ 48 ทีม ต้องมีผู้ชนะกลุ่ม B 12 ทีม แต่พบ {len(b_winners)} ทีม')
    if len(a_losers) < 8:
        raise ValueError(f'รอบที่ 4 แบบคัดตัวแทนทีมชาติ 48 ทีม ต้องมีผู้แพ้รอบ 3 กลุ่ม A 8 ทีม แต่พบ {len(a_losers)} ทีม')

    b_rows = []
    for idx, p in enumerate(b_winners[:12], start=1):
        row = dict(p)
        row['rank'] = idx
        row['seed'] = idx
        row['national_label'] = f'W-B{idx}'
        row['entry_label'] = f'W-B{idx}'
        row['source_label'] = f'ชนะรอบ 3 กลุ่ม B สาย {idx}'
        b_rows.append(row)

    roman = ['I A', 'II A', 'III A', 'IV A', 'V A', 'VI A', 'VII A', 'VIII A']
    a_rows = []
    for idx, p in enumerate(a_losers[:8], start=1):
        row = dict(p)
        row['rank'] = idx
        row['seed'] = 100 + idx
        row['national_label'] = f"L-{roman[idx - 1]}"
        row['entry_label'] = f"L-{roman[idx - 1]}"
        row['source_label'] = f'แพ้รอบ 3 กลุ่ม A สาย {idx}'
        a_rows.append(row)

    groups = [
        [b_rows[0], b_rows[1]],    # B1 : B2
        [b_rows[2], a_rows[0]],    # B3 : I A
        [b_rows[3], a_rows[1]],    # B4 : II A
        [b_rows[4], a_rows[2]],    # B5 : III A
        [b_rows[5], a_rows[3]],    # B6 : IV A
        [b_rows[6], a_rows[4]],    # B7 : V A
        [b_rows[7], a_rows[5]],    # B8 : VI A
        [b_rows[8], a_rows[6]],    # B9 : VII A
        [b_rows[9], a_rows[7]],    # B10 : VIII A
        [b_rows[10], b_rows[11]],  # B11 : B12
    ]
    meta = {
        'a_pair_count': 0,
        'b_pair_count': len(groups),
        'round_kind': 'national48_round4',
        'national48_round4': True,
        'note': 'รอบที่ 4 คัดตัวแทนทีมชาติ 48 ทีม: B1:B2, B3:I A, B4:II A, B5:III A ... B10:VIII A, B11:B12',
    }
    return groups, meta



def _make_regional64_round4_groups(latest_round_view):
    """สร้างรอบที่ 4 ตามผังคัดตัวแทนทีมชาติ 64 ทีม

    รอบก่อนหน้าให้ระบบทำงานตาม seed/bracket เดิมทั้งหมด
    เฉพาะตอนสร้าง "รอบที่ 4" ให้เอา:
    - ผู้ชนะรอบ 3 กลุ่ม B เรียงเป็น B1-B16
    - ผู้แพ้รอบ 3 กลุ่ม A เรียงเป็น I A - VIII A
    แล้วจัด 12 คู่:
      B1:B2, B3:B4,
      B5:I A, B6:II A, ... B12:VIII A,
      B13:B14, B15:B16
    """
    if not latest_round_view:
        raise ValueError('ยังไม่มีรอบก่อนหน้าให้สร้างรอบที่ 4')
    a_losers, b_winners = [], []
    for g in sorted(latest_round_view.get('group_views', []), key=lambda x: int(x.get('group_no') or 0)):
        zone = _ab_group_zone(latest_round_view, g.get('group_no'))
        res = g.get('result') or {}
        winner = res.get('winner')
        slots = [s for s in g.get('slots', []) if not s.get('is_bye')]
        loser = None
        if winner:
            for s in slots:
                if s.get('team_id') != winner.get('team_id') or s.get('team_name') != winner.get('team_name'):
                    loser = _slot_payload(s)
                    break
        if zone == 'A':
            if loser:
                a_losers.append(loser)
        else:
            if winner:
                b_winners.append(winner)

    if len(b_winners) < 16:
        raise ValueError(f'รอบที่ 4 แบบตัวแทนทีมชาติ ต้องมีผู้ชนะกลุ่ม B 16 ทีม แต่พบ {len(b_winners)} ทีม')
    if len(a_losers) < 8:
        raise ValueError(f'รอบที่ 4 แบบตัวแทนทีมชาติ ต้องมีผู้แพ้รอบ 3 กลุ่ม A 8 ทีม แต่พบ {len(a_losers)} ทีม')

    # ติดป้าย seed/label ใหม่เพื่อให้เรียงและดูง่ายตามผัง
    b_rows = []
    for idx, p in enumerate(b_winners[:16], start=1):
        row = dict(p)
        row['rank'] = idx
        row['seed'] = idx
        row['regional64_label'] = f'W-B{idx}'
        row['entry_label'] = f'W-B{idx}'
        row['source_label'] = f'ชนะรอบ 3 กลุ่ม B สาย {idx}'
        b_rows.append(row)

    a_rows = []
    roman = ['I A', 'II A', 'III A', 'IV A', 'V A', 'VI A', 'VII A', 'VIII A']
    for idx, p in enumerate(a_losers[:8], start=1):
        row = dict(p)
        row['rank'] = idx
        row['seed'] = 100 + idx
        row['regional64_label'] = f"L-{roman[idx - 1]}"
        row['entry_label'] = f"L-{roman[idx - 1]}"
        row['source_label'] = f'แพ้รอบ 3 กลุ่ม A สาย {idx}'
        a_rows.append(row)

    groups = [
        [b_rows[0], b_rows[1]],    # B1 : B2
        [b_rows[2], b_rows[3]],    # B3 : B4
        [b_rows[4], a_rows[0]],    # B5 : I A
        [b_rows[5], a_rows[1]],    # B6 : II A
        [b_rows[6], a_rows[2]],    # B7 : III A
        [b_rows[7], a_rows[3]],    # B8 : IV A
        [b_rows[8], a_rows[4]],    # B9 : V A
        [b_rows[9], a_rows[5]],    # B10 : VI A
        [b_rows[10], a_rows[6]],   # B11 : VII A
        [b_rows[11], a_rows[7]],   # B12 : VIII A
        [b_rows[12], b_rows[13]],  # B13 : B14
        [b_rows[14], b_rows[15]],  # B15 : B16
    ]
    meta = {
        'a_pair_count': 0,
        'b_pair_count': len(groups),
        'round_kind': 'regional64_round4',
        'regional64_round4': True,
        'note': 'รอบที่ 4 ตามผังคัดตัวแทนทีมชาติ 64 ทีม: B1:B2, B3:B4, B5:I A ... B12:VIII A, B13:B14, B15:B16',
    }
    return groups, meta


def _create_ab_ladder_competition(source_event, selected_rows, title, pairing_method, a_team_count, advance_a, advance_b, direct_rows=None, special_mode=None):
    """สร้างระบบ A/B ต่อจาก Standing:
    - ทีมบนตาม Standing เข้า A ตามจำนวนที่กำหนด
    - ทีมที่เหลือเข้า B
    - A ชนะอยู่ A, A แพ้ตกลง B, B แพ้ตกรอบ
    """
    selected_rows = list(selected_rows or [])
    direct_rows = list(direct_rows or [])
    total = len(selected_rows)
    if total < 2:
        raise ValueError('ระบบ A/B ต้องมีทีมอย่างน้อย 2 ทีม')
    a_team_count = max(1, min(total, int(a_team_count or max(1, total // 2))))
    advance_a = max(0, int(advance_a or 0))
    advance_b = max(0, int(advance_b or 0))
    a_rows = [_row_to_seed_payload(row, idx) for idx, row in enumerate(selected_rows[:a_team_count], start=1)]
    b_rows = [_row_to_seed_payload(row, idx + a_team_count) for idx, row in enumerate(selected_rows[a_team_count:], start=1)]
    groups, meta = _ab_make_round_groups(a_rows, b_rows, pairing_method)
    config = {
        'mode': special_mode or 'ab_ladder',
        'initial_total': total,
        'initial_a_count': a_team_count,
        'initial_b_count': len(b_rows),
        'advance_a': advance_a,
        'advance_b': advance_b,
        'pairing_method': pairing_method,
        # ทีมที่ไม่ต้องแข่งต่อหลัง Swiss เช่น อันดับ 1-2 ให้ถือว่าเข้ารอบ/ได้ลำดับก่อน A/B
        'regional64_ladder': bool(special_mode == 'regional64_ladder'),
        'direct_qualified': [
            {
                'rank': idx,
                'team_id': row.get('team_id'),
                'team_name': row.get('team_name'),
                'seed': row.get('seed') or row.get('rank') or idx,
            }
            for idx, row in enumerate(direct_rows, start=1)
        ],
    }
    res = db.session.execute(text("""
        INSERT INTO playoff_competitions (source_event_id, title, competition_type, pairing_method, config_json)
        VALUES (:source_event_id, :title, :competition_type, :pairing_method, :config_json)
    """), {
        'source_event_id': source_event.id,
        'title': title,
        'competition_type': 'ab_ladder',
        'pairing_method': pairing_method,
        'config_json': json.dumps(config, ensure_ascii=False),
    })
    db.session.flush()
    if db.engine.dialect.name == 'postgresql':
        playoff_id = db.session.execute(text("SELECT currval(pg_get_serial_sequence('playoff_competitions','id')) AS id")).mappings().first()['id']
    else:
        playoff_id = res.lastrowid
    meta.update({'a_team_count': len(a_rows), 'b_team_count': len(b_rows), 'round_kind': 'ab_ladder'})
    _create_playoff_round(playoff_id, 1, title or 'A/B รอบแรก', 'ab_ladder', groups, round_meta=meta)
    db.session.commit()
    return playoff_id


def _create_playoff_competition(source_event, selected_rows, title, competition_type, pairing_method, add_bye=False, direct_rows=None):
    config = {
        'direct_qualified': [
            {
                'rank': idx,
                'team_id': row.get('team_id'),
                'team_name': row.get('team_name'),
                'seed': row.get('seed') or row.get('rank') or idx,
            }
            for idx, row in enumerate((direct_rows or []), start=1)
        ]
    }
    res = db.session.execute(text("""
        INSERT INTO playoff_competitions (source_event_id, title, competition_type, pairing_method, config_json)
        VALUES (:source_event_id, :title, :competition_type, :pairing_method, :config_json)
    """), {
        'source_event_id': source_event.id,
        'title': title,
        'competition_type': competition_type,
        'pairing_method': pairing_method,
        'config_json': json.dumps(config, ensure_ascii=False),
    })
    db.session.flush()
    if db.engine.dialect.name == 'postgresql':
        playoff_id = db.session.execute(text("SELECT currval(pg_get_serial_sequence('playoff_competitions','id')) AS id")).mappings().first()['id']
    else:
        playoff_id = res.lastrowid
    if competition_type == 'double_knockout':
        groups = _playoff_double_groups(selected_rows, pairing_method, add_bye=add_bye)
    else:
        groups = _playoff_knockout_groups(selected_rows, pairing_method, add_bye=add_bye)
    _create_playoff_round(playoff_id, 1, title or 'รอบที่ 1', competition_type, groups)
    db.session.commit()
    return playoff_id


def _fetch_playoff(playoff_id):
    comp = db.session.execute(text("SELECT * FROM playoff_competitions WHERE id = :id"), {'id': playoff_id}).mappings().first()
    if not comp:
        return None
    rounds = db.session.execute(text("SELECT * FROM playoff_rounds WHERE playoff_id = :pid ORDER BY round_no"), {'pid': playoff_id}).mappings().all()
    round_views = []
    for rnd in rounds:
        slots = db.session.execute(text("SELECT * FROM playoff_slots WHERE round_id = :rid ORDER BY group_no, slot_no"), {'rid': rnd['id']}).mappings().all()
        scores = db.session.execute(text("SELECT * FROM playoff_scores WHERE round_id = :rid"), {'rid': rnd['id']}).mappings().all()
        manuals = db.session.execute(text("SELECT * FROM playoff_manual_results WHERE round_id = :rid"), {'rid': rnd['id']}).mappings().all()
        score_map = {(s['group_no'], s['slot_no'], s['stage_no']): s['score'] for s in scores}
        manual_map = {m['group_no']: m for m in manuals}
        group_nos = sorted({slot['group_no'] for slot in slots})
        group_views = []
        for group_no in group_nos:
            gslots = [dict(slot) for slot in slots if slot['group_no'] == group_no]
            _apply_bye_auto_scores_to_map(gslots, score_map)
            result = _compute_playoff_group_result(rnd['round_type'], gslots, score_map, manual_map.get(group_no))
            stage_state = _build_playoff_stage_state(rnd['round_type'], gslots, score_map)
            manual_options = [s for s in gslots if not s.get('is_bye')]
            group_views.append({'group_no': group_no, 'slots': gslots, 'result': result, 'stage_state': stage_state, 'manual_options': manual_options, 'manual_override': manual_map.get(group_no)})
        
        rnd_dict = dict(rnd)
        rnd_dict['round_meta'] = _safe_json_loads(rnd_dict.get('round_meta_json'), {})
        round_view = {'round': rnd_dict, 'score_map': score_map, 'group_views': group_views}
        if rnd_dict.get('round_type') == 'ab_ladder':
            round_view['ab_summary'] = _ab_participants_from_round(round_view)
        round_views.append(round_view)
    return {'competition': dict(comp), 'round_views': round_views}


def _slot_payload(slot):
    if not slot or slot.get('is_bye'):
        return None
    return {
        'team_id': slot.get('team_id'),
        'team_name': slot.get('team_name'),
        'seed': slot.get('seed') or slot.get('slot_no'),
        'rank': slot.get('seed') or slot.get('slot_no'),
        'entry_label': slot.get('entry_label'),
        'source_label': slot.get('source_label'),
    }


def _ab_local_group_no(round_view, group_no):
    """เลขสายย่อยภายในกลุ่ม A หรือ B (B ต้องเริ่มนับใหม่จาก 1)"""
    meta = (round_view.get('round') or {}).get('round_meta') or {}
    a_pair_count = int(meta.get('a_pair_count') or 0)
    group_no = int(group_no or 0)
    return group_no if group_no <= a_pair_count else max(1, group_no - a_pair_count)


def _ab_tag_flow(payload, round_view, group_no, zone, result_kind, entry_label=None):
    """แนบที่มาของทีมแบบอ่านย้อนหลังได้ เช่น ชนะรอบ 3 กลุ่ม B สาย 5"""
    if not payload:
        return None
    row = dict(payload)
    round_no = int(((round_view.get('round') or {}).get('round_no')) or 0)
    local_no = _ab_local_group_no(round_view, group_no)
    row['entry_label'] = entry_label or f"{'W' if result_kind == 'ชนะ' else 'L'}-{zone}{local_no}"
    row['source_label'] = f'{result_kind}รอบ {round_no} กลุ่ม {zone} สาย {local_no}'
    return row


def _apply_bye_auto_scores_to_map(slots, score_map):
    """Display/compute BYE pairs as 13:0 without asking to fill the X slot.

    X/BYE is a real bracket placeholder, not a missing team name.
    When a real team meets X in stage 1, the real team wins automatically 13:0.
    This fills the in-memory score map for UI colors, result calculation, and reports.
    It does not require the user to type a team name into the X slot.
    """
    by_slot = {int(s.get('slot_no')): s for s in slots}
    for a_no, b_no in ((1, 2), (3, 4)):
        a = by_slot.get(a_no)
        b = by_slot.get(b_no)
        if not a or not b:
            continue
        if bool(a.get('is_bye')) == bool(b.get('is_bye')):
            continue
        group_no = int(a.get('group_no') or b.get('group_no') or 0)
        if a.get('is_bye') and not b.get('is_bye'):
            score_map[(group_no, a_no, 1)] = 0
            score_map[(group_no, b_no, 1)] = 13
        elif b.get('is_bye') and not a.get('is_bye'):
            score_map[(group_no, a_no, 1)] = 13
            score_map[(group_no, b_no, 1)] = 0


def _playoff_score(score_map, slot, stage):
    if not slot:
        return None
    if slot.get('is_bye'):
        return 0
    return score_map.get((slot['group_no'], slot['slot_no'], stage))


def _decide_playoff_pair(a, b, stage_no, score_map):
    """Decide one pair like the original petanque_tournament engine.
    - BYE/X auto-advances the real team.
    - If both score boxes are blank, the pair is still incomplete.
    - If one side is keyed and the other is blank, blank is treated as 0.
      This allows either direct score entry or manual dropdown override.
    """
    if not a and not b:
        return {'a': a, 'b': b, 'stage': stage_no, 'winner': None, 'loser': None, 'complete': False, 'score_a': None, 'score_b': None}
    if a and a.get('is_bye') and b and not b.get('is_bye'):
        return {'a': a, 'b': b, 'stage': stage_no, 'winner': b, 'loser': a, 'complete': True, 'score_a': 0, 'score_b': 13}
    if b and b.get('is_bye') and a and not a.get('is_bye'):
        return {'a': a, 'b': b, 'stage': stage_no, 'winner': a, 'loser': b, 'complete': True, 'score_a': 13, 'score_b': 0}
    if a and not b:
        return {'a': a, 'b': b, 'stage': stage_no, 'winner': a, 'loser': None, 'complete': True, 'score_a': None, 'score_b': None}
    if b and not a:
        return {'a': a, 'b': b, 'stage': stage_no, 'winner': b, 'loser': None, 'complete': True, 'score_a': None, 'score_b': None}

    sa_raw = _playoff_score(score_map, a, stage_no)
    sb_raw = _playoff_score(score_map, b, stage_no)
    if sa_raw is None and sb_raw is None:
        return {'a': a, 'b': b, 'stage': stage_no, 'winner': None, 'loser': None, 'complete': False, 'score_a': None, 'score_b': None}

    sa = 0 if sa_raw is None else int(sa_raw)
    sb = 0 if sb_raw is None else int(sb_raw)
    if sa == sb:
        return {'a': a, 'b': b, 'stage': stage_no, 'winner': None, 'loser': None, 'complete': False, 'score_a': sa, 'score_b': sb}
    winner = a if sa > sb else b
    loser = b if winner is a else a
    return {'a': a, 'b': b, 'stage': stage_no, 'winner': winner, 'loser': loser, 'complete': True, 'score_a': sa, 'score_b': sb}


def _double_group_decisions(slots, score_map):
    by = {s['slot_no']: s for s in slots}
    qf1 = _decide_playoff_pair(by.get(1), by.get(2), 1, score_map)
    qf2 = _decide_playoff_pair(by.get(3), by.get(4), 1, score_map)
    wf = _decide_playoff_pair(qf1.get('winner'), qf2.get('winner'), 2, score_map) if qf1.get('complete') and qf2.get('complete') else None
    lf = _decide_playoff_pair(qf1.get('loser'), qf2.get('loser'), 2, score_map) if qf1.get('loser') and qf2.get('loser') else None
    final2 = _decide_playoff_pair(wf.get('loser'), lf.get('winner'), 3, score_map) if wf and wf.get('loser') and lf and lf.get('winner') else None
    return {'qf1': qf1, 'qf2': qf2, 'wf': wf, 'lf': lf, 'final2': final2}
def _blank_stage_state(slots):
    state = {}
    for s in slots:
        state[s['slot_no']] = {
            1: {'color': '', 'editable': True},
            2: {'color': '', 'editable': False},
            3: {'color': '', 'editable': False},
        }
    return state


def _mark_decision_state(state, dec):
    if not dec:
        return
    a, b = dec.get('a'), dec.get('b')
    st = int(dec.get('stage') or 1)
    for side in ('a', 'b'):
        s = dec.get(side)
        if s and s.get('slot_no') in state:
            state[s['slot_no']][st]['editable'] = True
    if dec.get('winner') and dec.get('loser'):
        w = dec['winner'].get('slot_no')
        l = dec['loser'].get('slot_no')
        if w in state:
            state[w][st]['color'] = 'win'
        if l in state:
            state[l][st]['color'] = 'loss'


def _build_playoff_stage_state(round_type, slots, score_map):
    state = _blank_stage_state(slots)
    if round_type in ('knockout', 'ab_ladder'):
        dec = _decide_playoff_pair(slots[0] if len(slots)>0 else None, slots[1] if len(slots)>1 else None, 1, score_map)
        _mark_decision_state(state, dec)
        return state
    decisions = _double_group_decisions(slots, score_map)
    for key in ('qf1','qf2'):
        _mark_decision_state(state, decisions.get(key))
    qf1, qf2 = decisions.get('qf1'), decisions.get('qf2')
    if qf1 and qf1.get('complete') and qf2 and qf2.get('complete'):
        for key in ('wf','lf'):
            d = decisions.get(key)
            if d:
                _mark_decision_state(state, d)
                for side in ('a','b'):
                    s = d.get(side)
                    if s and s.get('slot_no') in state:
                        state[s['slot_no']][2]['editable'] = True
    d = decisions.get('final2')
    if d:
        _mark_decision_state(state, d)
        for side in ('a','b'):
            s = d.get(side)
            if s and s.get('slot_no') in state:
                state[s['slot_no']][3]['editable'] = True
    return state

def _compute_playoff_group_result(round_type, slots, score_map, manual=None):
    """Compute winners using bracket flow from the original sample.
    Manual dropdown is optional; score boxes alone can finish the group.
    """
    valid = [s for s in slots if not s.get('is_bye')]
    by_slot = {s['slot_no']: s for s in slots}
    if manual and manual.get('winner_slot_no'):
        winner = by_slot.get(manual.get('winner_slot_no'))
        second = by_slot.get(manual.get('second_slot_no')) if round_type == 'double_knockout' else None
        if winner and not winner.get('is_bye'):
            if round_type != 'double_knockout' or (second and not second.get('is_bye') and second.get('slot_no') != winner.get('slot_no')):
                return {'winner': _slot_payload(winner), 'second': _slot_payload(second), 'complete': True, 'manual': True}

    if round_type in ('knockout', 'ab_ladder'):
        if len(valid) == 1:
            return {'winner': _slot_payload(valid[0]), 'second': None, 'complete': True, 'manual': False}
        if len(valid) < 2:
            return {'winner': None, 'second': None, 'complete': False, 'manual': False}
        dec = _decide_playoff_pair(slots[0] if len(slots) > 0 else None, slots[1] if len(slots) > 1 else None, 1, score_map)
        return {'winner': _slot_payload(dec.get('winner')) if dec and dec.get('winner') else None, 'second': None, 'complete': bool(dec and dec.get('winner')), 'manual': False}

    # Double knockout: original flow
    # stage 1: 1-2 and 3-4
    # stage 2: winners vs winners, losers vs losers
    # stage 3: loser of winners-final vs winner of losers-final for rank 2
    if len(valid) <= 2:
        if len(valid) < 2:
            return {'winner': _slot_payload(valid[0]) if valid else None, 'second': None, 'complete': bool(valid), 'manual': False}
        dec = _decide_playoff_pair(valid[0], valid[1], 1, score_map)
        return {'winner': _slot_payload(dec.get('winner')) if dec and dec.get('winner') else None,
                'second': _slot_payload(dec.get('loser')) if dec and dec.get('loser') else None,
                'complete': bool(dec and dec.get('complete')), 'manual': False}
    decisions = _double_group_decisions(slots, score_map)
    wf = decisions.get('wf')
    lf = decisions.get('lf')
    final2 = decisions.get('final2')
    top1 = wf.get('winner') if wf and wf.get('complete') else None
    top2 = final2.get('winner') if final2 and final2.get('complete') else None
    return {'winner': _slot_payload(top1) if top1 else None, 'second': _slot_payload(top2) if top2 else None, 'complete': bool(top1 and top2), 'manual': False}


def _ab_group_zone(round_view, group_no):
    meta = (round_view.get('round') or {}).get('round_meta') or _safe_json_loads((round_view.get('round') or {}).get('round_meta_json'), {})
    a_pair_count = int(meta.get('a_pair_count') or 0)
    return 'A' if int(group_no) <= a_pair_count else 'B'


def _ab_participants_from_round(round_view):
    """สรุปผลรอบ A/B: A ชนะไป A, A แพ้ลง B, B ชนะอยู่ B, B แพ้ตกรอบ"""
    next_a, next_b, eliminated = [], [], []
    a_winners, a_losers, b_winners, b_losers = [], [], [], []
    for g in round_view.get('group_views', []):
        zone = _ab_group_zone(round_view, g.get('group_no'))
        res = g.get('result') or {}
        winner = res.get('winner')
        slots = [s for s in g.get('slots', []) if not s.get('is_bye')]
        loser = None
        if winner:
            for s in slots:
                if s.get('team_id') != winner.get('team_id') or s.get('team_name') != winner.get('team_name'):
                    loser = _slot_payload(s)
                    break
        if zone == 'A':
            if winner:
                winner = _ab_tag_flow(winner, round_view, g.get('group_no'), 'A', 'ชนะ')
                a_winners.append(winner); next_a.append(winner)
            if loser:
                loser = _ab_tag_flow(loser, round_view, g.get('group_no'), 'A', 'แพ้')
                a_losers.append(loser); next_b.append(loser)
        else:
            if winner:
                winner = _ab_tag_flow(winner, round_view, g.get('group_no'), 'B', 'ชนะ')
                b_winners.append(winner); next_b.append(winner)
            if loser:
                loser = _ab_tag_flow(loser, round_view, g.get('group_no'), 'B', 'แพ้')
                b_losers.append(loser); eliminated.append(loser)
    return {
        'next_a': next_a, 'next_b': next_b, 'eliminated': eliminated,
        'a_winners': a_winners, 'a_losers': a_losers,
        'b_winners': b_winners, 'b_losers': b_losers,
    }


def _team_key_payload(p):
    if not p:
        return None
    return p.get('team_id') if p.get('team_id') is not None else p.get('team_name')


def _dedupe_payload_rows(rows):
    seen, out = set(), []
    for p in rows or []:
        key = _team_key_payload(p)
        if key in seen:
            continue
        seen.add(key)
        out.append(p)
    return out


def _ab_winners_by_zone(round_view, zone_name):
    """คืนผู้ชนะของ zone A/B เรียงตามเลขสายจริงในรอบนั้น ใช้เป็น คนที่ 1,2,3,4 ตามเลขสาย"""
    rows = []
    if not round_view:
        return rows
    for g in sorted(round_view.get('group_views', []), key=lambda x: int(x.get('group_no') or 0)):
        if _ab_group_zone(round_view, g.get('group_no')) != zone_name:
            continue
        winner = (g.get('result') or {}).get('winner')
        if not winner:
            continue
        payload = dict(winner)
        payload['group_no'] = int(g.get('group_no') or 0)
        rows.append(payload)
    return _dedupe_payload_rows(rows)


def _ab_survivors_by_zone(round_view, zone_name):
    """คืนทีมที่ยังอยู่ในสาย A/B หลังจบรอบนั้น เรียงตามเลขสายจริง
    - A = ผู้ชนะจากคู่สาย A
    - B = ผู้แพ้จากคู่สาย A ที่ตกลง B + ผู้ชนะจากคู่สาย B
    ใช้สำหรับรายงาน “เข้ารอบคนที่ 1,2,3...” เมื่อจำนวนเหลือตามเป้า
    """
    rows = []
    if not round_view:
        return rows
    for g in sorted(round_view.get('group_views', []), key=lambda x: int(x.get('group_no') or 0)):
        zone = _ab_group_zone(round_view, g.get('group_no'))
        if zone_name == 'A' and zone != 'A':
            continue
        if zone_name == 'B' and zone not in ('A', 'B'):
            continue
        res = g.get('result') or {}
        winner = res.get('winner')
        slots = [s for s in g.get('slots', []) if not s.get('is_bye')]
        loser = None
        if winner:
            for s in slots:
                if s.get('team_id') != winner.get('team_id') or s.get('team_name') != winner.get('team_name'):
                    loser = _slot_payload(s)
                    break
        payload = None
        if zone_name == 'A' and zone == 'A':
            payload = winner
        elif zone_name == 'B' and zone == 'A':
            payload = loser
        elif zone_name == 'B' and zone == 'B':
            payload = winner
        if payload:
            payload = dict(payload)
            payload['group_no'] = int(g.get('group_no') or 0)
            payload['from_zone'] = zone
            rows.append(payload)
    return _dedupe_payload_rows(rows)


def _ab_state_for_view(view):
    """สถานะระบบ A/B ทั้งชุด
    - ถ้า A เหลือถึงเป้าแล้ว ให้ถือว่า A จบ และเก็บรายชื่อผู้ผ่าน A จากรอบนั้นไว้
    - รอบต่อไปจะไม่เอา A ที่จบแล้วไปแข่งซ้ำ เหลือเฉพาะ B ที่ยังไม่ถึงเป้า
    """
    if not view or (view.get('competition') or {}).get('competition_type') != 'ab_ladder':
        return None
    config = _safe_json_loads((view.get('competition') or {}).get('config_json'), {})
    target_a = int(config.get('advance_a') or 0)
    target_b = int(config.get('advance_b') or 0)
    direct_qualified = list(config.get('direct_qualified') or [])
    direct_count = len(direct_qualified)
    final_a, final_b = [], []
    latest_complete_summary = {}
    latest_complete_round = None

    for rv in view.get('round_views', []):
        if not _playoff_round_complete(rv):
            continue
        latest_complete_round = rv
        summary = _ab_participants_from_round(rv)
        latest_complete_summary = summary
        a_candidates = _ab_survivors_by_zone(rv, 'A')
        b_candidates = _ab_survivors_by_zone(rv, 'B')
        if target_a > 0 and not final_a and a_candidates and len(a_candidates) <= target_a:
            final_a = a_candidates
        if target_b > 0 and not final_b and b_candidates and len(b_candidates) <= target_b:
            final_b = b_candidates

    latest = view.get('round_views', [])[-1] if view.get('round_views') else None
    latest_summary = _ab_participants_from_round(latest) if latest else {}
    latest_complete = bool(latest and _playoff_round_complete(latest))
    current_a = final_a if final_a else (latest_summary.get('next_a') or [])
    current_b = final_b if final_b else (latest_summary.get('next_b') or [])
    a_finished = bool(target_a > 0 and final_a)
    b_finished = bool(target_b > 0 and final_b)
    all_finished = bool((target_a <= 0 or a_finished) and (target_b <= 0 or b_finished))

    return {
        'config': config,
        'latest': latest_summary,
        'latest_complete_summary': latest_complete_summary,
        'current_a_count': len(current_a),
        'current_b_count': len(current_b),
        'target_a': target_a,
        'target_b': target_b,
        'a_finished': a_finished,
        'b_finished': b_finished,
        'all_finished': all_finished,
        'reached_target': all_finished,
        # รายงาน A/B ใช้ลำดับรวมต่อเนื่อง:
        # direct จาก Swiss มาก่อน แล้วตามด้วย A และ B เช่น direct 1-2, A 3-4, B 5-...
        'direct_qualified': [{'rank': i, 'team_name': p.get('team_name'), 'group_no': p.get('group_no')} for i, p in enumerate(direct_qualified, start=1)],
        'final_a': [{'rank': i + direct_count, 'team_name': p.get('team_name'), 'group_no': p.get('group_no')} for i, p in enumerate(final_a, start=1)],
        'final_b': [{'rank': i + direct_count + len(final_a), 'team_name': p.get('team_name'), 'group_no': p.get('group_no')} for i, p in enumerate(final_b, start=1)],
        'latest_round_complete': latest_complete,
        'latest_complete_round_id': (latest_complete_round or {}).get('round', {}).get('id') if latest_complete_round else None,
    }


def _ab_next_rows_for_creation(view, latest, pairing_method='seed'):
    """ทีมที่จะสร้างรอบ A/B ถัดไป: ถ้า A หรือ B จบตามเป้าแล้ว จะไม่ส่งสายนั้นไปแข่งต่อ"""
    state = _ab_state_for_view(view) or {}
    if state.get('all_finished'):
        return [], []
    ab = _ab_participants_from_round(latest)
    next_a = [] if state.get('a_finished') else (ab.get('next_a') or [])
    next_b = [] if state.get('b_finished') else (ab.get('next_b') or [])
    a_rows = [dict(p, rank=idx, seed=idx) for idx, p in enumerate(next_a, start=1)]
    b_rows = [dict(p, rank=idx + len(a_rows), seed=idx + len(a_rows)) for idx, p in enumerate(next_b, start=1)]
    return a_rows, b_rows


def _ab_status_for_view(view):
    return _ab_state_for_view(view)


def _playoff_round_complete(round_view):
    return all(g['result'].get('complete') for g in round_view.get('group_views', []))


def _participants_from_round(round_view):
    participants = []
    rtype = round_view['round']['round_type']
    if rtype == 'ab_ladder':
        ab = _ab_participants_from_round(round_view)
        participants = (ab.get('next_a') or []) + (ab.get('next_b') or [])
    else:
        for g in round_view.get('group_views', []):
            if g['result'].get('winner'):
                participants.append(g['result']['winner'])
            if rtype == 'double_knockout' and g['result'].get('second'):
                participants.append(g['result']['second'])
    seen, out = set(), []
    for p in participants:
        key = p.get('team_id') or p.get('team_name')
        if key not in seen:
            seen.add(key); out.append(p)
    return out


def _next_playoff_round_no(playoff_id):
    row = db.session.execute(text("SELECT COALESCE(MAX(round_no), 0) AS n FROM playoff_rounds WHERE playoff_id = :pid"), {'pid': playoff_id}).mappings().first()
    return int(row['n'] or 0) + 1



def _playoff_source_teams(playoff_id):
    rows = db.session.execute(text('''
        SELECT ps.team_id, ps.team_name, COALESCE(ps.seed, ps.slot_no) AS seed
        FROM playoff_slots ps
        JOIN playoff_rounds pr ON pr.id = ps.round_id
        WHERE pr.playoff_id = :pid AND pr.round_no = 1 AND ps.is_bye = false
        ORDER BY COALESCE(ps.seed, ps.slot_no), ps.id
    '''), {'pid': playoff_id}).mappings().all()
    return [{'id': r['team_id'], 'name': r['team_name'], 'seed': r['seed']} for r in rows]



def _source_event_full_report(event_id, qualified_team_ids=None, status_by_team_id=None):
    """รายงานรอบแรก Swiss: รายชื่อทีม, ผลการแข่งขันแต่ละครั้ง, ผลจัดลำดับ."""
    if not event_id:
        return {'teams': [], 'rounds': [], 'standings': []}
    qids = {int(x) for x in (qualified_team_ids or []) if x is not None}
    status_by_team_id = status_by_team_id or {}
    teams = Team.query.filter_by(event_id=event_id).order_by(Team.id.asc()).all()
    matches = Match.query.filter_by(event_id=event_id).order_by(Match.round.asc(), Match.field.asc(), Match.id.asc()).all()
    round_map = {}
    for m in matches:
        t1 = m.team1.name if m.team1 else '-'
        t2 = m.team2.name if m.team2 else 'BYE'
        s1 = '0' if m.team1_score is None else str(m.team1_score)
        s2 = '0' if m.team2_score is None else str(m.team2_score)
        if not m.is_locked:
            result = 'รอยืนยันผล'
        elif m.team2_id is None:
            result = f'{t1} ชนะ BYE'
        elif int(m.team1_score or 0) > int(m.team2_score or 0):
            result = f'{t1} ชนะ'
        elif int(m.team2_score or 0) > int(m.team1_score or 0):
            result = f'{t2} ชนะ'
        else:
            result = 'เสมอ'
        round_map.setdefault(m.round or 1, []).append({'field': m.field or '', 'team1': t1, 'team2': t2, 'score': f'{s1} - {s2}' if m.team2_id else 'BYE', 'result': result})
    standings_rows = []
    for idx, row in enumerate(calculate_standings(event_id), start=1):
        tid = row.get('team_id')
        status = status_by_team_id.get(tid) or ('เข้ารอบ' if tid in qids else 'ตกรอบ')
        standings_rows.append({'rank': idx, 'team_name': row.get('team_name'), 'score': row.get('score'), 'buchholz': row.get('buchholz'), 'final_buchholz': row.get('final_buchholz'), 'point_for': row.get('point_for'), 'point_against': row.get('point_against'), 'status': status})
    return {'teams': [{'id': t.id, 'name': t.name} for t in teams], 'rounds': [{'round_no': rn, 'matches': rows} for rn, rows in sorted(round_map.items())], 'standings': standings_rows}


def _slot_name(slot):
    if not slot:
        return '-'
    return 'X' if slot.get('is_bye') else (slot.get('team_name') or '-')


def _slot_score(score_map, group_no, slot, stage):
    if not slot:
        return ''
    v = score_map.get((group_no, slot.get('slot_no'), stage))
    return '' if v is None else str(v)


def _slot_score_for_report(score_map, group_no, slot, stage):
    """คะแนนสำหรับรายงาน: ช่องว่างให้นับเป็น 0 ตามที่ใช้ตรวจผลจริง."""
    if not slot:
        return '-'
    if slot.get('is_bye'):
        return '0'
    v = score_map.get((group_no, slot.get('slot_no'), stage))
    return '0' if v is None else str(v)


def _report_match_row(group_no, court, label, a, b, stage, score_map, winner=None):
    sa = _slot_score_for_report(score_map, group_no, a, stage)
    sb = _slot_score_for_report(score_map, group_no, b, stage)
    if a and b and (a.get('is_bye') or b.get('is_bye')):
        if a.get('is_bye') and not b.get('is_bye'):
            sa, sb = '0', '13'
        elif b.get('is_bye') and not a.get('is_bye'):
            sa, sb = '13', '0'
    win_key = None
    if isinstance(winner, dict):
        win_key = winner.get('slot_no') or winner.get('team_id') or winner.get('team_name')
    def is_win(slot):
        if not slot or not win_key:
            return False
        return win_key in {slot.get('slot_no'), slot.get('team_id'), slot.get('team_name')}
    return {
        'group_no': group_no,
        'court': court or '',
        'label': label,
        'team1': _slot_name(a),
        'team2': _slot_name(b),
        'score': f'{sa} - {sb}',
        'result': (winner or {}).get('team_name') if isinstance(winner, dict) else (winner or ''),
        'team1_class': 'report-team-win' if is_win(a) else ('report-team-loss' if winner and b else ''),
        'team2_class': 'report-team-win' if is_win(b) else ('report-team-loss' if winner and a else ''),
    }


def _playoff_round_report_pages(view):
    """รายงานเพลย์ออฟ: 1 รอบต่อ 1 หน้า."""
    pages = []
    if not view:
        return pages
    pairing_label = {'seed': 'ตาม seed', 'random': 'สุ่ม', 'manual': 'แอดมินกำหนด'}
    comp_pairing = pairing_label.get(view.get('competition', {}).get('pairing_method'), view.get('competition', {}).get('pairing_method') or '')
    for rv in view.get('round_views', []):
        rtype = rv['round']['round_type']
        system = 'ดับเบิ้ลน็อคเอ้าท์' if rtype == 'double_knockout' else ('A/B หลัง Swiss' if rtype == 'ab_ladder' else 'น็อคเอ้าท์')
        rows = []
        for g in rv.get('group_views', []):
            slots = {int(s['slot_no']): s for s in g.get('slots', [])}
            score_map = rv.get('score_map', {})
            def add(label, a, b, stage, winner=None):
                if not a and not b:
                    return
                rows.append(_report_match_row(g['group_no'], (a or b or {}).get('court_name') or '', label, a, b, stage, score_map, winner))
            if rtype == 'double_knockout':
                dec = _double_group_decisions(list(slots.values()), score_map)
                for key, label, stage in [('qf1','ผลการแข่งขันครั้งที่ 1: คู่บน',1),('qf2','ผลการแข่งขันครั้งที่ 1: คู่ล่าง',1),('wf','ผลการแข่งขันครั้งที่ 2: ผู้ชนะพบผู้ชนะ',2),('lf','ผลการแข่งขันครั้งที่ 2: ผู้แพ้พบผู้แพ้',2),('final2','ผลการแข่งขันครั้งที่ 3: หาอันดับ 2',3)]:
                    d = dec.get(key)
                    if d:
                        add(label, d.get('a'), d.get('b'), stage, d.get('winner'))
            else:
                d = _decide_playoff_pair(slots.get(1), slots.get(2), 1, score_map)
                add('ผลการแข่งขัน', slots.get(1), slots.get(2), 1, d.get('winner') if d else None)
        pages.append({'round_name': rv['round']['round_name'], 'system': system, 'pairing_method': comp_pairing, 'rows': rows})
    return pages


def _latest_knockout_losers(round_view):
    losers = []
    if not round_view:
        return losers
    for g in round_view.get('group_views', []):
        if round_view['round']['round_type'] == 'knockout':
            slots = {int(s['slot_no']): s for s in g.get('slots', [])}
            d = _decide_playoff_pair(slots.get(1), slots.get(2), 1, round_view.get('score_map', {}))
            if d and d.get('loser') and not d.get('loser').get('is_bye'):
                losers.append(_slot_payload(d.get('loser')))
        elif g['result'].get('second'):
            losers.append(g['result']['second'])
    seen, out = set(), []
    for p in losers:
        key = p.get('team_id') or p.get('team_name')
        if key not in seen:
            seen.add(key); out.append(p)
    return out


def _final_ranking_preview(view):
    """สรุปอันดับ 1-4 สำหรับหน้ารายงาน.
    - รอบชิง 1 คู่: ผู้ชนะ=1 ผู้แพ้=2 และผู้แพ้รอบรอง 2 ทีมเป็นอันดับ 3 ร่วม
    - รอบชิงที่มี 2 คู่: คู่แรกเป็นชิง 1/2, คู่สองเป็นชิง 3/4
    - ดับเบิ้ล: ใช้ผล winner/second ของกลุ่มสุดท้าย
    """
    if not view or not view.get('round_views'):
        return []
    latest = view['round_views'][-1]
    if not _playoff_round_complete(latest):
        return []
    rows = []
    if view.get('competition', {}).get('competition_type') == 'ab_ladder':
        state = _ab_state_for_view(view) or {}
        out = []
        if state.get('direct_qualified'):
            for r in state.get('direct_qualified') or []:
                out.append({'rank': r.get('rank'), 'team_name': r.get('team_name'), 'group_no': r.get('group_no'), 'label': 'ตัวแทนอัตโนมัติ'})
        if state.get('final_a'):
            for r in state.get('final_a') or []:
                out.append({'rank': r.get('rank'), 'team_name': r.get('team_name'), 'group_no': r.get('group_no'), 'label': 'เข้ารอบสาย A'})
        if state.get('final_b'):
            for r in state.get('final_b') or []:
                out.append({'rank': r.get('rank'), 'team_name': r.get('team_name'), 'group_no': r.get('group_no'), 'label': 'เข้ารอบสาย B'})
        return out
    if latest['round']['round_type'] == 'double_knockout':
        for g in latest.get('group_views', []):
            if g['result'].get('winner'):
                rows.append({'rank': 1, 'team_name': g['result']['winner']['team_name']})
            if g['result'].get('second'):
                rows.append({'rank': 2, 'team_name': g['result']['second']['team_name']})
        return rows

    groups = latest.get('group_views', [])
    # กรณีสร้างรอบชิงแบบ "ชิงอันดับ 3" จะมี 2 คู่ในรอบเดียวกัน
    if len(groups) >= 2:
        # คู่ที่ 1 = ชิง 1/2
        g = groups[0]
        slots = {int(s['slot_no']): s for s in g.get('slots', [])}
        d = _decide_playoff_pair(slots.get(1), slots.get(2), 1, latest.get('score_map', {}))
        if d and d.get('winner'):
            rows.append({'rank': 1, 'team_name': d['winner']['team_name']})
        if d and d.get('loser') and not d['loser'].get('is_bye'):
            rows.append({'rank': 2, 'team_name': d['loser']['team_name']})
        # คู่ที่ 2 = ชิงอันดับ 3/4
        g = groups[1]
        slots = {int(s['slot_no']): s for s in g.get('slots', [])}
        d = _decide_playoff_pair(slots.get(1), slots.get(2), 1, latest.get('score_map', {}))
        if d and d.get('winner'):
            rows.append({'rank': 3, 'team_name': d['winner']['team_name']})
        if d and d.get('loser') and not d['loser'].get('is_bye'):
            rows.append({'rank': 4, 'team_name': d['loser']['team_name']})
        return rows

    # รอบชิงปกติ 1 คู่: อันดับ 3 ร่วม = ผู้แพ้รอบรอง 2 ทีม
    if len(groups) == 1:
        g = groups[0]
        slots = {int(s['slot_no']): s for s in g.get('slots', [])}
        d = _decide_playoff_pair(slots.get(1), slots.get(2), 1, latest.get('score_map', {}))
        if d and d.get('winner'):
            rows.append({'rank': 1, 'team_name': d['winner']['team_name']})
        if d and d.get('loser') and not d['loser'].get('is_bye'):
            rows.append({'rank': 2, 'team_name': d['loser']['team_name']})
        if len(view.get('round_views', [])) >= 2:
            semi_losers = _latest_knockout_losers(view['round_views'][-2])[:2]
            for p in semi_losers:
                rows.append({'rank': 3, 'team_name': p.get('team_name')})
    return rows


def _playoff_report_rows(view):
    rows = []
    if not view:
        return rows
    for rv in view.get('round_views', []):
        rtype = rv['round']['round_type']
        system = 'ดับเบิ้ลน็อคเอ้าท์' if rtype == 'double_knockout' else ('A/B หลัง Swiss' if rtype == 'ab_ladder' else 'น็อคเอ้าท์')
        round_name = rv['round']['round_name']
        for g in rv.get('group_views', []):
            slots = {int(s['slot_no']): s for s in g.get('slots', [])}
            score_map = rv.get('score_map', {})
            def nm(slot):
                if not slot:
                    return '-'
                return 'X' if slot.get('is_bye') else (slot.get('team_name') or slot.get('display_name') or '-')
            def add(label, a, b, stage, result=''):
                if not a and not b:
                    return
                winner = None
                if result:
                    for cand in (a, b):
                        if cand and cand.get('team_name') == result:
                            winner = cand
                            break
                row = _report_match_row(g['group_no'], (a or b or {}).get('court_name') or '', label, a, b, stage, score_map, winner)
                row.update({'round_name': round_name, 'system': system})
                rows.append(row)
            if rtype == 'double_knockout':
                dec = _double_group_decisions(list(slots.values()), score_map)
                for key,label,stage in [('qf1','รอบ 1 คู่บน',1),('qf2','รอบ 1 คู่ล่าง',1),('wf','ผู้ชนะเจอผู้ชนะ',2),('lf','ผู้แพ้เจอผู้แพ้',2),('final2','หาอันดับ 2',3)]:
                    d = dec.get(key)
                    if d:
                        add(label, d.get('a'), d.get('b'), stage, (d.get('winner') or {}).get('team_name') if d.get('winner') else '')
            else:
                d = _decide_playoff_pair(slots.get(1), slots.get(2), 1, score_map)
                add('คู่แข่งขัน', slots.get(1), slots.get(2), 1, (d.get('winner') or {}).get('team_name') if d and d.get('winner') else '')
    return rows



# ------------------------- Playoff score sheets / online scorecards -------------------------
def _playoff_scorecard_serializer():
    return URLSafeSerializer(app.config.get("SECRET_KEY", "your_secret_key"), salt="playoff-scorecard-v1")


def _make_playoff_scorecard_token(playoff_id, round_id, group_no, stage_no, a_slot_no, b_slot_no, pair_key=None):
    data = {
        'p': int(playoff_id), 'r': int(round_id), 'g': int(group_no),
        'st': int(stage_no), 'a': int(a_slot_no), 'b': int(b_slot_no),
    }
    if pair_key:
        data['k'] = str(pair_key)
    return _playoff_scorecard_serializer().dumps(data)


def _load_playoff_scorecard_token(token):
    try:
        data = _playoff_scorecard_serializer().loads(token)
        return {
            'playoff_id': _as_int(data.get('p')),
            'round_id': _as_int(data.get('r')),
            'group_no': _as_int(data.get('g')),
            'stage_no': _as_int(data.get('st')),
            'a_slot_no': _as_int(data.get('a')),
            'b_slot_no': _as_int(data.get('b')),
            'pair_key': data.get('k'),
        }
    except BadSignature:
        return None


def _playoff_system_label(round_type):
    if round_type == 'double_knockout':
        return 'ดับเบิ้ลน็อคเอาท์'
    if round_type == 'ab_ladder':
        return 'A/B หลัง Swiss'
    return 'น็อคเอาท์'


def _playoff_stage_label(round_type, stage_no, pair_key=None):
    if round_type == 'ab_ladder':
        return 'ผลการแข่งขัน A/B'
    if round_type == 'double_knockout':
        labels = {
            (1, 'qf1'): 'ครั้งที่ 1: คู่บน',
            (1, 'qf2'): 'ครั้งที่ 1: คู่ล่าง',
            (2, 'wf'): 'ครั้งที่ 2: ผู้ชนะพบผู้ชนะ',
            (2, 'lf'): 'ครั้งที่ 2: ผู้แพ้พบผู้แพ้',
            (3, 'final2'): 'ครั้งที่ 3: หาอันดับ 2',
        }
        return labels.get((stage_no, pair_key), f'ครั้งที่ {stage_no}')
    return 'ผลการแข่งขัน'


def _playoff_slot_dict_by_no(round_id, group_no):
    slots = db.session.execute(text("""
        SELECT * FROM playoff_slots
        WHERE round_id=:rid AND group_no=:g
        ORDER BY slot_no
    """), {'rid': round_id, 'g': group_no}).mappings().all()
    return {int(s['slot_no']): dict(s) for s in slots}


def _playoff_score_map_for_round(round_id):
    rows = db.session.execute(text("SELECT * FROM playoff_scores WHERE round_id=:rid"), {'rid': round_id}).mappings().all()
    return {(int(s['group_no']), int(s['slot_no']), int(s['stage_no'])): s['score'] for s in rows}



def _playoff_placeholder_slot(name):
    return {'slot_no': 0, 'team_name': name, 'is_bye': False, 'is_placeholder': True, 'court_name': ''}


def _playoff_token_pair(playoff_id, rnd, group_no, pair_key, a_slot_no=0, b_slot_no=0):
    """Resolve a printed/QR playoff scorecard pair.
    Future pairs keep the same QR: before the real teams are known it returns blank placeholders;
    after prior matches are scored it resolves to the real winning/losing slots.
    """
    slots = _playoff_slot_dict_by_no(rnd['id'], group_no)
    score_map = _playoff_score_map_for_round(rnd['id'])
    if pair_key == 'qf1':
        return slots.get(1), slots.get(2)
    if pair_key == 'qf2':
        return slots.get(3), slots.get(4)
    if rnd['round_type'] == 'double_knockout':
        dec = _double_group_decisions(list(slots.values()), score_map)
        if pair_key == 'wf':
            return (dec.get('wf') or {}).get('a') or _playoff_placeholder_slot(''), (dec.get('wf') or {}).get('b') or _playoff_placeholder_slot('')
        if pair_key == 'lf':
            return (dec.get('lf') or {}).get('a') or _playoff_placeholder_slot(''), (dec.get('lf') or {}).get('b') or _playoff_placeholder_slot('')
        if pair_key == 'final2':
            return (dec.get('final2') or {}).get('a') or _playoff_placeholder_slot(''), (dec.get('final2') or {}).get('b') or _playoff_placeholder_slot('')
    return slots.get(a_slot_no), slots.get(b_slot_no)


def _playoff_is_real_score_slot(slot):
    return bool(slot and int(slot.get('slot_no') or 0) > 0 and not slot.get('is_placeholder'))


def _playoff_pair_sheet(playoff_id, comp, rnd, this_group_no, pair_key, label, stage_no, a, b, score_map):
    if not a and not b:
        return None
    # ไม่สร้างสกอร์ชีทคู่ X vs X / คู่ว่างทั้งคู่
    if a and b and a.get('is_bye') and b.get('is_bye'):
        return None

    a_slot = int((a or {}).get('slot_no') or 0)
    b_slot = int((b or {}).get('slot_no') or 0)
    can_score_online = _playoff_is_real_score_slot(a) and _playoff_is_real_score_slot(b)
    # ให้มี QR รอไว้ทุกใบ แม้คู่อนาคตยังไม่รู้ชื่อทีมจริง
    token = _make_playoff_scorecard_token(playoff_id, rnd['id'], this_group_no, stage_no, a_slot, b_slot, pair_key)
    scorecard_url = url_for('public_playoff_scorecard', token=token, _external=True)
    court = (a or b or {}).get('court_name') or ''
    return {
        'playoff_id': playoff_id,
        'round_id': rnd['id'],
        'round_name': rnd['round_name'],
        'round_type': rnd['round_type'],
        'system_label': _playoff_system_label(rnd['round_type']),
        'group_no': this_group_no,
        'pair_key': pair_key,
        'stage_no': stage_no,
        'stage_label': label,
        'court_name': court,
        'team1_name': _slot_name(a),
        'team2_name': _slot_name(b),
        'team1_print_name': '' if (a or {}).get('is_placeholder') else _slot_name(a),
        'team2_print_name': '' if (b or {}).get('is_placeholder') else _slot_name(b),
        'team1_slot_no': a_slot,
        'team2_slot_no': b_slot,
        'team1_score': score_map.get((this_group_no, a_slot, stage_no), '') if can_score_online else '',
        'team2_score': score_map.get((this_group_no, b_slot, stage_no), '') if can_score_online else '',
        'token': token,
        'scorecard_url': scorecard_url,
        'can_score_online': can_score_online,
    }


def _playoff_score_sheet_rows(playoff_id, round_id=None, group_no=None):
    view = _fetch_playoff(playoff_id)
    if not view:
        return None, []
    sheets = []
    for rv in view.get('round_views', []):
        rnd = rv['round']
        if round_id and int(rnd['id']) != int(round_id):
            continue
        score_map = rv.get('score_map', {})
        for group in rv.get('group_views', []):
            this_group_no = int(group['group_no'])
            if group_no and this_group_no != int(group_no):
                continue
            slots = {int(s['slot_no']): s for s in group.get('slots', [])}
            if rnd['round_type'] == 'double_knockout':
                dec = _double_group_decisions(list(slots.values()), score_map)

                def ph(name):
                    return _playoff_placeholder_slot(name)

                # ทำสกอร์ชีทให้ครบตามต้นฉบับตั้งแต่ยังไม่กรอกคะแนน:
                # 1) คู่แรก 1-2, 2) คู่แรก 3-4, 3) ผู้ชนะพบผู้ชนะ, 4) ผู้แพ้พบผู้แพ้, 5) ชิงอันดับ 2
                pair_defs = [
                    ('qf1', 1, slots.get(1), slots.get(2)),
                    ('qf2', 1, slots.get(3), slots.get(4)),
                    ('wf', 2, (dec.get('wf') or {}).get('a') or ph('ผู้ชนะคู่ที่ 1'), (dec.get('wf') or {}).get('b') or ph('ผู้ชนะคู่ที่ 2')),
                    ('lf', 2, (dec.get('lf') or {}).get('a') or ph('ผู้แพ้คู่ที่ 1'), (dec.get('lf') or {}).get('b') or ph('ผู้แพ้คู่ที่ 2')),
                    ('final2', 3, (dec.get('final2') or {}).get('a') or ph('ผู้แพ้ผู้ชนะพบผู้ชนะ'), (dec.get('final2') or {}).get('b') or ph('ผู้ชนะผู้แพ้พบผู้แพ้')),
                ]
                for pair_key, stage_no, a, b in pair_defs:
                    if not a or not b:
                        continue
                    sheet = _playoff_pair_sheet(playoff_id, view['competition'], rnd, this_group_no, pair_key, _playoff_stage_label(rnd['round_type'], stage_no, pair_key), stage_no, a, b, score_map)
                    if sheet:
                        sheets.append(sheet)
            else:
                sheet = _playoff_pair_sheet(playoff_id, view['competition'], rnd, this_group_no, 'knockout', _playoff_stage_label(rnd['round_type'], 1), 1, slots.get(1), slots.get(2), score_map)
                if sheet:
                    sheets.append(sheet)
    return view, sheets


def _playoff_match_table_rows(playoff_id, round_id=None, group_no=None):
    """Prepare printable playoff competition tables by round/group.
    One table = one group/สาย, similar to the official blank competition table.
    """
    view = _fetch_playoff(playoff_id)
    if not view:
        return None, []

    tables = []

    for rv in view.get('round_views', []):
        rnd = rv['round']

        if round_id and int(rnd['id']) != int(round_id):
            continue

        round_type = rnd.get('round_type') or ''
        max_stage = 3 if round_type == 'double_knockout' else 1

        for group in rv.get('group_views', []):
            this_group_no = int(group['group_no'])

            if group_no and this_group_no != int(group_no):
                continue

            slots = list(group.get('slots', []))
            if not slots:
                continue

            court_1 = ''
            court_2 = ''

            for slot in slots:
                slot_no = int(slot.get('slot_no') or 0)

                if slot_no in (1, 2) and slot.get('court_name') and not court_1:
                    court_1 = slot.get('court_name') or ''

                if slot_no in (3, 4) and slot.get('court_name') and not court_2:
                    court_2 = slot.get('court_name') or ''

            # ✅ แยกกลุ่ม A/B สำหรับระบบ ab_ladder
            group_label = ''
            if round_type == 'ab_ladder':
                group_label = _ab_group_zone(rv, this_group_no)

            # เลขผู้เข้ารอบบนใบพิมพ์ต้องเดินต่อกันทั้งระบบ
            # - knockout / ab_ladder: 1 สาย = ผู้เข้ารอบ 1 ทีม จึงใช้เลขเดียวกับเลขสายรวม
            # - double_knockout: 1 สาย = ผู้เข้ารอบ 2 ทีม จึงใช้เลขคู่ต่อเนื่อง 1-2, 3-4, ...
            if round_type == 'double_knockout':
                qualified_numbers = [((this_group_no - 1) * 2) + 1, ((this_group_no - 1) * 2) + 2]
            else:
                qualified_numbers = [this_group_no]

            tables.append({
                'round_id': int(rnd['id']),
                'round_name': rnd.get('round_name') or '',
                'round_type': round_type,
                'system_label': _playoff_system_label(round_type),
                # ใช้เลขสายจริงแบบต่อเนื่อง ไม่รีเซ็ตเมื่อเปลี่ยน A/B
                'group_no': this_group_no,

                # ส่งไปให้ playoff_match_tables.html ใช้แสดง กลุ่ม A / กลุ่ม B
                'group_label': group_label,
                # ใช้เลขทีมที่เข้ารอบแบบต่อเนื่อง ไม่ให้กลุ่ม B กลับไปเริ่ม 1
                'qualified_numbers': qualified_numbers,

                'max_stage': max_stage,
                'slots': slots,
                'court_1': court_1,
                'court_2': court_2,
                'result': group.get('result') or {},
            })

    return view, tables


@app.route('/playoff/<int:playoff_id>/match-tables')
@login_required
def playoff_match_tables_print(playoff_id):
    selected_round = request.args.get('round_id', type=int)
    selected_group = request.args.get('group_no', type=int)

    per_page_raw = request.args.get("per_page", "1")
    if per_page_raw == "all":
        per_page = "all"
    else:
        try:
            per_page = int(per_page_raw)
        except (TypeError, ValueError):
            per_page = 1

        if per_page not in [1, 2, 3, 4, 6, 8, 10, 12, 14]:
            per_page = 1

    view, tables = _playoff_match_table_rows(playoff_id, selected_round, selected_group)

    if not view:
        flash('ไม่พบระบบเพลย์ออฟ', 'danger')
        return redirect(url_for('index'))

    if per_page != "all" and any(t.get("round_type") == "double_knockout" for t in tables if t) and per_page > 4:
        per_page = 4

    source_event = Event.query.get(view['competition']['source_event_id']) if view.get('competition') else None

    return render_template(
        'playoff_match_tables.html',
        view=view,
        source_event=source_event,
        tables=tables,
        selected_round=selected_round,
        selected_group=selected_group,
        per_page=per_page,
    )


@app.route('/playoff/<int:playoff_id>/round/<int:round_id>/match-tables')
@login_required
def playoff_round_match_tables_print(playoff_id, round_id):
    return redirect(url_for('playoff_match_tables_print', playoff_id=playoff_id, round_id=round_id, per_page=request.args.get('per_page', 1)))


@app.route('/playoff/<int:playoff_id>/round/<int:round_id>/group/<int:group_no>/match-tables')
@login_required
def playoff_group_match_tables_print(playoff_id, round_id, group_no):
    return redirect(url_for('playoff_match_tables_print', playoff_id=playoff_id, round_id=round_id, group_no=group_no, per_page=request.args.get('per_page', 1)))


@app.route('/playoff/<int:playoff_id>/score-sheet')
@login_required
def playoff_score_sheet(playoff_id):
    selected_round = request.args.get('round_id', type=int)
    selected_group = request.args.get('group_no', type=int)
    # สกอร์ชีทเพลย์ออฟใช้แพทเทิร์น Swiss: 2 ใบต่อแผ่นคงที่ / ไม่แสดงตัวเลือกจำนวนใบต่อแผ่น
    per_page = 2
    view, sheets = _playoff_score_sheet_rows(playoff_id, selected_round, selected_group)
    if not view:
        flash('ไม่พบระบบเพลย์ออฟ', 'danger')
        return redirect(url_for('index'))
    source_event = Event.query.get(view['competition']['source_event_id']) if view.get('competition') else None
    return render_template('playoff_score_sheet.html', view=view, source_event=source_event, sheets=sheets, selected_round=selected_round, selected_group=selected_group, per_page=per_page)


@app.route('/playoff/<int:playoff_id>/round/<int:round_id>/score-sheet')
@login_required
def playoff_round_score_sheet(playoff_id, round_id):
    return redirect(url_for('playoff_score_sheet', playoff_id=playoff_id, round_id=round_id))

@app.route('/playoff/<int:playoff_id>/round/<int:round_id>/group/<int:group_no>/score-sheet')
@login_required
def playoff_group_score_sheet(playoff_id, round_id, group_no):
    return redirect(url_for('playoff_score_sheet', playoff_id=playoff_id, round_id=round_id, group_no=group_no))


@app.route('/playoff-scorecard/<token>', methods=['GET'])
def public_playoff_scorecard(token):
    data = _load_playoff_scorecard_token(token)
    if not data:
        return 'ลิงก์สกอร์การ์ดไม่ถูกต้อง', 404
    context = _playoff_online_context(data, token)
    if not context:
        return 'ไม่พบคู่แข่งขันเพลย์ออฟ', 404
    return render_template('playoff_online_scorecard.html', **context)


@app.route('/playoff-scorecard/<token>/qr.png')
def public_playoff_scorecard_qr(token):
    data = _load_playoff_scorecard_token(token)
    if not data:
        return 'invalid token', 404
    url = url_for('public_playoff_scorecard', token=token, _external=True)
    img = qrcode.make(url)
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return send_file(buf, mimetype='image/png', max_age=0)


def _playoff_online_context(data, token):
    comp = db.session.execute(text("SELECT * FROM playoff_competitions WHERE id=:id"), {'id': data['playoff_id']}).mappings().first()
    rnd = db.session.execute(text("SELECT * FROM playoff_rounds WHERE id=:rid AND playoff_id=:pid"), {'rid': data['round_id'], 'pid': data['playoff_id']}).mappings().first()
    if not comp or not rnd:
        return None
    pair_key = data.get('pair_key')
    a, b = _playoff_token_pair(data['playoff_id'], dict(rnd), data['group_no'], pair_key, data['a_slot_no'], data['b_slot_no'])
    if not a or not b:
        return None
    score_map = _playoff_score_map_for_round(data['round_id'])
    source_event = Event.query.get(comp['source_event_id'])
    a_slot = int((a or {}).get('slot_no') or 0)
    b_slot = int((b or {}).get('slot_no') or 0)
    can_score_online = _playoff_is_real_score_slot(a) and _playoff_is_real_score_slot(b)
    return {
        'token': token,
        'competition': dict(comp),
        'round': dict(rnd),
        'source_event': source_event,
        'group_no': data['group_no'],
        'stage_no': data['stage_no'],
        'stage_label': _playoff_stage_label(rnd['round_type'], data['stage_no'], pair_key),
        'team1': dict(a),
        'team2': dict(b),
        'team1_score': score_map.get((data['group_no'], a_slot, data['stage_no']), '') if can_score_online else '',
        'team2_score': score_map.get((data['group_no'], b_slot, data['stage_no']), '') if can_score_online else '',
        'can_score_online': can_score_online,
        'autosave_url': url_for('public_playoff_scorecard_autosave', token=token),
        'finish_url': url_for('public_playoff_scorecard_finish', token=token),
        'scorecard_public_url': url_for('public_playoff_scorecard', token=token, _external=True),
    }


def _save_playoff_scorecard_values(data, score1_raw, score2_raw):
    try:
        score1 = max(0, min(13, int(score1_raw)))
        score2 = max(0, min(13, int(score2_raw)))
    except Exception:
        return False, 'กรุณากรอกคะแนน 0-13'
    # QR ของคู่อนาคตเปิดรอได้ แต่จะบันทึกคะแนนได้เมื่อระบบทราบชื่อทีมจริงแล้ว
    rnd = db.session.execute(text("SELECT * FROM playoff_rounds WHERE id=:rid AND playoff_id=:pid"), {'rid': data['round_id'], 'pid': data['playoff_id']}).mappings().first()
    if not rnd:
        return False, 'ไม่พบรอบแข่งขัน'
    a, b = _playoff_token_pair(data['playoff_id'], dict(rnd), data['group_no'], data.get('pair_key'), data['a_slot_no'], data['b_slot_no'])
    if not (_playoff_is_real_score_slot(a) and _playoff_is_real_score_slot(b)):
        return False, 'คู่นี้ยังรอทราบทีมจริง ยังบันทึกคะแนนไม่ได้'
    a_slot_no = int(a['slot_no'])
    b_slot_no = int(b['slot_no'])
    if a and a.get('is_bye'):
        score1 = 0
    if b and b.get('is_bye'):
        score2 = 0
    for slot_no, score in ((a_slot_no, score1), (b_slot_no, score2)):
        db.session.execute(text("DELETE FROM playoff_scores WHERE round_id=:rid AND group_no=:g AND slot_no=:s AND stage_no=:st"), {
            'rid': data['round_id'], 'g': data['group_no'], 's': slot_no, 'st': data['stage_no']
        })
        db.session.execute(text("""
            INSERT INTO playoff_scores (round_id, group_no, slot_no, stage_no, score)
            VALUES (:rid, :g, :s, :st, :score)
        """), {'rid': data['round_id'], 'g': data['group_no'], 's': slot_no, 'st': data['stage_no'], 'score': score})
    db.session.commit()
    socketio.emit('playoff_score_updated', {
        'playoff_id': data['playoff_id'], 'round_id': data['round_id'], 'group_no': data['group_no'],
        'slot_no': a_slot_no, 'stage_no': data['stage_no'], 'score': score1,
    }, to=f"playoff_{data['playoff_id']}")
    socketio.emit('playoff_score_updated', {
        'playoff_id': data['playoff_id'], 'round_id': data['round_id'], 'group_no': data['group_no'],
        'slot_no': b_slot_no, 'stage_no': data['stage_no'], 'score': score2,
    }, to=f"playoff_{data['playoff_id']}")
    return True, 'บันทึกคะแนนออนไลน์แล้ว'


@app.route('/playoff-scorecard/<token>/autosave', methods=['POST'])
def public_playoff_scorecard_autosave(token):
    data = _load_playoff_scorecard_token(token)
    if not data:
        return jsonify({'ok': False, 'message': 'ลิงก์ไม่ถูกต้อง'}), 404
    payload = request.get_json(silent=True) or request.form
    ok, msg = _save_playoff_scorecard_values(data, payload.get('team1_score', ''), payload.get('team2_score', ''))
    return jsonify({'ok': ok, 'message': msg}), 200 if ok else 400


@app.route('/playoff-scorecard/<token>/finish', methods=['POST'])
def public_playoff_scorecard_finish(token):
    data = _load_playoff_scorecard_token(token)
    if not data:
        return 'ลิงก์ไม่ถูกต้อง', 404
    ok, msg = _save_playoff_scorecard_values(data, request.form.get('team1_score', ''), request.form.get('team2_score', ''))
    flash(msg, 'success' if ok else 'danger')
    return redirect(url_for('public_playoff_scorecard', token=token))



def _playoff_source_status_map(view, source_team_ids=None):
    """คืนสถานะรายงาน Swiss เดิม: direct = ตัวแทนคนที่ n, playoff teams = จัดแข่งต่อ."""
    config = _safe_json_loads((view.get('competition') or {}).get('config_json'), {}) if view else {}
    status_map = {}
    for idx, row in enumerate(config.get('direct_qualified') or [], start=1):
        tid = row.get('team_id')
        if tid is not None:
            status_map[int(tid)] = f'ตัวแทนคนที่ {idx}'
    for tid in (source_team_ids or []):
        if tid is not None and int(tid) not in status_map:
            status_map[int(tid)] = 'เข้ารอบ'
    return status_map

@app.route('/playoff/<int:playoff_id>')
@login_required
def playoff_detail(playoff_id):
    view = _fetch_playoff(playoff_id)
    if not view:
        flash('ไม่พบระบบแข่งขันต่อ', 'danger')
        return redirect(url_for('index'))
    source_event = Event.query.get(view['competition']['source_event_id'])
    latest = view['round_views'][-1] if view['round_views'] else None
    latest_complete = _playoff_round_complete(latest) if latest else False
    if view['competition'].get('competition_type') == 'ab_ladder' and latest and latest_complete:
        a_rows, b_rows = _ab_next_rows_for_creation(view, latest, view['competition'].get('pairing_method') or 'seed')
        next_participants = a_rows + b_rows
    else:
        next_participants = _participants_from_round(latest) if latest and latest_complete else []
    source_teams = _playoff_source_teams(playoff_id)
    qualified_ids = [t.get('id') for t in source_teams]
    source_status_map = _playoff_source_status_map(view, qualified_ids)
    source_report = _source_event_full_report(source_event.id if source_event else None, qualified_ids, source_status_map)
    latest_losers = _latest_knockout_losers(latest) if latest and latest_complete else []
    return render_template(
        'playoff_detail.html',
        view=view,
        source_event=source_event,
        latest_complete=latest_complete,
        next_participants=next_participants,
        latest_losers=latest_losers,
        source_teams=source_teams,
        source_report=source_report,
        playoff_report_pages=_playoff_round_report_pages(view),
        final_ranking_preview=_final_ranking_preview(view),
        full_report_rows=_playoff_report_rows(view),
        ab_status=_ab_status_for_view(view),
    )


def _playoff_print_report_context(playoff_id, report_type='all'):
    """Build the existing playoff print-report data without rendering a page.

    The single-event report and the multi-event report book both use this helper,
    so report logic stays in one place and the old report remains unchanged.
    """
    view = _fetch_playoff(playoff_id)
    if not view:
        return None
    source_event = Event.query.get(view['competition']['source_event_id'])
    source_teams = _playoff_source_teams(playoff_id)
    qualified_ids = [t.get('id') for t in source_teams]
    source_status_map = _playoff_source_status_map(view, qualified_ids)
    source_report = _source_event_full_report(
        source_event.id if source_event else None,
        qualified_ids,
        source_status_map,
    )
    allowed_report_types = {'all', 'summary', 'standings', 'swiss', 'playoff', 'final'}
    report_type = (report_type or 'all').strip().lower()
    if report_type not in allowed_report_types:
        report_type = 'all'
    return {
        'playoff_id': playoff_id,
        'view': view,
        'source_event': source_event,
        'source_teams': source_teams,
        'source_report': source_report,
        'playoff_report_pages': _playoff_round_report_pages(view),
        'final_ranking_preview': _final_ranking_preview(view),
        'report_type': report_type,
    }


@app.route('/playoff/<int:playoff_id>/print')
@login_required
def playoff_print_report(playoff_id):
    context = _playoff_print_report_context(playoff_id, request.args.get('section') or 'all')
    if not context:
        flash('ไม่พบระบบแข่งขันต่อ', 'danger')
        return redirect(url_for('index'))
    return render_template('playoff_print_report.html', **context)


@app.route('/multi-manage/print-final-reports')
@login_required
@roles_required('superadmin')
def multi_event_print_final_reports():
    """Print one report book containing the latest playoff report of each selected event."""
    selected_event_ids = _selected_event_ids_from_request()
    events = _beta_selected_manage_events(selected_event_ids)
    report_type = (request.args.get('section') or 'all').strip().lower()

    latest_playoffs = _beta_latest_playoffs_for_events(selected_event_ids)
    playoff_by_event = {
        int(row['source_event_id']): int(row['id'])
        for row in latest_playoffs
        if row.get('source_event_id') is not None and row.get('id') is not None
    }

    reports = []
    missing_events = []
    for event in events:
        playoff_id = playoff_by_event.get(int(event.id))
        if not playoff_id:
            missing_events.append(event)
            continue
        context = _playoff_print_report_context(playoff_id, report_type)
        if context:
            reports.append(context)
        else:
            missing_events.append(event)

    if not reports:
        flash('อีเว้นท์ที่เลือกยังไม่มีระบบจัดต่อ/รายงานเพลย์ออฟสำหรับพิมพ์', 'warning')
        return redirect(url_for('multi_event_manager', events=','.join(map(str, selected_event_ids))))

    return render_template(
        'multi_event_print_final_reports.html',
        reports=reports,
        missing_events=missing_events,
        selected_event_ids=selected_event_ids,
        report_type=report_type,
    )


@app.route('/playoff/<int:playoff_id>/round/<int:round_id>/save', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def save_playoff_round(playoff_id, round_id):
    # ปุ่มบันทึกผลมี 2 แบบ:
    # - save_group=<เลขสาย> บันทึกเฉพาะสายนั้น
    # - save_all=1 หรือไม่ส่ง save_group บันทึกทุกสาย
    target_group = _as_int(request.form.get('save_group'), 0)
    slots = db.session.execute(text("SELECT * FROM playoff_slots WHERE round_id = :rid ORDER BY group_no, slot_no"), {'rid': round_id}).mappings().all()
    if target_group:
        slots_to_save = [slot for slot in slots if int(slot['group_no']) == int(target_group)]
    else:
        slots_to_save = list(slots)

    if not slots_to_save:
        flash('ไม่พบสายที่ต้องการบันทึก', 'warning')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id) + f"#round-{round_id}")

    for slot in slots_to_save:
        court_key = f"court_{slot['id']}"
        if court_key in request.form:
            court = request.form.get(court_key, '').strip()
            pair_start = 1 if int(slot['slot_no']) <= 2 else 3
            pair_end = pair_start + 1
            db.session.execute(text("""
                UPDATE playoff_slots SET court_name = :court
                WHERE round_id = :rid AND group_no = :g AND slot_no BETWEEN :a AND :b
            """), {'court': court or None, 'rid': round_id, 'g': slot['group_no'], 'a': pair_start, 'b': pair_end})
        for stage in (1, 2, 3):
            raw = request.form.get(f"score_{slot['id']}_{stage}", '')
            db.session.execute(text("DELETE FROM playoff_scores WHERE round_id=:rid AND group_no=:g AND slot_no=:s AND stage_no=:st"), {'rid': round_id, 'g': slot['group_no'], 's': slot['slot_no'], 'st': stage})
            if raw != '':
                try:
                    val = max(0, min(13, int(raw)))
                except Exception:
                    continue
                db.session.execute(text("""
                    INSERT INTO playoff_scores (round_id, group_no, slot_no, stage_no, score)
                    VALUES (:rid, :g, :s, :st, :score)
                """), {'rid': round_id, 'g': slot['group_no'], 's': slot['slot_no'], 'st': stage, 'score': val})

    group_nos = sorted({slot['group_no'] for slot in slots_to_save})
    for group_no in group_nos:
        w = _as_int(request.form.get(f'manual_winner_{group_no}'), 0)
        sec = _as_int(request.form.get(f'manual_second_{group_no}'), 0)
        db.session.execute(text("DELETE FROM playoff_manual_results WHERE round_id=:rid AND group_no=:g"), {'rid': round_id, 'g': group_no})
        if w:
            db.session.execute(text("""
                INSERT INTO playoff_manual_results (round_id, group_no, winner_slot_no, second_slot_no)
                VALUES (:rid, :g, :w, :s)
            """), {'rid': round_id, 'g': group_no, 'w': w, 's': sec or None})
    db.session.commit()
    socketio.emit('playoff_reload', {'playoff_id': playoff_id}, to=f'playoff_{playoff_id}')
    if target_group:
        flash(f'บันทึกผลสายที่ {target_group} แล้ว', 'success')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id) + f"#round-{round_id}-group-{target_group}")
    flash('บันทึกผลทุกสายในรอบนี้แล้ว', 'success')
    return redirect(url_for('playoff_detail', playoff_id=playoff_id) + f"#round-{round_id}")


@app.route('/playoff/<int:playoff_id>/renumber-courts', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def renumber_playoff_courts(playoff_id):
    round_id = _as_int(request.form.get('round_id'))
    start = max(1, _as_int(request.form.get('start_court'), 1))
    skip = { _as_int(x.strip()) for x in (request.form.get('skip_courts') or '').split(',') if x.strip() }
    slots = db.session.execute(text("SELECT * FROM playoff_slots WHERE round_id=:rid ORDER BY group_no, slot_no"), {'rid': round_id}).mappings().all()
    court = start
    assigned = {}
    for slot in slots:
        pair_key = (slot['group_no'], 1 if slot['slot_no'] <= 2 else 3)
        if pair_key not in assigned:
            while court in skip:
                court += 1
            assigned[pair_key] = str(court); court += 1
        db.session.execute(text("UPDATE playoff_slots SET court_name=:court WHERE id=:id"), {'court': assigned[pair_key], 'id': slot['id']})
    db.session.commit()
    socketio.emit('playoff_reload', {'playoff_id': playoff_id}, to=f'playoff_{playoff_id}')
    flash('จัดเลขสนามอัตโนมัติแล้ว', 'success')
    return redirect(url_for('playoff_detail', playoff_id=playoff_id) + f"#round-{round_id}")





@app.route('/playoff/<int:playoff_id>/live-scores')
@login_required
def playoff_live_scores(playoff_id):
    rows = db.session.execute(text("""
        SELECT ps.round_id, ps.group_no, ps.slot_no, ps.stage_no, ps.score
        FROM playoff_scores ps
        JOIN playoff_rounds pr ON pr.id = ps.round_id
        WHERE pr.playoff_id = :pid
        ORDER BY ps.round_id, ps.group_no, ps.stage_no, ps.slot_no
    """), {'pid': playoff_id}).mappings().all()
    return jsonify({
        'ok': True,
        'playoff_id': playoff_id,
        'scores': [
            {
                'round_id': int(r['round_id']),
                'group_no': int(r['group_no']),
                'slot_no': int(r['slot_no']),
                'stage_no': int(r['stage_no']),
                'score': r['score'] if r['score'] is None else int(r['score']),
            } for r in rows
        ]
    })

@app.route('/playoff/<int:playoff_id>/autosave-score', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def autosave_playoff_score(playoff_id):
    data = request.get_json(silent=True) or request.form
    round_id = _as_int(data.get('round_id'))
    group_no = _as_int(data.get('group_no'))
    slot_no = _as_int(data.get('slot_no'))
    stage_no = _as_int(data.get('stage_no'))
    raw = (str(data.get('score')) if data.get('score') is not None else '').strip()
    db.session.execute(text("DELETE FROM playoff_scores WHERE round_id=:rid AND group_no=:g AND slot_no=:s AND stage_no=:st"), {'rid': round_id, 'g': group_no, 's': slot_no, 'st': stage_no})
    score = None
    if raw != '':
        try:
            score = max(0, min(13, int(raw)))
        except Exception:
            return jsonify({'ok': False, 'message': 'กรอกคะแนน 0-13'}), 400
        db.session.execute(text("""
            INSERT INTO playoff_scores (round_id, group_no, slot_no, stage_no, score)
            VALUES (:rid, :g, :s, :st, :score)
        """), {'rid': round_id, 'g': group_no, 's': slot_no, 'st': stage_no, 'score': score})
    db.session.commit()
    socketio.emit('playoff_score_updated', {
        'playoff_id': playoff_id, 'round_id': round_id, 'group_no': group_no,
        'slot_no': slot_no, 'stage_no': stage_no, 'score': score
    }, to=f'playoff_{playoff_id}')
    return jsonify({'ok': True})


@app.route('/playoff/<int:playoff_id>/autosave-court', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def autosave_playoff_court(playoff_id):
    data = request.get_json(silent=True) or request.form
    slot_id = _as_int(data.get('slot_id'))
    court = (data.get('court') or data.get('court_name') or '').strip()
    slot = db.session.execute(text("SELECT round_id, group_no, slot_no FROM playoff_slots WHERE id=:id"), {'id': slot_id}).mappings().first()
    if not slot:
        return jsonify({'ok': False, 'message': 'ไม่พบช่องสนาม'}), 404
    pair_start = 1 if int(slot['slot_no']) <= 2 else 3
    pair_end = pair_start + 1
    pair_slots = db.session.execute(text("""
        SELECT id FROM playoff_slots
        WHERE round_id=:rid AND group_no=:g AND slot_no BETWEEN :a AND :b
    """), {'rid': slot['round_id'], 'g': slot['group_no'], 'a': pair_start, 'b': pair_end}).mappings().all()
    db.session.execute(text("""
        UPDATE playoff_slots SET court_name=:court
        WHERE round_id=:rid AND group_no=:g AND slot_no BETWEEN :a AND :b
    """), {'court': court or None, 'rid': slot['round_id'], 'g': slot['group_no'], 'a': pair_start, 'b': pair_end})
    db.session.commit()
    for ps in pair_slots:
        socketio.emit('playoff_court_updated', {'playoff_id': playoff_id, 'slot_id': ps['id'], 'court': court}, to=f'playoff_{playoff_id}')
    return jsonify({'ok': True})


@app.route('/playoff/<int:playoff_id>/round/<int:round_id>/delete', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def delete_playoff_round(playoff_id, round_id):
    latest = db.session.execute(text("""
        SELECT id FROM playoff_rounds WHERE playoff_id=:pid ORDER BY round_no DESC, id DESC LIMIT 1
    """), {'pid': playoff_id}).mappings().first()
    if not latest or int(latest['id']) != int(round_id):
        flash('ลบได้เฉพาะรอบล่าสุดเท่านั้น', 'warning')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    db.session.execute(text("DELETE FROM playoff_scores WHERE round_id=:rid"), {'rid': round_id})
    db.session.execute(text("DELETE FROM playoff_manual_results WHERE round_id=:rid"), {'rid': round_id})
    db.session.execute(text("DELETE FROM playoff_slots WHERE round_id=:rid"), {'rid': round_id})
    db.session.execute(text("DELETE FROM playoff_rounds WHERE id=:rid"), {'rid': round_id})
    db.session.commit()
    socketio.emit('playoff_reload', {'playoff_id': playoff_id}, to=f'playoff_{playoff_id}')
    flash('ลบรอบล่าสุดแล้ว', 'success')
    return redirect(url_for('playoff_detail', playoff_id=playoff_id))


@app.route('/playoff/<int:playoff_id>/slot/<int:slot_id>/fill-bye', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def fill_playoff_bye_slot(playoff_id, slot_id):
    team_name = (request.form.get('team_name') or '').strip()
    if not team_name:
        flash('กรุณากรอกชื่อทีม', 'warning')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    db.session.execute(text("UPDATE playoff_slots SET team_name=:name, is_bye=false WHERE id=:id"), {'name': team_name, 'id': slot_id})
    db.session.commit()
    socketio.emit('playoff_reload', {'playoff_id': playoff_id}, to=f'playoff_{playoff_id}')
    flash('เพิ่มทีมแทน X แล้ว', 'success')
    return redirect(url_for('playoff_detail', playoff_id=playoff_id))


@app.route('/playoff/<int:playoff_id>/next-round', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def create_playoff_next_round(playoff_id):
    view = _fetch_playoff(playoff_id)
    if not view or not view['round_views']:
        flash('ไม่พบข้อมูลรอบเดิม', 'danger')
        return redirect(url_for('index'))
    latest = view['round_views'][-1]
    if not _playoff_round_complete(latest):
        flash('กรุณาบันทึกผลให้ครบทุกสายก่อนสร้างรอบต่อไป', 'warning')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    participants = _participants_from_round(latest)
    if len(participants) < 2:
        flash('เหลือผู้ชนะแล้ว ไม่ต้องสร้างรอบต่อไป', 'info')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    next_type = (request.form.get('next_type') or 'knockout').strip()
    if view['competition'].get('competition_type') == 'ab_ladder':
        next_type = 'ab_ladder'
    pairing_method = (request.form.get('pairing_method') or 'seed').strip()
    if pairing_method not in {'seed', 'random', 'manual', 'bracket', 'national_qualifier'}:
        pairing_method = 'seed'
    add_bye = request.form.get('add_bye') == '1'
    user_round_name = (request.form.get('round_name') or '').strip()
    round_name = user_round_name or f"รอบที่ {_next_playoff_round_no(playoff_id)}"
    final_third_mode = (request.form.get('final_third_mode') or 'joint_third').strip()
    if next_type == 'ab_ladder':
        ab_state = _ab_state_for_view(view) or {}
        if ab_state.get('all_finished'):
            flash('ระบบ A/B ถึงจำนวนตามกำหนดแล้ว จบการแข่งขันและใช้รายงานผลได้เลย', 'success')
            return redirect(url_for('playoff_detail', playoff_id=playoff_id) + '#report')
        next_round_no = _next_playoff_round_no(playoff_id)
        try:
            # ใช้ลอจิกคัดตัวแทนทีมชาติได้ 2 กรณี:
            # 1) ถูกเลือกไว้ตั้งแต่สร้างเพลย์ออฟครั้งแรก
            # 2) ผู้ใช้มาเลือก "ประกบคู่คัดตัวแทนทีมชาติ" ในหน้าสร้างรอบถัดไป
            national_mode_now = _regional64_enabled(view) or pairing_method == 'national_qualifier'
            national_size = _national_qualifier_size(view, force=national_mode_now)
            if national_mode_now and national_size not in (48, 64):
                raise ValueError('ประกบคู่คัดตัวแทนทีมชาติใช้ได้กับเพลย์ออฟ 64 ทีม หรือ 48 ทีมเท่านั้น')

            if national_mode_now and not _regional64_round3_created(view) and next_round_no >= 2:
                if national_size == 48:
                    groups, meta = _make_national48_round3_groups(latest)
                    meta['round_description'] = 'คัดตัวแทนทีมชาติ 48 ทีม'
                    if not user_round_name:
                        round_name = 'รอบที่ 3'
                else:
                    groups, meta = _make_regional64_round3_groups(latest)
                    meta['round_description'] = 'คัดตัวแทนทีมชาติ 64 ทีม'
                    if not user_round_name:
                        round_name = 'รอบที่ 3'
                a_rows, b_rows = [], [item for pair in groups for item in pair if item]
            elif national_mode_now and national_size == 48 and not _national48_round4_created(view) and next_round_no >= 3:
                groups, meta = _make_national48_round4_groups(latest)
                a_rows, b_rows = [], [item for pair in groups for item in pair if item]
                meta['round_description'] = 'คัดตัวแทนทีมชาติ 48 ทีม'
                if not user_round_name:
                    round_name = 'รอบที่ 4'
            elif national_mode_now and national_size == 64 and not _regional64_round4_created(view) and next_round_no >= 3:
                groups, meta = _make_regional64_round4_groups(latest)
                a_rows, b_rows = [], [item for pair in groups for item in pair if item]
                meta['round_description'] = 'คัดตัวแทนทีมชาติ 64 ทีม'
                if not user_round_name:
                    round_name = 'รอบที่ 4'
            else:
                a_rows, b_rows = _ab_next_rows_for_creation(view, latest, pairing_method)
                if len(a_rows) + len(b_rows) < 2:
                    flash('ระบบ A/B เหลือทีมไม่พอสร้างรอบต่อไปแล้ว ให้ดูผลเข้ารอบในรายงาน', 'info')
                    return redirect(url_for('playoff_detail', playoff_id=playoff_id) + '#report')
                groups, meta = _ab_make_round_groups(a_rows, b_rows, pairing_method)
                meta.update({'a_team_count': len(a_rows), 'b_team_count': len(b_rows), 'round_kind': 'ab_ladder'})
                meta['round_description'] = f"A/B รอบที่ {next_round_no} — A {len(a_rows)} ทีม / B {len(b_rows)} ทีม"
                if not user_round_name and not round_name:
                    round_name = f"รอบที่ {next_round_no}"
        except ValueError as exc:
            flash(str(exc), 'warning')
            return redirect(url_for('playoff_detail', playoff_id=playoff_id))
        _create_playoff_round(playoff_id, next_round_no, round_name, 'ab_ladder', groups, round_meta=meta)
        db.session.commit()
        socketio.emit('playoff_reload', {'playoff_id': playoff_id}, to=f'playoff_{playoff_id}')
        flash('สร้างรอบ A/B ต่อไปแล้ว', 'success')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))

    if len(participants) == 2:
        next_type = 'knockout'
    if next_type == 'swiss':
        comp = view['competition']
        source_event = Event.query.get(comp['source_event_id'])
        if source_event:
            rows = [{'team_id': p.get('team_id'), 'team_name': p.get('team_name'), 'rank': idx} for idx, p in enumerate(participants, start=1)]
            new_event = _make_next_event_from_selected(source_event, rows, round_name, rounds=max(1, _as_int(request.form.get('swiss_rounds'), 3)))
            flash('สร้าง Swiss ใหม่จากทีมที่ผ่านเข้ารอบแล้ว', 'success')
            return redirect(url_for('event_detail', event_id=new_event.id))
    if next_type == 'round_robin':
        try:
            new_rr_id = _create_round_robin_competition(source_event, participants, round_name, direct_rows=None)
        except Exception as exc:
            flash(f'สร้าง Round Robin ไม่สำเร็จ: {exc}', 'warning')
            return redirect(url_for('playoff_detail', playoff_id=playoff_id))
        flash('สร้าง Round Robin รอบถัดไปแล้ว', 'success')
        return redirect(url_for('playoff_detail', playoff_id=new_rr_id))
    rows = [{'team_id': p.get('team_id'), 'team_name': p.get('team_name'), 'rank': idx, 'seed': idx} for idx, p in enumerate(participants, start=1)]
    if pairing_method == 'manual' and len(participants) > 2:
        session[f'playoff_next_round_manual_{playoff_id}'] = {
            'next_type': next_type,
            'round_name': round_name,
            'add_bye': add_bye,
            'rows': rows,
            'created_at': datetime.utcnow().isoformat(),
        }
        flash('เลือก MANUAL แล้ว กรุณาจิ้มทีมลงคู่/ลงสายเองก่อนสร้างรอบถัดไป', 'info')
        return redirect(url_for('playoff_manual_pairing_next_round', playoff_id=playoff_id))
    if len(participants) == 2 and final_third_mode == 'third_place':
        losers = _latest_knockout_losers(latest)[:2]
        groups = [rows[:2]]
        if len(losers) >= 2:
            bronze_rows = [{'team_id': p.get('team_id'), 'team_name': p.get('team_name'), 'rank': idx + 3, 'seed': idx + 3} for idx, p in enumerate(losers, start=1)]
            groups.append(bronze_rows[:2])
    else:
        try:
            groups = _playoff_double_groups(rows, pairing_method, add_bye=add_bye) if next_type == 'double_knockout' else _playoff_knockout_groups(rows, pairing_method, add_bye=add_bye)
        except ValueError as exc:
            flash(str(exc), 'warning')
            return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    _create_playoff_round(playoff_id, _next_playoff_round_no(playoff_id), round_name, next_type, groups)
    db.session.commit()
    socketio.emit('playoff_reload', {'playoff_id': playoff_id}, to=f'playoff_{playoff_id}')
    flash('สร้างรอบต่อไปแล้ว', 'success')
    return redirect(url_for('playoff_detail', playoff_id=playoff_id))





@app.route('/playoff/<int:playoff_id>/manual-pairing-next', methods=['GET', 'POST'])
@login_required
@roles_required('admin', 'superadmin')
def playoff_manual_pairing_next_round(playoff_id):
    view = _fetch_playoff(playoff_id)
    if not view:
        flash('ไม่พบข้อมูลเพลย์ออฟ', 'danger')
        return redirect(url_for('index'))
    payload = session.get(f'playoff_next_round_manual_{playoff_id}')
    if not payload:
        flash('ไม่พบข้อมูลสร้างรอบแบบ MANUAL กรุณากดสร้างรอบต่อไปใหม่อีกครั้ง', 'warning')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    selected_rows = payload.get('rows') or []
    competition_type = payload.get('next_type') or 'knockout'
    if competition_type not in {'knockout', 'double_knockout'}:
        flash('MANUAL ใช้กับ Knockout หรือ Double knockout เท่านั้น', 'warning')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    try:
        plan = _manual_pairing_plan(selected_rows, competition_type)
    except ValueError as exc:
        flash(str(exc), 'warning')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    teams = [_row_to_seed_payload(row, idx) for idx, row in enumerate(selected_rows, start=1)]
    source_event = Event.query.get(view['competition'].get('source_event_id')) if view['competition'].get('source_event_id') else None
    if request.method == 'POST':
        try:
            groups = _manual_groups_from_request(selected_rows, competition_type)
        except ValueError as exc:
            flash(str(exc), 'warning')
            return render_template('playoff_manual_pairing.html', mode='next', source_event=source_event, view=view, payload=payload, teams=teams, plan=plan)
        _create_playoff_round(playoff_id, _next_playoff_round_no(playoff_id), payload.get('round_name') or 'รอบถัดไป', competition_type, groups)
        db.session.commit()
        session.pop(f'playoff_next_round_manual_{playoff_id}', None)
        socketio.emit('playoff_reload', {'playoff_id': playoff_id}, to=f'playoff_{playoff_id}')
        flash('สร้างรอบต่อไปแบบ MANUAL ตามที่แอดมินจิ้มคู่แล้ว', 'success')
        return redirect(url_for('playoff_detail', playoff_id=playoff_id))
    return render_template('playoff_manual_pairing.html', mode='next', source_event=source_event, view=view, payload=payload, teams=teams, plan=plan)

@app.route('/playoff/<int:playoff_id>/delete', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def delete_playoff_competition(playoff_id):
    comp = db.session.execute(text('SELECT * FROM playoff_competitions WHERE id=:id'), {'id': playoff_id}).mappings().first()
    if not comp:
        flash('ไม่พบเพลย์ออฟ', 'warning')
        return redirect(url_for('index'))
    round_ids = [r['id'] for r in db.session.execute(text('SELECT id FROM playoff_rounds WHERE playoff_id=:pid'), {'pid': playoff_id}).mappings().all()]
    for rid in round_ids:
        db.session.execute(text('DELETE FROM playoff_scores WHERE round_id=:rid'), {'rid': rid})
        db.session.execute(text('DELETE FROM playoff_manual_results WHERE round_id=:rid'), {'rid': rid})
        db.session.execute(text('DELETE FROM playoff_slots WHERE round_id=:rid'), {'rid': rid})
    db.session.execute(text('DELETE FROM playoff_rounds WHERE playoff_id=:pid'), {'pid': playoff_id})
    db.session.execute(text('DELETE FROM playoff_competitions WHERE id=:id'), {'id': playoff_id})
    db.session.commit()
    flash('ลบเพลย์ออฟนี้แล้ว', 'success')
    if comp.get('source_event_id'):
        return redirect(url_for('event_standings', event_id=comp.get('source_event_id')))
    return redirect(url_for('index'))

@app.route('/playoff/<int:playoff_id>/save-report', methods=['POST'])
@login_required
@roles_required('admin', 'superadmin')
def save_playoff_report(playoff_id):
    note = request.form.get('report_note', '')
    db.session.execute(text("UPDATE playoff_competitions SET report_note=:note WHERE id=:id"), {'note': note, 'id': playoff_id})
    db.session.commit()
    flash('บันทึกข้อความรายงานแล้ว', 'success')
    return redirect(url_for('playoff_detail', playoff_id=playoff_id) + '#report')


def calculate_end_time(duration: str):
    now = datetime.utcnow()
    if duration == '1d':
        return now + timedelta(days=1)
    elif duration == '1w':
        return now + timedelta(weeks=1)
    elif duration == '1m':
        return now + timedelta(days=30)
    elif duration == '3m':
        return now + timedelta(days=90)
    elif duration == '6m':
        return now + timedelta(days=180)
    elif duration == '1y':
        return now + timedelta(days=365)
    elif duration == 'forever':
        return None
    else:
        # ถ้าค่าที่ส่งมาไม่ตรงรายการ ให้ตั้ง 1 เดือนแทน ไม่ปล่อยเป็นถาวรโดยไม่ตั้งใจ
        return now + timedelta(days=30)



#--------------------------------------------------------------------------------------

def setup():   
    with app.app_context():
        db.create_all()  # สร้างตารางถ้ายังไม่มี 

        # สร้าง superadmin ถ้ายังไม่มี
        if not User.query.filter_by(username='superadmin').first():
            superadmin = User(username='superadmin', role='superadmin')
            superadmin.set_password('yagami1225')
            db.session.add(superadmin)
            print("Superadmin user created.")

        # สร้าง admin ถ้ายังไม่มี
        if not User.query.filter_by(username='admin').first():
            admin = User(username='admin', role='admin')
            admin.set_password('Admin1234!')
            db.session.add(admin)
            print("Admin user created.")

        db.session.commit()


# -----------------------------------------------------------------------------
# Public Live Report API for Live Report Board
# -----------------------------------------------------------------------------
@app.after_request
def live_report_public_headers(response):
    """Allow the separated Live Report Board to read public JSON and embed public pages."""
    response.headers.setdefault("Access-Control-Allow-Origin", "*")
    response.headers.setdefault("Access-Control-Allow-Headers", "Content-Type, Authorization")
    response.headers.setdefault("Access-Control-Allow-Methods", "GET, OPTIONS")
    # ให้ Report Board / จอทีวีฝังหน้า public ได้ ถ้าในอนาคตต้องใช้ iframe
    response.headers.pop("X-Frame-Options", None)
    response.headers.setdefault("Content-Security-Policy", "frame-ancestors *")
    return response


def _lr_team_payload(team):
    if not team:
        return None
    return {"id": team.id, "name": team.name}


def _lr_match_payload(match):
    score1 = match.pending_team1_score if match.pending_team1_score is not None else match.team1_score
    score2 = match.pending_team2_score if match.pending_team2_score is not None else match.team2_score
    if match.team2_id is None:
        status = "bye"
    elif match.is_locked:
        status = "finished"
    elif match.pending_team1_score is not None or match.pending_team2_score is not None:
        status = "playing"
    else:
        status = "waiting"
    leader = "draw"
    try:
        if int(score1 or 0) > int(score2 or 0):
            leader = "team1"
        elif int(score2 or 0) > int(score1 or 0):
            leader = "team2"
    except Exception:
        pass
    return {
        "id": match.id,
        "event_id": match.event_id,
        "round": match.round,
        "round_label": f"รอบแรก ครั้งที่ {match.round}",
        "field": match.field,
        "team1": match.team1.name if match.team1 else "",
        "team2": match.team2.name if match.team2 else "BYE",
        "team1_obj": _lr_team_payload(match.team1),
        "team2_obj": _lr_team_payload(match.team2),
        "score1": score1 or 0,
        "score2": score2 or 0,
        "official_score1": match.team1_score or 0,
        "official_score2": match.team2_score or 0,
        "pending_score1": match.pending_team1_score,
        "pending_score2": match.pending_team2_score,
        "has_pending": match.pending_team1_score is not None or match.pending_team2_score is not None,
        "status": status,
        "leader": leader,
        "is_locked": bool(match.is_locked),
    }


def _lr_event_payload(event):
    return {
        "id": event.id,
        "name": event.name,
        "location": event.location,
        "category": event.category,
        "sex": event.sex,
        "age_group": event.age_group,
        "rounds": event.rounds,
        "current_round": event.current_round,
        "date": event.date.isoformat() if event.date else None,
        "team_count": Team.query.filter_by(event_id=event.id).count(),
        "match_count": Match.query.filter_by(event_id=event.id).count(),
    }


@app.route('/api/public/events')
def api_public_events():
    events = Event.query.order_by(Event.created_at.desc(), Event.id.desc()).all()
    return jsonify({"ok": True, "events": [_lr_event_payload(e) for e in events]})


@app.route('/api/public/event/<int:event_id>/live')
def api_public_event_live(event_id):
    event = Event.query.get_or_404(event_id)
    round_no = request.args.get('round', type=int)
    query = Match.query.filter_by(event_id=event.id)
    if round_no:
        query = query.filter_by(round=round_no)
    matches = query.order_by(Match.round.asc(), Match.field.asc().nullslast(), Match.id.asc()).all()
    rounds = sorted({m.round for m in Match.query.filter_by(event_id=event.id).all() if m.round})
    live_matches = [m for m in matches if not m.is_locked and m.team2_id is not None]
    finished_matches = [m for m in matches if m.is_locked]
    return jsonify({
        "ok": True,
        "source": "tournoi",
        "event": _lr_event_payload(event),
        "rounds": rounds,
        "current_round": event.current_round,
        "matches": [_lr_match_payload(m) for m in matches],
        "live_matches": [_lr_match_payload(m) for m in live_matches],
        "latest_results": [_lr_match_payload(m) for m in finished_matches[-20:]],
    })


@app.route('/api/public/event/<int:event_id>/standings')
def api_public_event_standings(event_id):
    event = Event.query.get_or_404(event_id)
    rows = calculate_standings(event.id)
    standings = []
    for idx, row in enumerate(rows, start=1):
        standings.append({
            "rank": idx,
            "team_id": row.get("team_id"),
            "team": row.get("team_name"),
            "win": row.get("score", 0),
            "played": row.get("played", 0),
            "point_for": row.get("point_for", 0),
            "point_against": row.get("point_against", 0),
            "diff": row.get("point_diff", 0),
            "buchholz": row.get("buchholz", 0),
            "final_buchholz": row.get("final_buchholz", 0),
            "raw": row,
        })
    return jsonify({"ok": True, "source": "tournoi", "event": _lr_event_payload(event), "standings": standings})


@app.route('/api/public/event/<int:event_id>/playoffs')
def api_public_event_playoffs(event_id):
    event = Event.query.get_or_404(event_id)
    playoffs = []
    try:
        for p in _active_playoffs_for_event(event.id):
            view = _fetch_playoff(p.get('id'))
            playoffs.append({"id": p.get('id'), "title": p.get('title'), "competition_type": p.get('competition_type'), "view": view})
    except Exception as exc:
        return jsonify({"ok": False, "source": "tournoi", "event": _lr_event_payload(event), "error": str(exc), "playoffs": []}), 200
    return jsonify({"ok": True, "source": "tournoi", "event": _lr_event_payload(event), "playoffs": playoffs})


@app.route('/api/public/event/<int:event_id>/report')
def api_public_event_report(event_id):
    """One-call endpoint for Report Board."""
    event = Event.query.get_or_404(event_id)
    matches = Match.query.filter_by(event_id=event.id).order_by(Match.round.asc(), Match.field.asc().nullslast(), Match.id.asc()).all()
    standings = []
    for idx, row in enumerate(calculate_standings(event.id), start=1):
        standings.append({"rank": idx, "team": row.get("team_name"), "win": row.get("score", 0), "diff": row.get("point_diff", 0), "raw": row})
    return jsonify({
        "ok": True,
        "source": "tournoi",
        "event": _lr_event_payload(event),
        "matches": [_lr_match_payload(m) for m in matches],
        "live_matches": [_lr_match_payload(m) for m in matches if not m.is_locked and m.team2_id is not None],
        "latest_results": [_lr_match_payload(m) for m in matches if m.is_locked][-20:],
        "standings": standings[:32],
    })


# ---------- BETA: multi-event manager ----------
def _beta_field_labels(prefix='', start=1, field_max=27, exclude_text='', count=0):
    """สร้างเลขสนามตามลำดับ ใช้กับหน้าบริหารหลายอีเว้นท์"""
    prefix = prefix or ''
    try:
        start = int(start or 1)
    except Exception:
        start = 1
    try:
        field_max = int(field_max or 27)
    except Exception:
        field_max = 27
    exclude = set(x.strip() for x in (exclude_text or '').split(',') if x.strip())
    labels = []
    n = start
    guard = 0
    while len(labels) < count and guard < max(200, field_max * 20):
        label = f"{prefix}{n}"
        if label not in exclude:
            labels.append(label)
        n += 1
        guard += 1
    return labels


def _beta_field_pool(prefix='', start=1, field_max=27, exclude_text=''):
    """คืนรายการสนามที่ใช้ได้ เช่น 1..27 โดยตัดสนามที่เว้นออก ใช้สำหรับวนจัดสนาม"""
    prefix = prefix or ''
    try:
        start = int(start or 1)
    except Exception:
        start = 1
    try:
        field_max = int(field_max or 27)
    except Exception:
        field_max = 27
    exclude = set(x.strip() for x in (exclude_text or '').split(',') if x.strip())
    labels = []
    for n in range(start, start + max(field_max, 1)):
        label = f"{prefix}{n}"
        if label not in exclude and str(n) not in exclude:
            labels.append(label)
    return labels or [str(start)]


def _beta_match_team_ids(match):
    ids = []
    if getattr(match, 'team1_id', None):
        ids.append(match.team1_id)
    if getattr(match, 'team2_id', None):
        ids.append(match.team2_id)
    return ids


def _beta_is_bye_match(match):
    """คู่ BYE ไม่ต้องใช้สนาม เพราะไม่มีการแข่งขันจริงในสนาม"""
    return bool(getattr(match, 'team1_id', None) and not getattr(match, 'team2_id', None))


def _beta_assignable_match_count(matches):
    """นับเฉพาะคู่ที่มี 2 ทีมจริง ต้องลงสนามแข่ง"""
    return sum(1 for m in (matches or []) if not _beta_is_bye_match(m))


def _beta_team_used_fields(event_id, exclude_match_ids=None):
    """สร้าง map team_id -> set(สนามที่เคยเล่น) เพื่อหลีกเลี่ยงซ้ำสนามเดิมในรอบต่อไป"""
    exclude_match_ids = set(exclude_match_ids or [])
    used = {}
    qs = Match.query.filter_by(event_id=event_id).all()
    for m in qs:
        if m.id in exclude_match_ids or m.field in (None, ''):
            continue
        field_label = str(m.field)
        for tid in _beta_match_team_ids(m):
            used.setdefault(tid, set()).add(field_label)
    return used


def _beta_assign_fields_to_matches(matches, prefix='', start=1, field_max=27, exclude_text='', continue_numbering=True, avoid_team_repeat=True):
    """กำหนดเลขสนามให้ match list ของหลายอีเว้นท์พร้อมกัน
    - สนามหนึ่งใช้ได้แค่ 1 คู่ต่อครั้งที่เลือก เพื่อกันหลายอีเว้นท์ชนสนามกัน
    - ถ้าเป็นไปได้ จะไม่ให้ทีมกลับไปเล่นสนามที่เคยเล่นในอีเว้นท์เดิม
    - ถ้าจำนวนคู่มากกว่าจำนวนสนาม จะปล่อยคู่ที่เกินเป็นค่าว่าง/ยังไม่เลือก ไม่วนซ้ำสนาม
    """
    if not matches:
        return 0
    labels = _beta_field_pool(prefix, start, field_max, exclude_text)
    assigned = 0
    used_global = set()

    # ล้างสนามเดิมของชุดที่เลือกก่อน แล้วค่อยจัดใหม่แบบรวมทุกอีเว้นท์
    current_ids = {m.id for m in matches if getattr(m, 'id', None)}
    team_used_cache = {}

    for match in matches:
        # ล้างสนามทุกคู่ก่อน รวมถึงคู่ BYE ที่เคยถูกใส่สนามจากเวอร์ชันเก่า
        match.field = None

    for match in matches:
        # BYE ไม่มีการแข่งขันจริง จึงไม่ต้องจองเลขสนาม และไม่ให้นับเป็นสนามที่ยังไม่เลือก
        if _beta_is_bye_match(match):
            match.field = None
            continue

        available_labels = [label for label in labels if label not in used_global]
        if not available_labels:
            # สนามไม่พอสำหรับแข่งพร้อมกัน: ไม่ใส่เลขสนาม เพื่อให้ขึ้นว่า "ยังไม่เลือก"
            match.field = None
            continue

        event_id = match.event_id
        if event_id not in team_used_cache:
            team_used_cache[event_id] = _beta_team_used_fields(event_id, exclude_match_ids=current_ids) if avoid_team_repeat else {}
        used_by_team = team_used_cache[event_id]
        team_ids = _beta_match_team_ids(match)

        chosen = None
        if avoid_team_repeat and team_ids:
            candidates = []
            for label in available_labels:
                conflict = any(label in used_by_team.get(tid, set()) for tid in team_ids)
                if not conflict:
                    candidates.append(label)
            if candidates:
                chosen = candidates[0]

        if chosen is None:
            chosen = available_labels[0]

        match.field = _beta_db_field_value(chosen)
        used_global.add(str(chosen))
        for tid in team_ids:
            used_by_team.setdefault(tid, set()).add(str(chosen))
        assigned += 1


    return assigned


def _beta_event_field_start_map_from_request(events, default_start=1):
    """อ่านช่องกำหนดสนามเริ่มรายอีเว้นท์จากหน้า multi-manage
    ถ้าไม่กรอก จะให้ระบบต่อเลขสนามอัตโนมัติจากอีเว้นท์ก่อนหน้า
    """
    mapping = {}
    raw_param = request.values.get('event_field_starts') or ''
    for item in raw_param.split(','):
        if ':' not in item:
            continue
        eid_s, st_s = item.split(':', 1)
        try:
            value = int(float(st_s))
            if value > 0:
                mapping[int(eid_s)] = value
        except Exception:
            pass
    for event in events or []:
        raw = request.values.get(f'event_field_start_{event.id}')
        if raw is None or str(raw).strip() == '':
            continue
        try:
            value = int(float(str(raw).strip()))
            if value > 0:
                mapping[event.id] = value
        except Exception:
            pass
    return mapping


def _beta_event_field_starts_param(start_map):
    return ','.join(f'{int(eid)}:{int(st)}' for eid, st in (start_map or {}).items())


def _beta_assign_fields_by_event_groups(event_match_groups, prefix='', start=1, field_max=27, exclude_text='', event_start_map=None):
    """จัดสนามแบบแบ่งเป็นบล็อกต่ออีเว้นท์
    - อีเว้นท์หนึ่งจะได้เลขสนามเรียงติดกัน ไม่กระโดดสลับกับอีเว้นท์อื่น
    - ถ้ากรอกสนามเริ่มรายอีเว้นท์ จะเริ่มตามนั้น
    - ถ้าไม่กรอก จะต่อเลขอัตโนมัติจากอีเว้นท์ก่อนหน้า
    - คู่ BYE ไม่กินเลขสนาม
    """
    event_start_map = event_start_map or {}
    try:
        start = int(start or 1)
    except Exception:
        start = 1
    try:
        field_max = int(field_max or 27)
    except Exception:
        field_max = 27
    prefix = prefix or ''
    exclude = set(x.strip() for x in (exclude_text or '').split(',') if x.strip())
    min_num = start
    max_num = start + max(field_max, 1) - 1
    used_global = set()
    assigned = 0
    next_auto = start

    def make_label(n):
        return f"{prefix}{n}"

    def is_allowed(n):
        label = make_label(n)
        return n <= max_num and label not in exclude and str(n) not in exclude and str(label) not in used_global

    for event, matches in event_match_groups:
        matches = list(matches or [])
        for match in matches:
            match.field = None

        current = int(event_start_map.get(getattr(event, 'id', None), next_auto) or next_auto)
        if current < min_num:
            current = min_num

        for match in matches:
            if _beta_is_bye_match(match):
                match.field = None
                continue
            n = current
            while n <= max_num and not is_allowed(n):
                n += 1
            if n > max_num:
                match.field = None
                current = n
                continue
            label = make_label(n)
            match.field = _beta_db_field_value(label)
            used_global.add(str(label))
            assigned += 1
            current = n + 1

        # ถ้าไม่ได้ระบุสนามเริ่มของอีเว้นท์ถัดไป ให้ต่อจากเลขสุดท้ายของอีเว้นท์นี้
        if getattr(event, 'id', None) not in event_start_map:
            next_auto = max(next_auto, current)
        else:
            # กรณีระบุเอง ให้ระบบอัตโนมัติตัวถัดไปต่อจากช่วงที่ใช้จริงด้วย เพื่อกันชนกัน
            next_auto = max(next_auto, current)

    return assigned


def _beta_match_field_repeat_count(event_id, round_no):
    """นับคู่ในรอบที่เลือกที่มีทีมเคยเล่นสนามนี้มาก่อนแล้ว ใช้โชว์เตือนในหน้าบริหารรวม"""
    matches = Match.query.filter_by(event_id=event_id, round=round_no).all()
    if not matches:
        return 0
    current_ids = {m.id for m in matches}
    before_used = _beta_team_used_fields(event_id, exclude_match_ids=current_ids)
    repeat = 0
    for m in matches:
        if m.field in (None, ''):
            continue
        label = str(m.field)
        if any(label in before_used.get(tid, set()) for tid in _beta_match_team_ids(m)):
            repeat += 1
    return repeat


def _beta_event_round_field_text(event_id, round_no):
    fields = [str(m.field) for m in Match.query.filter_by(event_id=event_id, round=round_no).all() if m.field not in (None, '')]
    if not fields:
        return '-'
    def sort_key(x):
        try:
            return (0, int(x))
        except Exception:
            return (1, x)
    fields = sorted(set(fields), key=sort_key)
    if len(fields) <= 8:
        return ', '.join(fields)
    return f"{fields[0]}-{fields[-1]} ({len(fields)} สนาม)"


def _beta_db_field_value(label):
    """แปลงเลขสนามก่อนบันทึกลง DB
    field ในฐานข้อมูลเดิมเป็น Integer หลายเครื่อง/หลาย DB โดยเฉพาะ PostgreSQL จะ Error ถ้าส่งค่าว่างหรือข้อความเข้าไป
    ดังนั้นถ้าเป็นเลขล้วนให้เก็บเป็น int ถ้าไม่ใช่เลขล้วนค่อยเก็บเป็นข้อความสำหรับ DB ที่รองรับ
    """
    if label in (None, ''):
        return None
    label_s = str(label).strip()
    if label_s.isdigit():
        return int(label_s)
    return label_s


def _beta_event_round_unassigned_count(event_id, round_no):
    """นับเฉพาะคู่ที่ต้องแข่งจริงแต่ยังไม่มีเลขสนาม; คู่ BYE ไม่ต้องใส่สนาม
    ห้าม query เทียบ Match.field == '' เพราะ PostgreSQL จะ Internal Server Error เมื่อ field เป็น Integer
    """
    matches = Match.query.filter(
        Match.event_id == event_id,
        Match.round == round_no,
        Match.team1_id != None,
        Match.team2_id != None,
    ).all()
    return sum(1 for m in matches if getattr(m, 'field', None) in (None, ''))


def _beta_pair_first_round(event, separate_same_name=False):
    """จับคู่รอบแรกให้ 1 อีเว้นท์ คืน (ok, message, count) ไม่ redirect"""
    existing_matches = Match.query.filter_by(event_id=event.id, round=1).count()
    if existing_matches > 0:
        return False, "มีคู่รอบแรกอยู่แล้ว", existing_matches

    teams = Team.query.filter_by(event_id=event.id).all()
    if len(teams) < 2:
        return False, "ทีมไม่พอสำหรับจับคู่", 0

    def extract_base_name(name):
        return re.split(r'[\s\-]*\d+$', (name or '').strip())[0]

    teams = teams[:]
    random.shuffle(teams)
    pairings = []
    used_ids = set()

    for i, team1 in enumerate(teams):
        if team1.id in used_ids:
            continue
        for j in range(i + 1, len(teams)):
            team2 = teams[j]
            if team2.id in used_ids:
                continue
            if separate_same_name and extract_base_name(team1.name) == extract_base_name(team2.name):
                continue
            pairings.append((team1, team2))
            used_ids.add(team1.id)
            used_ids.add(team2.id)
            break

    remaining = [t for t in teams if t.id not in used_ids]
    if remaining:
        pairings.append((remaining[0], None))

    team_lookup = {team.id: team for team in teams}
    created = 0
    for team1, team2 in pairings:
        team1_id, team2_id = _order_pair_by_alphabet(team1.id, team2.id if team2 else None, team_lookup)
        match = Match(event_id=event.id, round=1, team1_id=team1_id, team2_id=team2_id, is_locked=False)
        if team2_id is None:
            match.team1_score = 1
            match.team2_score = 0
            match.is_locked = True
        db.session.add(match)
        created += 1
    return True, "จับคู่รอบแรกแล้ว", created




def _beta_pair_selected_round(event, target_round=1, separate_same_name=False):
    """จับคู่ครั้งที่เลือกให้ 1 อีเว้นท์แบบปลอดภัย
    - ครั้งที่ 1 ใช้ตัวจับรอบแรกเดิมของหน้า beta
    - ครั้งที่ 2+ ใช้ swiss_pairing เดิมของระบบ
    - ถ้ารอบเป้าหมายมีคู่แล้ว จะข้ามเพื่อกันลบงานเดิม
    คืน (ok, message, count)
    """
    try:
        target_round = int(target_round or 1)
    except Exception:
        target_round = 1
    if target_round <= 1:
        return _beta_pair_first_round(event, separate_same_name=separate_same_name)

    existing_target = Match.query.filter_by(event_id=event.id, round=target_round).count()
    if existing_target > 0:
        return False, f"มีคู่ครั้งที่ {target_round} อยู่แล้ว", existing_target

    previous_round = target_round - 1
    previous_count = Match.query.filter_by(event_id=event.id, round=previous_round).count()
    if previous_count <= 0:
        return False, f"ยังไม่มีคู่ครั้งที่ {previous_round}", 0

    unlocked_count = Match.query.filter_by(event_id=event.id, round=previous_round, is_locked=False).count()
    if unlocked_count > 0:
        return False, f"ครั้งที่ {previous_round} ยังลงคะแนน/ล็อกผลไม่ครบ เหลือ {unlocked_count} คู่", 0

    if getattr(event, 'rounds', None) and target_round > int(event.rounds or 0):
        return False, f"เกินจำนวนรอบที่ตั้งไว้ ({event.rounds})", 0

    ok, msg = swiss_pairing(event.id, target_round, separate_same_name=separate_same_name)
    if not ok:
        return False, msg, 0

    created = Match.query.filter_by(event_id=event.id, round=target_round).count()
    return True, f"จับคู่ครั้งที่ {target_round} แล้ว", created



def _beta_event_rounds_map_from_request(events, default_round=1):
    """อ่านรอบต่ออีเว้นท์จาก form/query เพื่อให้หน้ารวมจัดการหลายอีเว้นท์ที่อยู่คนละรอบได้"""
    try:
        default_round = int(default_round or 1)
    except Exception:
        default_round = 1
    default_round = max(1, min(default_round, 4))
    mapping = {}

    raw = request.values.get('event_rounds') or ''
    for item in raw.split(','):
        if ':' not in item:
            continue
        eid_s, rnd_s = item.split(':', 1)
        try:
            mapping[int(eid_s)] = max(1, min(int(rnd_s), 4))
        except Exception:
            pass

    for event in events or []:
        key = f'event_round_{event.id}'
        val = request.values.get(key, type=int)
        if val:
            mapping[event.id] = max(1, min(int(val), 4))
        mapping.setdefault(event.id, default_round)
    return mapping


def _beta_event_rounds_param(round_map):
    return ','.join(f'{int(eid)}:{int(rnd)}' for eid, rnd in (round_map or {}).items())


def _beta_safe_int_field_order_value(value):
    try:
        return int(value)
    except Exception:
        return 999999


def _beta_playoff_count_for_event(event_id):
    try:
        row = db.session.execute(text('SELECT COUNT(*) AS c FROM playoff_competitions WHERE source_event_id=:eid'), {'eid': event_id}).mappings().first()
        return int((row or {}).get('c') or 0)
    except Exception:
        return 0


def _beta_first_playoff_id_for_event(event_id):
    try:
        row = db.session.execute(text('SELECT id FROM playoff_competitions WHERE source_event_id=:eid ORDER BY id DESC LIMIT 1'), {'eid': event_id}).mappings().first()
        return int(row['id']) if row else None
    except Exception:
        return None


def _beta_event_all_matches_locked(event_id):
    total = Match.query.filter_by(event_id=event_id).count()
    if total <= 0:
        return False, 'ยังไม่มีคู่แข่งขัน'
    unlocked = Match.query.filter_by(event_id=event_id, is_locked=False).count()
    if unlocked > 0:
        return False, f'ยังมีคู่ที่ยังไม่ล็อกผล {unlocked} คู่'
    return True, ''


def _beta_int_form(name, default=0, minimum=0):
    try:
        val = int(request.form.get(name, default) or default)
    except Exception:
        val = default
    return max(minimum, val)


def _beta_multi_create_next_for_event(event):
    """สร้างระบบจัดต่อจาก Standing สำหรับ 1 อีเว้นท์ตามค่าที่กรอกในแถว multi-manage"""
    eid = event.id
    allow_duplicate = request.form.get(f'event_allow_duplicate_next_{eid}') == 'on'
    existing = _beta_playoff_count_for_event(eid)
    if existing and not allow_duplicate:
        return False, f'มีเพลย์ออฟ/รอบต่ออยู่แล้ว {existing} รายการ', None

    ok_locked, locked_msg = _beta_event_all_matches_locked(eid)
    if not ok_locked:
        return False, locked_msg, None

    standings_rows = calculate_standings(eid)
    if not standings_rows:
        return False, 'ยังไม่มี Standing', None

    direct_count = _beta_int_form(f'event_direct_count_{eid}', 0, 0)
    continue_count = _beta_int_form(f'event_continue_count_{eid}', 8, 0)
    ab_a_count = _beta_int_form(f'event_ab_a_count_{eid}', max(1, continue_count // 2) if continue_count else 0, 0)
    swiss_rounds = _beta_int_form(f'event_swiss_rounds_{eid}', 3, 1)

    competition_type = (request.form.get(f'event_competition_type_{eid}') or 'knockout').strip()
    if competition_type not in {'knockout', 'double_knockout', 'swiss', 'ab_ladder', 'regional64_ladder', 'round_robin'}:
        competition_type = 'knockout'
    pairing_method = (request.form.get(f'event_pairing_method_{eid}') or 'seed').strip()
    if pairing_method not in {'seed', 'random', 'bracket', 'national_qualifier'}:
        pairing_method = 'seed'
    if competition_type == 'regional64_ladder':
        pairing_method = 'national_qualifier'
    add_bye = request.form.get(f'event_add_bye_{eid}') == 'on'
    include_direct = request.form.get(f'event_include_direct_{eid}') == 'on'
    next_stage_name = (request.form.get(f'event_next_stage_name_{eid}') or '').strip() or 'รอบต่อไป'

    total_rows = len(standings_rows)
    direct_count = max(0, min(direct_count, total_rows))
    continue_count = max(0, min(continue_count, total_rows - direct_count))

    direct_rows = [dict(r, rank=i+1) for i, r in enumerate(standings_rows[:direct_count])]
    selected_rows = [dict(r, rank=direct_count+i+1) for i, r in enumerate(standings_rows[direct_count:direct_count+continue_count])]

    if include_direct and direct_rows:
        merged = []
        seen = set()
        for row in direct_rows + selected_rows:
            key = row.get('team_id') or row.get('team_name')
            if key in seen:
                continue
            seen.add(key)
            merged.append(row)
        selected_rows = merged
        direct_rows_for_report = []
    else:
        direct_rows_for_report = direct_rows

    if not selected_rows:
        return False, 'ไม่มีทีมสำหรับจัดต่อ กรุณาตั้งอันดับถัดไปมากกว่า 0', None

    try:
        if competition_type == 'round_robin':
            playoff_id = _create_round_robin_competition(event, selected_rows, next_stage_name, direct_rows=direct_rows_for_report)
            return True, f'สร้าง Round Robin #{playoff_id}', ('playoff', playoff_id)
        if competition_type == 'swiss':
            new_event = _make_next_event_from_selected(event, selected_rows, next_stage_name, rounds=swiss_rounds)
            return True, f'สร้าง Swiss ใหม่ #{new_event.id}', ('event', new_event.id)
        if competition_type in {'ab_ladder', 'regional64_ladder'}:
            if competition_type == 'regional64_ladder':
                playoff_id = _create_ab_ladder_competition(event, selected_rows, next_stage_name, 'national_qualifier', 32, 8, 12, direct_rows=direct_rows_for_report, special_mode='regional64_ladder')
            else:
                playoff_id = _create_ab_ladder_competition(event, selected_rows, next_stage_name, pairing_method, ab_a_count or max(1, len(selected_rows)//2), 0, 0, direct_rows=direct_rows_for_report)
            return True, f'สร้าง A/B #{playoff_id}', ('playoff', playoff_id)
        playoff_id = _create_playoff_competition(event, selected_rows, next_stage_name, competition_type, pairing_method, add_bye=add_bye, direct_rows=direct_rows_for_report)
        return True, f'สร้างเพลย์ออฟ #{playoff_id}', ('playoff', playoff_id)
    except Exception as exc:
        db.session.rollback()
        return False, str(exc), None


def _beta_selected_manage_events(selected_event_ids):
    if not selected_event_ids:
        return []
    events = Event.query.filter(Event.id.in_(selected_event_ids)).all()
    events.sort(key=lambda e: (e.date or date.min, e.name or '', e.id))
    return events


@app.route('/multi-manage', methods=['GET', 'POST'])
@login_required
@roles_required('superadmin')
def multi_event_manager():
    """บริหารจัดการหลายอีเว้นท์: เลือกอีเว้นท์ จับคู่ ใส่สนาม ลบคู่ และพิมพ์รวม
    รองรับกรณีที่แต่ละอีเว้นท์อยู่คนละรอบ โดยเลือกครั้ง/รอบแยกต่ออีเว้นท์ได้
    """
    today = date.today()
    selected_round = request.values.get('round', type=int) or 1
    selected_round = max(1, min(int(selected_round or 1), 4))
    selected_event_ids = _selected_event_ids_from_request()
    field_start = request.values.get('field_start', type=int) or 1
    field_max = request.values.get('field_max', type=int) or 27
    field_prefix = (request.values.get('field_prefix') or '').strip()
    field_exclude = (request.values.get('field_exclude') or '').strip()
    separate_same_name = request.values.get('separate_same_name') == 'on'
    force_delete_locked = request.values.get('force_delete_locked') == 'on'
    event_field_start_map = _beta_event_field_start_map_from_request([], field_start)

    # ใช้ order_by แบบปลอดภัยกับ SQLite/Railway หลายเวอร์ชัน ไม่ใช้ nullslast() กัน Internal Server Error
    all_events = Event.query.order_by(Event.date.desc(), Event.id.desc()).all()
    if not selected_event_ids:
        selected_event_ids = [e.id for e in all_events if e.date == today][:12]

    selected_events = _beta_selected_manage_events(selected_event_ids)
    event_round_map = _beta_event_rounds_map_from_request(selected_events, selected_round)
    event_field_start_map = _beta_event_field_start_map_from_request(selected_events, field_start)

    if request.method == 'POST':
        action = request.form.get('action') or 'refresh'
        selected_events = _beta_selected_manage_events(selected_event_ids)
        selected_round = request.form.get('round', type=int) or selected_round
        selected_round = max(1, min(int(selected_round or 1), 4))
        field_start = request.form.get('field_start', type=int) or field_start
        field_max = request.form.get('field_max', type=int) or field_max
        field_prefix = (request.form.get('field_prefix') or '').strip()
        field_exclude = (request.form.get('field_exclude') or '').strip()
        separate_same_name = request.form.get('separate_same_name') == 'on'
        force_delete_locked = request.form.get('force_delete_locked') == 'on'
        event_round_map = _beta_event_rounds_map_from_request(selected_events, selected_round)
        event_field_start_map = _beta_event_field_start_map_from_request(selected_events, field_start)

        if not selected_events:
            flash('กรุณาเลือกอีเว้นท์ก่อน', 'warning')
            return redirect(url_for('multi_event_manager'))

        if action == 'refresh':
            pass

        elif action in ('pair_first_round', 'pair_selected_round'):
            ok_count = skip_count = match_count = 0
            messages = []
            changed = []
            for event in selected_events:
                target_round = 1 if action == 'pair_first_round' else event_round_map.get(event.id, selected_round)
                if target_round <= 1:
                    ok, msg, created = _beta_pair_first_round(event, separate_same_name=separate_same_name)
                    target_round = 1
                else:
                    ok, msg, created = _beta_pair_selected_round(event, target_round=target_round, separate_same_name=separate_same_name)
                if ok:
                    ok_count += 1
                    match_count += created
                    changed.append((event.id, target_round))
                else:
                    skip_count += 1
                    messages.append(f"{event.name} ครั้งที่ {target_round}: {msg}")
            db.session.commit()
            for event_id, round_no in changed:
                try:
                    _emit_round_live_changed(event_id, round_no, force_reload=True, reason='multi_event_manager_pair_selected_round')
                except Exception:
                    pass
            if messages:
                flash(f"จับคู่สำเร็จ {ok_count} อีเว้นท์ / ข้าม {skip_count} อีเว้นท์ | " + ' | '.join(messages[:8]), 'warning')
            else:
                flash(f"จับคู่สำเร็จ {ok_count} อีเว้นท์ รวม {match_count} คู่", 'success')

        elif action == 'assign_fields':
            event_match_groups = []
            all_matches = []
            changed = []
            for event in selected_events:
                event.auto_field_enabled = True
                event.field_start = event_field_start_map.get(event.id, field_start)
                event.field_max = field_max
                event.field_prefix = field_prefix
                event.field_exclude = field_exclude
                round_no = event_round_map.get(event.id, selected_round)
                matches = Match.query.filter_by(event_id=event.id, round=round_no).order_by(Match.id.asc()).all()
                event_match_groups.append((event, matches))
                all_matches.extend(matches)
                if matches:
                    changed.append((event.id, round_no))
            if not all_matches:
                flash('ยังไม่มีคู่แข่งขันในอีเว้นท์/ครั้งที่เลือก', 'warning')
            else:
                assigned = _beta_assign_fields_by_event_groups(event_match_groups, field_prefix, field_start, field_max, field_exclude, event_field_start_map)
                unassigned = _beta_assignable_match_count(all_matches) - assigned
                db.session.commit()
                for event_id, round_no in changed:
                    try:
                        _emit_round_live_changed(event_id, round_no, force_reload=True, reason='multi_event_manager_assign_fields')
                    except Exception:
                        pass
                flash(f'กำหนดเลขสนามให้แล้ว {assigned} คู่ / ยังไม่เลือกสนาม {unassigned} คู่' + (' (สนามไม่พอหรือเว้นสนามไว้)' if unassigned else ''), 'success' if not unassigned else 'warning')

        elif action in ('pair_and_assign_first_round', 'pair_and_assign_selected_round'):
            ok_count = skip_count = match_count = 0
            messages = []
            changed = []
            for event in selected_events:
                target_round = 1 if action == 'pair_and_assign_first_round' else event_round_map.get(event.id, selected_round)
                if target_round <= 1:
                    ok, msg, created = _beta_pair_first_round(event, separate_same_name=separate_same_name)
                    target_round = 1
                else:
                    ok, msg, created = _beta_pair_selected_round(event, target_round=target_round, separate_same_name=separate_same_name)
                if ok:
                    ok_count += 1
                    match_count += created
                    changed.append((event.id, target_round))
                else:
                    skip_count += 1
                    messages.append(f"{event.name} ครั้งที่ {target_round}: {msg}")
            db.session.flush()
            event_match_groups = []
            all_matches = []
            for event in selected_events:
                event.auto_field_enabled = True
                event.field_start = event_field_start_map.get(event.id, field_start)
                event.field_max = field_max
                event.field_prefix = field_prefix
                event.field_exclude = field_exclude
                round_no = event_round_map.get(event.id, selected_round)
                matches = Match.query.filter_by(event_id=event.id, round=round_no).order_by(Match.id.asc()).all()
                event_match_groups.append((event, matches))
                all_matches.extend(matches)
            assigned = _beta_assign_fields_by_event_groups(event_match_groups, field_prefix, field_start, field_max, field_exclude, event_field_start_map) if all_matches else 0
            unassigned = _beta_assignable_match_count(all_matches) - assigned if all_matches else 0
            db.session.commit()
            for event_id, round_no in set(changed):
                try:
                    _emit_round_live_changed(event_id, round_no, force_reload=True, reason='multi_event_manager_pair_assign_selected')
                except Exception:
                    pass
            tail = (' | ' + ' | '.join(messages[:8])) if messages else ''
            flash(f'จับคู่+กำหนดสนามเสร็จ: จับใหม่ {ok_count} อีเว้นท์ / ข้าม {skip_count} / ใส่สนาม {assigned} คู่ / ยังไม่เลือกสนาม {unassigned} คู่{tail}', 'success' if ok_count and not unassigned else 'warning')

        elif action == 'clear_fields':
            cleared = 0
            changed = []
            for event in selected_events:
                round_no = event_round_map.get(event.id, selected_round)
                matches = Match.query.filter_by(event_id=event.id, round=round_no).all()
                for m in matches:
                    if m.field not in (None, ''):
                        m.field = None
                        cleared += 1
                if matches:
                    changed.append((event.id, round_no))
            db.session.commit()
            for event_id, round_no in changed:
                try:
                    _emit_round_live_changed(event_id, round_no, force_reload=True, reason='multi_event_manager_clear_fields')
                except Exception:
                    pass
            flash(f'ล้างเลขสนามแล้ว {cleared} คู่', 'success')

        elif action == 'delete_round_matches':
            deleted = 0
            skipped = []
            changed = []
            for event in selected_events:
                round_no = event_round_map.get(event.id, selected_round)
                q = Match.query.filter_by(event_id=event.id, round=round_no)
                locked_count = q.filter(Match.is_locked == True).count()
                total_count = q.count()
                if total_count <= 0:
                    skipped.append(f'{event.name} ครั้งที่ {round_no}: ไม่มีคู่ให้ลบ')
                    continue
                if locked_count and not force_delete_locked:
                    skipped.append(f'{event.name} ครั้งที่ {round_no}: มีคู่ล็อกผลแล้ว {locked_count} คู่')
                    continue
                for m in q.all():
                    db.session.delete(m)
                    deleted += 1
                changed.append((event.id, round_no))
            db.session.commit()
            for event_id, round_no in changed:
                try:
                    _emit_round_live_changed(event_id, round_no, force_reload=True, reason='multi_event_manager_delete_round')
                except Exception:
                    pass
            tail = (' | ' + ' | '.join(skipped[:8])) if skipped else ''
            flash(f'ลบคู่ในครั้งที่เลือกแล้ว {deleted} คู่{tail}', 'warning' if skipped else 'success')

        elif action == 'create_next_competitions':
            ok_count = skip_count = 0
            messages = []
            for event in selected_events:
                ok, msg, created_ref = _beta_multi_create_next_for_event(event)
                if ok:
                    ok_count += 1
                else:
                    skip_count += 1
                messages.append(f'{event.name}: {msg}')
            level = 'success' if ok_count and not skip_count else ('warning' if ok_count else 'danger')
            flash(f'สร้างระบบจัดต่อสำเร็จ {ok_count} อีเว้นท์ / ข้าม {skip_count} อีเว้นท์ | ' + ' | '.join(messages[:10]), level)

        query = {
            'events': ','.join(map(str, selected_event_ids)),
            'round': selected_round,
            'field_start': field_start,
            'field_max': field_max,
            'field_prefix': field_prefix,
            'field_exclude': field_exclude,
            'event_rounds': _beta_event_rounds_param(event_round_map),
            'event_field_starts': _beta_event_field_starts_param(event_field_start_map),
        }
        if separate_same_name:
            query['separate_same_name'] = 'on'
        if force_delete_locked:
            query['force_delete_locked'] = 'on'
        return redirect(url_for('multi_event_manager', **query))

    rows = []
    for event in selected_events:
        round_no = event_round_map.get(event.id, selected_round)
        rounds = [r[0] for r in db.session.query(Match.round).filter_by(event_id=event.id).distinct().order_by(Match.round.asc()).all()]
        match_count = Match.query.filter_by(event_id=event.id, round=round_no).count()
        waiting_count = Match.query.filter_by(event_id=event.id, round=round_no, is_locked=False).count()
        team_count = Team.query.filter_by(event_id=event.id).count()
        rows.append({
            'event': event,
            'selected_round': round_no,
            'team_count': team_count,
            'rounds': rounds,
            'match_count': match_count,
            'waiting_count': waiting_count,
            'field_text': _beta_event_round_field_text(event.id, round_no),
            'field_repeat_count': _beta_match_field_repeat_count(event.id, round_no),
            'field_unassigned_count': _beta_event_round_unassigned_count(event.id, round_no),
            'playoff_count': _beta_playoff_count_for_event(event.id),
            'latest_playoff_id': _beta_first_playoff_id_for_event(event.id),
        })

    selected_events_param = ','.join(map(str, selected_event_ids))
    event_rounds_param = _beta_event_rounds_param(event_round_map)
    event_field_starts_param = _beta_event_field_starts_param(event_field_start_map)
    summary = {
        'events': len(selected_events),
        'teams': sum(r['team_count'] for r in rows),
        'matches': sum(r['match_count'] for r in rows),
        'waiting': sum(r['waiting_count'] for r in rows),
        'unassigned_fields': sum(r['field_unassigned_count'] for r in rows),
        'usable_fields': len(_beta_field_pool(field_prefix, field_start, field_max, field_exclude)),
    }

    return render_template(
        'multi_event_manager.html',
        all_events=all_events,
        selected_events=selected_events,
        selected_event_ids=selected_event_ids,
        selected_events_param=selected_events_param,
        event_rounds_param=event_rounds_param,
        event_field_starts_param=event_field_starts_param,
        event_field_start_map=event_field_start_map,
        event_round_map=event_round_map,
        selected_round=selected_round,
        rows=rows,
        summary=summary,
        today=today,
        field_start=field_start,
        field_max=field_max,
        field_prefix=field_prefix,
        field_exclude=field_exclude,
        separate_same_name=separate_same_name,
        force_delete_locked=force_delete_locked,
    )


@app.route('/multi-manage/print-pairings')
@login_required
@roles_required('superadmin')
def multi_event_print_pairings():
    selected_round = request.args.get('round', type=int) or 1
    selected_event_ids = _selected_event_ids_from_request()
    events = _beta_selected_manage_events(selected_event_ids)
    event_round_map = _beta_event_rounds_map_from_request(events, selected_round)
    rows = []
    for event in events:
        round_no = event_round_map.get(event.id, selected_round)
        try:
            event.logo_list = json.loads(event.logo_filename) if event.logo_filename else []
        except Exception:
            event.logo_list = []
        matches = Match.query.filter_by(event_id=event.id, round=round_no).order_by(Match.id.asc()).all()
        matches.sort(key=lambda m: (_beta_safe_int_field_order_value(m.field) if m.field not in (None, '') else 999999, m.id))
        teams = {t.id: t.name for t in Team.query.filter_by(event_id=event.id).all()}
        rows.append({'event': event, 'matches': matches, 'teams': teams, 'team_count': len(teams), 'selected_round': round_no})
    return render_template('multi_event_print_pairings.html', events=events, rows=rows, selected_round=selected_round, event_round_map=event_round_map)


@app.route('/multi-manage/print-score-sheets')
@login_required
@roles_required('superadmin')
def multi_event_print_score_sheets():
    selected_round = request.args.get('round', type=int) or 1
    selected_event_ids = _selected_event_ids_from_request()
    events = _beta_selected_manage_events(selected_event_ids)
    event_round_map = _beta_event_rounds_map_from_request(events, selected_round)
    rows = []
    for event in events:
        round_no = event_round_map.get(event.id, selected_round)
        try:
            event.logo_list = json.loads(event.logo_filename) if event.logo_filename else []
        except Exception:
            event.logo_list = []
        matches = Match.query.filter_by(event_id=event.id, round=round_no).order_by(Match.id.asc()).all()
        matches.sort(key=lambda m: (_beta_safe_int_field_order_value(m.field) if m.field not in (None, '') else 999999, m.id))
        ensure_match_tokens(matches)
        teams = {t.id: t.name for t in Team.query.filter_by(event_id=event.id).all()}
        rows.append({'event': event, 'matches': matches, 'teams': teams, 'team_count': len(teams), 'selected_round': round_no})
    return render_template('multi_event_print_score_sheets.html', events=events, rows=rows, selected_round=selected_round, event_round_map=event_round_map)


@app.route('/multi-manage/print-standings')
@login_required
@roles_required('superadmin')
def multi_event_print_standings():
    selected_event_ids = _selected_event_ids_from_request()
    events = _beta_selected_manage_events(selected_event_ids)
    direct_advance_count = request.args.get('direct_advance_count', type=int) or 0
    continue_count = request.args.get('continue_count', type=int) or 8
    ab_a_team_count = request.args.get('ab_a_team_count', type=int)
    competition_type = request.args.get('competition_type', '')
    rows = []
    for event in events:
        standings_data = calculate_standings(event.id)
        total_rows = len(standings_data)
        ev_direct = request.args.get(f'event_direct_count_{event.id}', type=int)
        ev_continue = request.args.get(f'event_continue_count_{event.id}', type=int)
        ev_ab = request.args.get(f'event_ab_a_count_{event.id}', type=int)
        ev_comp = (request.args.get(f'event_competition_type_{event.id}') or competition_type or '').strip()
        dac = max(0, min(total_rows, int(ev_direct if ev_direct is not None else (direct_advance_count or 0))))
        cc = max(0, min(total_rows - dac, int(ev_continue if ev_continue is not None else (continue_count or 0))))
        num_qualified = dac + cc
        a_count = ev_ab if ev_ab is not None else ab_a_team_count
        if a_count is None:
            a_count = (cc // 2) if cc > 1 else min(1, cc)
        a_count = max(0, min(cc, int(a_count or 0)))
        rows.append({
            'event': event,
            'standings': standings_data,
            'direct_advance_count': dac,
            'continue_count': cc,
            'num_qualified_teams': num_qualified,
            'ab_a_team_count': a_count,
            'competition_type': ev_comp,
        })
    return render_template(
        'multi_event_print_standings.html',
        rows=rows,
        direct_advance_count=direct_advance_count,
        continue_count=continue_count,
        ab_a_team_count=ab_a_team_count,
        competition_type=competition_type,
    )



def _beta_latest_playoffs_for_events(event_ids):
    if not event_ids:
        return []
    ids = []
    for x in event_ids:
        try:
            ids.append(int(x))
        except Exception:
            pass
    if not ids:
        return []
    placeholders = ','.join(f':eid{i}' for i in range(len(ids)))
    params = {f'eid{i}': eid for i, eid in enumerate(ids)}
    rows = db.session.execute(text(f"""
        SELECT pc.*
        FROM playoff_competitions pc
        JOIN (
            SELECT source_event_id, MAX(id) AS max_id
            FROM playoff_competitions
            WHERE source_event_id IN ({placeholders})
            GROUP BY source_event_id
        ) latest ON latest.max_id = pc.id
        ORDER BY pc.source_event_id, pc.id
    """), params).mappings().all()
    return [dict(r) for r in rows]


def _beta_playoff_pairs_for_round(playoff_id, round_id):
    comp = db.session.execute(text("SELECT * FROM playoff_competitions WHERE id=:id"), {'id': playoff_id}).mappings().first()
    rnd = db.session.execute(text("SELECT * FROM playoff_rounds WHERE id=:rid"), {'rid': round_id}).mappings().first()
    slots = db.session.execute(text("SELECT * FROM playoff_slots WHERE round_id=:rid ORDER BY group_no, slot_no"), {'rid': round_id}).mappings().all()
    score_rows = db.session.execute(text("SELECT * FROM playoff_scores WHERE round_id=:rid"), {'rid': round_id}).mappings().all()
    score_map = {(int(s['group_no']), int(s['slot_no']), int(s['stage_no'])): s['score'] for s in score_rows}
    by_group = {}
    for slot in slots:
        by_group.setdefault(int(slot['group_no']), []).append(dict(slot))
    pairs = []
    for group_no, gslots in by_group.items():
        by_slot = {int(s['slot_no']): s for s in gslots}
        # แสดงคู่ที่ต้องคีย์แบบง่าย: คู่ 1-2 และถ้ามี 3-4 ก็แสดงด้วย
        for a_no, b_no, stage_no, label in ((1, 2, 1, 'คู่ที่ 1'), (3, 4, 1, 'คู่ที่ 2')):
            a = by_slot.get(a_no); b = by_slot.get(b_no)
            if not a or not b:
                continue
            if a.get('is_bye') and b.get('is_bye'):
                continue
            if a.get('is_bye') or b.get('is_bye'):
                continue
            pairs.append({
                'playoff': comp, 'round': rnd, 'group_no': group_no, 'stage_no': stage_no,
                'pair_label': label, 'a': a, 'b': b,
                'score_a': score_map.get((group_no, a_no, stage_no)),
                'score_b': score_map.get((group_no, b_no, stage_no)),
                'court': a.get('court_name') or b.get('court_name') or '',
            })
    return pairs


@app.route('/multi-playoff-score', methods=['GET', 'POST'])
@login_required
@roles_required('superadmin')
def multi_playoff_score_center():
    selected_event_ids = _selected_event_ids_from_request()
    all_events = Event.query.order_by(Event.date.desc(), Event.id.desc()).all()
    if not selected_event_ids:
        today = date.today()
        selected_event_ids = [e.id for e in all_events if e.date == today][:12]
    selected_events = _beta_selected_manage_events(selected_event_ids)
    playoffs = _beta_latest_playoffs_for_events(selected_event_ids)

    if request.method == 'POST':
        saved = 0
        changed_playoffs = set()
        for key, raw in request.form.items():
            if not key.startswith('score_'):
                continue
            # score_<playoff_id>_<round_id>_<group_no>_<slot_no>_<stage_no>
            parts = key.split('_')
            if len(parts) != 6:
                continue
            _, pid_s, rid_s, group_s, slot_s, stage_s = parts
            pid = _as_int(pid_s); rid = _as_int(rid_s); group_no = _as_int(group_s); slot_no = _as_int(slot_s); stage_no = _as_int(stage_s)
            db.session.execute(text("DELETE FROM playoff_scores WHERE round_id=:rid AND group_no=:g AND slot_no=:s AND stage_no=:st"), {'rid': rid, 'g': group_no, 's': slot_no, 'st': stage_no})
            raw = (raw or '').strip()
            if raw != '':
                try:
                    val = max(0, min(13, int(raw)))
                except Exception:
                    continue
                db.session.execute(text("""
                    INSERT INTO playoff_scores (round_id, group_no, slot_no, stage_no, score)
                    VALUES (:rid, :g, :s, :st, :score)
                """), {'rid': rid, 'g': group_no, 's': slot_no, 'st': stage_no, 'score': val})
                saved += 1
            changed_playoffs.add(pid)
        db.session.commit()
        for pid in changed_playoffs:
            try:
                socketio.emit('playoff_reload', {'playoff_id': pid}, to=f'playoff_{pid}')
            except Exception:
                pass
        flash(f'บันทึกคะแนนเพลย์ออฟรวมแล้ว {saved} ช่อง', 'success')
        return redirect(url_for('multi_playoff_score_center', events=','.join(map(str, selected_event_ids))))

    rows = []
    for comp in playoffs:
        source_event = Event.query.get(comp.get('source_event_id'))
        latest_round = db.session.execute(text("SELECT * FROM playoff_rounds WHERE playoff_id=:pid ORDER BY round_no DESC, id DESC LIMIT 1"), {'pid': comp['id']}).mappings().first()
        if not latest_round:
            continue
        pairs = _beta_playoff_pairs_for_round(comp['id'], latest_round['id'])
        for pair in pairs:
            pair['source_event'] = source_event
            rows.append(pair)
    return render_template('multi_playoff_score_center.html', all_events=all_events, selected_event_ids=selected_event_ids, selected_events=selected_events, rows=rows)


# ---------- end beta multi-event manager ----------


if __name__ == '__main__':
    if not os.path.exists(app.config['UPLOAD_FOLDER']):
        os.makedirs(app.config['UPLOAD_FOLDER'])

    setup()
    socketio.run(app, host="0.0.0.0", port=5001, debug=True)